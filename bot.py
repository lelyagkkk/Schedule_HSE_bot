import logging
import os
import re
import argparse
import json
import html
import asyncio
import atexit
import time as time_module
from http.cookies import SimpleCookie
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from io import BytesIO
from threading import Lock
from urllib.parse import parse_qsl, unquote, urlencode, urlparse, urlunparse
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
from telegram import (
    BotCommand,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    MenuButtonCommands,
    ReplyKeyboardMarkup,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)
import yadisk

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

excel_lock = Lock()
labshake_auth_lock = Lock()
labshake_cookie_cache: dict[str, str] = {}
single_instance_lock_fd: int | None = None
single_instance_lock_path: str | None = None

ROLE_PARTICIPANT_BUTTON_TEXT = "Я участник"
ROLE_RESEARCHER_BUTTON_TEXT = "Я исследователь"
EXPERIMENT_SELECTOR_BUTTON_TEXT = "Выбрать эксперимент"
SLOT_WINDOW_DAYS = 14
MAX_WINDOW_STEPS = 52
SLOT_HORIZON_DAYS = 31
BOOKED_FILL = PatternFill(fill_type="solid", start_color="FFE2F0D9", end_color="FFE2F0D9")
LABSHAKE_BUSY_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD9D9D9",
    end_color="FFD9D9D9",
)
EMPTY_FILL = PatternFill(fill_type=None)
LABSHAKE_BUSY_MARKER = "system:labshake_busy"
RESEARCHER_BLOCK_MARKER = "system:researcher_block"
SYNC_REMOVED_SLOTS_LIMIT = 30
BOT_COMMANDS = [
    BotCommand("experiments", "Выбрать эксперимент"),
    BotCommand("researcher", "Режим исследователя"),
    BotCommand("menu", "Открыть меню"),
    BotCommand("book", "Записаться"),
    BotCommand("move", "Перенести запись"),
    BotCommand("cancel", "Отменить ввод"),
    BotCommand("start", "Старт"),
]
DEFAULT_TERMS_TEXT = (
    "Перед записью прочитайте условия эксперимента.\n"
    "Если вы подходите требованиям, нажмите кнопку подтверждения."
)
VALID_BORDER_STYLES = {
    "thin",
    "hair",
    "mediumDashDot",
    "medium",
    "dotted",
    "dashDot",
    "mediumDashed",
    "mediumDashDotDot",
    "thick",
    "slantDashDot",
    "dashDotDot",
    "dashed",
    "double",
}
WEEKDAY_NAMES_RU = (
    "Понедельник",
    "Вторник",
    "Среда",
    "Четверг",
    "Пятница",
    "Суббота",
    "Воскресенье",
)
WEEKDAY_RAINBOW_MARKERS = {
    0: "🟥",  # Monday
    1: "🟧",  # Tuesday
    2: "🟨",  # Wednesday
    3: "🟩",  # Thursday
    4: "🟦",  # Friday
    5: "🟪",  # Saturday
    6: "🟫",  # Sunday
}
WEEKDAY_TOKEN_MAP = {
    "понедельник": 0,
    "пн": 0,
    "monday": 0,
    "mon": 0,
    "вторник": 1,
    "вт": 1,
    "tuesday": 1,
    "tue": 1,
    "среда": 2,
    "ср": 2,
    "wednesday": 2,
    "wed": 2,
    "четверг": 3,
    "чт": 3,
    "thursday": 3,
    "thu": 3,
    "пятница": 4,
    "пт": 4,
    "friday": 4,
    "fri": 4,
    "суббота": 5,
    "сб": 5,
    "saturday": 5,
    "sat": 5,
    "воскресенье": 6,
    "вс": 6,
    "sunday": 6,
    "sun": 6,
}
LABSHAKE_DATE_RE = re.compile(r"\b(20\d{2}-\d{2}-\d{2})\b")
LABSHAKE_DATE_TOKEN_RE = re.compile(
    r"\b(20\d{2}[./-]\d{1,2}[./-]\d{1,2}|\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?)\b"
)
ADMIN_FIELD_LABELS = {
    "participant_visible": "Видимость для участников",
    "working_hours": "Рабочее окно (часы)",
    "excluded_days": "Исключенные дни",
    "max_weekly_hours": "Лимит часов в неделю",
    "slot_mode": "Режим слотов",
    "available_days_ahead": "Количество доступных дней",
    "slot_duration_hours": "Длительность слота (ч)",
    "slot_step_minutes": "Перерыв между слотами (мин)",
    "labshake_booking_comment": "Комментарий к записям LabShake",
    "scientist_id": "Telegram исследователя",
}
ADMIN_FIELD_ORDER = [
    "participant_visible",
    "working_hours",
    "excluded_days",
    "max_weekly_hours",
    "slot_mode",
    "available_days_ahead",
    "slot_duration_hours",
    "slot_step_minutes",
    "labshake_booking_comment",
    "scientist_id",
]


@dataclass
class ExcelStorage:
    mode: str
    excel_path: str | None = None
    yadisk_path: str | None = None
    yadisk_client: yadisk.YaDisk | None = None


@dataclass(frozen=True)
class SheetColumns:
    date: int
    time: int
    full_name: int
    phone: int
    telegram: int
    booked_at: int | None = None


@dataclass
class ExperimentConfig:
    experiment_id: str
    title: str
    terms_text: str
    scientist_id: str
    storage: ExcelStorage
    participant_visible: bool = False
    max_weekly_hours: float | None = None
    default_slot_duration_hours: float | None = None
    slot_mode: str = "manual"
    working_hours: str | None = None
    excluded_weekdays: set[int] = field(default_factory=set)
    slot_duration_hours: float | None = None
    min_gap_hours: float = 0.0
    slot_step_minutes: int = 60
    available_days_ahead: int = SLOT_HORIZON_DAYS
    labshake_schedule_url: str | None = None
    labshake_booking_comment: str | None = None
    labshake_cookie_env: str = "LABSHAKE_COOKIE"
    labshake_days_ahead: int = SLOT_HORIZON_DAYS


@dataclass
class ResearcherAccess:
    usernames: set[str] = field(default_factory=set)
    user_ids: set[int] = field(default_factory=set)


def create_empty_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Slots"
    ws.append(["Date", "Time", "Telegram", "FullName", "Phone", "BookedAt"])
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def sanitize_styles_xml(styles_xml: bytes) -> tuple[bytes, bool]:
    try:
        root = ET.fromstring(styles_xml)
    except ET.ParseError:
        return styles_xml, False

    changed = False
    border_side_tags = {
        "left",
        "right",
        "top",
        "bottom",
        "diagonal",
        "vertical",
        "horizontal",
        "start",
        "end",
    }

    for element in root.iter():
        tag = element.tag.split("}", 1)[-1]
        if tag not in border_side_tags:
            continue

        style = element.attrib.get("style")
        if style is not None and style not in VALID_BORDER_STYLES:
            element.attrib.pop("style", None)
            changed = True

    if not changed:
        return styles_xml, False

    new_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return new_xml, True


def sanitize_workbook_bytes(data: bytes) -> tuple[bytes, bool]:
    try:
        source = ZipFile(BytesIO(data), "r")
    except Exception:
        return data, False

    with source:
        if "xl/styles.xml" not in source.namelist():
            return data, False

        styles_xml = source.read("xl/styles.xml")
        new_styles_xml, changed = sanitize_styles_xml(styles_xml)
        if not changed:
            return data, False

        output = BytesIO()
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as target:
            for item in source.infolist():
                payload = new_styles_xml if item.filename == "xl/styles.xml" else source.read(item.filename)
                target.writestr(item, payload)

    return output.getvalue(), True


def ensure_workbook_exists(path: str) -> None:
    if os.path.exists(path):
        return

    parent_dir = os.path.dirname(path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)

    wb = create_empty_workbook()
    wb.save(path)
    wb.close()
    logger.info("Created new Excel file at %s", path)


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return "".join(ch for ch in text if ch.isalnum())


def detect_sheet_columns(ws) -> SheetColumns:
    max_col = max(ws.max_column, 6)
    headers: dict[str, int] = {}

    for col in range(1, max_col + 1):
        raw = ws.cell(row=1, column=col).value
        if is_empty(raw):
            continue
        key = normalize_header(raw)
        if key and key not in headers:
            headers[key] = col

    def pick(keys: tuple[str, ...]) -> int | None:
        for key in keys:
            if key in headers:
                return headers[key]
        return None

    date_col = pick(("день", "дата", "date", "day"))
    time_col = pick(("время", "time"))
    full_name_col = pick(("фио", "fullname", "name", "фиоучастника"))
    phone_col = pick(("номер", "телефон", "phone", "phonenumber", "номертелефона"))
    tg_col = pick(("тг", "tg", "telegram", "телеграм", "username"))
    booked_at_col = pick(
        (
            "bookedat",
            "bookedtime",
            "времязаписи",
            "датазаписи",
            "датавремязаписи",
            "createdat",
        )
    )

    if date_col and time_col and full_name_col and phone_col and tg_col:
        return SheetColumns(
            date=date_col,
            time=time_col,
            full_name=full_name_col,
            phone=phone_col,
            telegram=tg_col,
            booked_at=booked_at_col,
        )

    # Fallback for legacy/default table layout.
    return SheetColumns(
        date=1,
        time=2,
        full_name=4,
        phone=5,
        telegram=3,
        booked_at=6 if ws.max_column >= 6 else None,
    )


def booking_style_columns(cols: SheetColumns) -> list[int]:
    columns = [cols.date, cols.time, cols.full_name, cols.phone, cols.telegram]
    if cols.booked_at:
        columns.append(cols.booked_at)
    return sorted(set(columns))


def set_booking_row_style(ws, row: int, cols: SheetColumns, booked: bool) -> None:
    tg_value = ws.cell(row=row, column=cols.telegram).value
    if is_empty(tg_value):
        fill = EMPTY_FILL
    elif is_labshake_busy_value(tg_value):
        fill = LABSHAKE_BUSY_FILL
    else:
        fill = BOOKED_FILL if booked else EMPTY_FILL
    for col in booking_style_columns(cols):
        ws.cell(row=row, column=col).fill = fill


def normalize_yadisk_path(raw_value: str) -> str:
    value = raw_value.strip()
    if not value:
        return "disk:/slots.xlsx"

    if value.startswith("http://") or value.startswith("https://"):
        marker = "/edit/disk/"
        if marker in value:
            encoded_path = value.split(marker, 1)[1].split("?", 1)[0]
            decoded_path = unquote(encoded_path)
            if decoded_path.startswith("disk/"):
                return "disk:/" + decoded_path[len("disk/") :]
            if decoded_path.startswith("disk:/"):
                return decoded_path

    if value.startswith("disk:/"):
        return value
    if value.startswith("disk/"):
        return "disk:/" + value[len("disk/") :]
    if value.startswith("/"):
        return f"disk:{value}"
    return f"disk:/{value}"


def build_storage_from_values(
    mode: str,
    excel_path: str | None,
    yadisk_path: str | None,
    yadisk_token: str | None,
) -> ExcelStorage:
    normalized_mode = (mode or "local").strip().lower()

    if normalized_mode == "local":
        return ExcelStorage(mode="local", excel_path=(excel_path or "slots.xlsx"))

    if normalized_mode == "yadisk":
        token = (yadisk_token or "").strip()
        if not token:
            raise RuntimeError("YADISK_TOKEN not found. Add it to .env file.")

        raw_path = yadisk_path or excel_path or "slots.xlsx"
        path = normalize_yadisk_path(raw_path)
        return ExcelStorage(
            mode="yadisk", yadisk_path=path, yadisk_client=yadisk.YaDisk(token=token)
        )

    raise RuntimeError("Unsupported STORAGE_MODE. Use local or yadisk.")


def build_storage_config() -> ExcelStorage:
    return build_storage_from_values(
        mode=os.getenv("STORAGE_MODE", "local"),
        excel_path=os.getenv("EXCEL_PATH", "slots.xlsx"),
        yadisk_path=os.getenv("YADISK_PATH", os.getenv("EXCEL_PATH", "slots.xlsx")),
        yadisk_token=os.getenv("YADISK_TOKEN", ""),
    )


def make_experiment_id(value: str, fallback_index: int) -> str:
    slug = re.sub(r"[^a-z0-9_]+", "_", value.strip().lower())
    slug = slug.strip("_")
    if slug:
        return slug
    return f"experiment_{fallback_index}"


def parse_optional_positive_float(
    raw: object, *, field_name: str, experiment_title: str
) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, str) and raw.strip() == "":
        return None

    try:
        value = float(raw)
    except (TypeError, ValueError):
        raise RuntimeError(
            f"{experiment_title}: '{field_name}' must be a number."
        ) from None

    if value <= 0:
        raise RuntimeError(
            f"{experiment_title}: '{field_name}' must be greater than 0."
        )
    return value


def parse_optional_non_negative_float(
    raw: object, *, field_name: str, experiment_title: str
) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, str) and raw.strip() == "":
        return None

    try:
        value = float(raw)
    except (TypeError, ValueError):
        raise RuntimeError(
            f"{experiment_title}: '{field_name}' must be a number."
        ) from None

    if value < 0:
        raise RuntimeError(
            f"{experiment_title}: '{field_name}' must be >= 0."
        )
    return value


def parse_optional_positive_int(
    raw: object, *, field_name: str, experiment_title: str
) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, str) and raw.strip() == "":
        return None

    try:
        value = int(raw)
    except (TypeError, ValueError):
        raise RuntimeError(
            f"{experiment_title}: '{field_name}' must be an integer."
        ) from None

    if value <= 0:
        raise RuntimeError(
            f"{experiment_title}: '{field_name}' must be greater than 0."
        )
    return value


def parse_optional_bool(
    raw: object, *, field_name: str, experiment_title: str, default: bool = False
) -> bool:
    if raw is None:
        return default
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, int):
        if raw in {0, 1}:
            return bool(raw)
    if isinstance(raw, float):
        if raw in {0.0, 1.0}:
            return bool(int(raw))

    text = str(raw).strip().lower()
    if not text:
        return default
    if text in {"1", "true", "yes", "y", "on", "да"}:
        return True
    if text in {"0", "false", "no", "n", "off", "нет"}:
        return False
    raise RuntimeError(
        f"{experiment_title}: '{field_name}' must be boolean (true/false)."
    )


def normalize_weekday_token(token: str) -> str:
    text = token.strip().lower().replace("ё", "е")
    return "".join(ch for ch in text if ch.isalnum())


def parse_excluded_weekdays_value(raw_value: object) -> set[int]:
    if raw_value is None:
        return set()

    def parse_token(token: object) -> int:
        if isinstance(token, bool):
            raise ValueError("Булевы значения не поддерживаются.")
        if isinstance(token, int):
            value = int(token)
            if 0 <= value <= 6:
                return value
            raise ValueError("Номер дня должен быть от 0 до 6.")
        if isinstance(token, float):
            if token.is_integer():
                value = int(token)
                if 0 <= value <= 6:
                    return value
            raise ValueError("Номер дня должен быть от 0 до 6.")

        text = str(token).strip()
        if not text:
            raise ValueError("Пустое значение дня.")
        if text.isdigit():
            value = int(text)
            if 0 <= value <= 6:
                return value
            raise ValueError("Номер дня должен быть от 0 до 6.")

        normalized = normalize_weekday_token(text)
        if normalized in {"none", "null", "нет", "пусто"}:
            raise ValueError("none_marker")
        day = WEEKDAY_TOKEN_MAP.get(normalized)
        if day is None:
            raise ValueError(f"Неизвестный день: '{text}'.")
        return day

    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return set()
        if normalize_weekday_token(text) in {"none", "null", "нет", "пусто"}:
            return set()
        separators_normalized = re.sub(r"[;\n|/]+", ",", text)
        parts = [part.strip() for part in separators_normalized.split(",") if part.strip()]
        if len(parts) == 1 and " " in parts[0]:
            parts = [part.strip() for part in parts[0].split() if part.strip()]
        tokens: list[object] = parts
    elif isinstance(raw_value, (list, tuple, set)):
        tokens = list(raw_value)
    else:
        tokens = [raw_value]

    result: set[int] = set()
    for token in tokens:
        try:
            day_idx = parse_token(token)
        except ValueError as exc:
            if str(exc) == "none_marker":
                continue
            raise
        result.add(day_idx)
    return result


def format_weekday_set(weekdays: set[int]) -> str:
    if not weekdays:
        return "не заданы"
    ordered = sorted(day for day in weekdays if 0 <= day <= 6)
    return ", ".join(WEEKDAY_NAMES_RU[day] for day in ordered)


def load_experiments_config(default_storage: ExcelStorage) -> list[ExperimentConfig]:
    experiments_file = os.getenv("EXPERIMENTS_FILE", "experiments.json").strip()
    default_terms = os.getenv("EXPERIMENT_TERMS", DEFAULT_TERMS_TEXT).replace("\\n", "\n")
    default_title = os.getenv(
        "DEFAULT_EXPERIMENT_TITLE", "Исследование процессов принятия решений (fNIRS + tDCS)"
    )
    default_scientist_id = os.getenv("DEFAULT_SCIENTIST_ID", "scientist_id").strip() or "scientist_id"

    if not os.path.exists(experiments_file):
        return [
            ExperimentConfig(
                experiment_id="default",
                title=default_title,
                terms_text=default_terms,
                scientist_id=default_scientist_id,
                storage=default_storage,
                participant_visible=False,
                max_weekly_hours=None,
                default_slot_duration_hours=None,
                slot_mode="manual",
                working_hours=None,
                excluded_weekdays=set(),
                slot_duration_hours=None,
                min_gap_hours=0.0,
                slot_step_minutes=60,
                available_days_ahead=SLOT_HORIZON_DAYS,
                labshake_booking_comment=None,
                labshake_days_ahead=SLOT_HORIZON_DAYS,
            )
        ]

    try:
        with open(experiments_file, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception as exc:
        raise RuntimeError(f"Could not read {experiments_file}: {exc}") from exc

    if isinstance(raw, dict):
        items = raw.get("experiments")
    elif isinstance(raw, list):
        items = raw
    else:
        items = None

    if not isinstance(items, list) or not items:
        raise RuntimeError(f"{experiments_file} must contain non-empty experiments list.")

    result: list[ExperimentConfig] = []
    used_ids: set[str] = set()

    for idx, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise RuntimeError(f"{experiments_file}: experiment #{idx} must be an object.")

        title = str(item.get("title", "")).strip()
        if not title:
            raise RuntimeError(f"{experiments_file}: experiment #{idx} has empty title.")

        raw_id = str(item.get("id", "")).strip() or title
        experiment_id = make_experiment_id(raw_id, idx)
        if experiment_id in used_ids:
            raise RuntimeError(f"{experiments_file}: duplicate experiment id '{experiment_id}'.")
        used_ids.add(experiment_id)

        # For multi-experiment mode each experiment should define its own terms.
        terms_source = (
            item.get("default_terms_text")
            or item.get("terms_text")
            or item.get("terms")
        )
        if terms_source is None or str(terms_source).strip() == "":
            raise RuntimeError(
                f"{experiments_file}: experiment '{title}' must define "
                f"'default_terms_text' (or 'terms_text'/'terms')."
            )
        terms_text = str(terms_source).replace("\\n", "\n")
        scientist_id = str(item.get("scientist_id", "")).strip()
        if not scientist_id:
            raise RuntimeError(
                f"{experiments_file}: experiment '{title}' must define 'scientist_id'."
            )
        extra_params = item.get("extra_params", {})
        if extra_params is None:
            extra_params = {}
        if not isinstance(extra_params, dict):
            raise RuntimeError(
                f"{experiments_file}: experiment '{title}' has invalid 'extra_params' (must be object)."
            )
        participant_visible = parse_optional_bool(
            item.get(
                "participant_visible",
                extra_params.get("participant_visible"),
            ),
            field_name="participant_visible",
            experiment_title=title,
            default=False,
        )

        max_weekly_hours = parse_optional_positive_float(
            item.get("max_weekly_hours", extra_params.get("max_weekly_hours")),
            field_name="max_weekly_hours",
            experiment_title=title,
        )
        default_slot_duration_hours = parse_optional_positive_float(
            item.get(
                "default_slot_duration_hours",
                extra_params.get("default_slot_duration_hours"),
            ),
            field_name="default_slot_duration_hours",
            experiment_title=title,
        )
        slot_mode = str(
            item.get("slot_mode", extra_params.get("slot_mode", "manual"))
        ).strip().lower()
        if slot_mode not in {"manual", "day_windows"}:
            raise RuntimeError(
                f"{experiments_file}: experiment '{title}' has invalid 'slot_mode'. "
                "Use 'manual' or 'day_windows'."
            )

        working_hours = item.get("working_hours", extra_params.get("working_hours"))
        if isinstance(working_hours, str):
            working_hours = working_hours.strip()
            if not working_hours:
                working_hours = None
        elif working_hours is not None:
            working_hours = str(working_hours).strip() or None

        if working_hours is not None and parse_time_range(working_hours) is None:
            raise RuntimeError(
                f"{experiments_file}: experiment '{title}' has invalid 'working_hours'. "
                "Use format like '10:00-17:00'."
            )

        excluded_days_raw = item.get("excluded_days", extra_params.get("excluded_days"))
        try:
            excluded_weekdays = parse_excluded_weekdays_value(excluded_days_raw)
        except ValueError as exc:
            raise RuntimeError(
                f"{experiments_file}: experiment '{title}' has invalid 'excluded_days': {exc}"
            ) from None

        slot_duration_hours = parse_optional_positive_float(
            item.get("slot_duration_hours", extra_params.get("slot_duration_hours")),
            field_name="slot_duration_hours",
            experiment_title=title,
        )
        min_gap_hours_legacy = parse_optional_non_negative_float(
            item.get("min_gap_hours", extra_params.get("min_gap_hours")),
            field_name="min_gap_hours",
            experiment_title=title,
        )
        slot_step_minutes = parse_optional_positive_int(
            item.get("slot_step_minutes", extra_params.get("slot_step_minutes")),
            field_name="slot_step_minutes",
            experiment_title=title,
        )
        if slot_step_minutes is None:
            if min_gap_hours_legacy is not None:
                slot_step_minutes = max(1, int(round(min_gap_hours_legacy * 60)))
            else:
                slot_step_minutes = 60

        # Keep one source of truth for pause settings.
        min_gap_hours = slot_step_minutes / 60.0

        labshake_schedule_url = item.get(
            "labshake_schedule_url", extra_params.get("labshake_schedule_url")
        )
        if isinstance(labshake_schedule_url, str):
            labshake_schedule_url = labshake_schedule_url.strip() or None
        elif labshake_schedule_url is not None:
            labshake_schedule_url = str(labshake_schedule_url).strip() or None

        labshake_booking_comment = item.get(
            "labshake_booking_comment", extra_params.get("labshake_booking_comment")
        )
        if labshake_booking_comment is None:
            normalized_comment = None
        else:
            comment_text = str(labshake_booking_comment).strip()
            normalized_comment = comment_text or None

        labshake_cookie_env = str(
            item.get("labshake_cookie_env", extra_params.get("labshake_cookie_env", "LABSHAKE_COOKIE"))
        ).strip()
        if not labshake_cookie_env:
            labshake_cookie_env = "LABSHAKE_COOKIE"

        available_days_ahead = parse_optional_positive_int(
            item.get(
                "available_days_ahead",
                extra_params.get(
                    "available_days_ahead",
                    item.get("labshake_days_ahead", extra_params.get("labshake_days_ahead")),
                ),
            ),
            field_name="available_days_ahead",
            experiment_title=title,
        )
        if available_days_ahead is None:
            available_days_ahead = SLOT_HORIZON_DAYS
        # Synchronize LabShake and participant horizon using one source of truth.
        labshake_days_ahead = available_days_ahead

        if slot_mode == "day_windows" and slot_duration_hours is None:
            raise RuntimeError(
                f"{experiments_file}: experiment '{title}' with slot_mode='day_windows' "
                "must define 'slot_duration_hours'."
            )

        storage_mode = str(item.get("storage_mode", default_storage.mode)).strip().lower()

        excel_path = item.get("excel_path")
        if excel_path is None:
            excel_path = default_storage.excel_path
        excel_path = str(excel_path) if excel_path is not None else None

        yadisk_path = item.get("yadisk_path")
        if yadisk_path is None:
            yadisk_path = default_storage.yadisk_path
        yadisk_path = str(yadisk_path) if yadisk_path is not None else None

        token_env = str(item.get("yadisk_token_env", "YADISK_TOKEN")).strip() or "YADISK_TOKEN"
        yadisk_token = os.getenv(token_env, "")

        if storage_mode == "local" and not excel_path:
            excel_path = f"{experiment_id}.xlsx"

        storage = build_storage_from_values(
            mode=storage_mode,
            excel_path=excel_path,
            yadisk_path=yadisk_path,
            yadisk_token=yadisk_token,
        )

        result.append(
            ExperimentConfig(
                experiment_id=experiment_id,
                title=title,
                terms_text=terms_text,
                scientist_id=scientist_id,
                storage=storage,
                participant_visible=participant_visible,
                max_weekly_hours=max_weekly_hours,
                default_slot_duration_hours=default_slot_duration_hours,
                slot_mode=slot_mode,
                working_hours=working_hours,
                excluded_weekdays=excluded_weekdays,
                slot_duration_hours=slot_duration_hours,
                min_gap_hours=min_gap_hours,
                slot_step_minutes=slot_step_minutes,
                available_days_ahead=available_days_ahead,
                labshake_schedule_url=labshake_schedule_url,
                labshake_booking_comment=normalized_comment,
                labshake_cookie_env=labshake_cookie_env,
                labshake_days_ahead=labshake_days_ahead,
            )
        )

    return result


def normalize_username(value: object) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.startswith("@"):
        text = text[1:]
    text = text.strip().lower()
    return text or None


def load_researchers_access(experiments: list[ExperimentConfig]) -> ResearcherAccess:
    access = ResearcherAccess()
    researchers_file = os.getenv("RESEARCHERS_FILE", "researchers.json").strip()
    default_items = sorted({exp.scientist_id for exp in experiments if exp.scientist_id})

    if not os.path.exists(researchers_file):
        try:
            with open(researchers_file, "w", encoding="utf-8") as f:
                json.dump({"researchers": default_items}, f, ensure_ascii=False, indent=2)
            logger.info("Created %s", researchers_file)
        except Exception:
            logger.exception("Could not create %s", researchers_file)

    if not os.path.exists(researchers_file):
        return access

    try:
        with open(researchers_file, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        logger.exception("Could not read %s", researchers_file)
        return access

    if isinstance(raw, dict):
        items = raw.get("researchers")
        if items is None:
            items = []
            for value in raw.get("usernames", []):
                items.append({"username": value})
            for value in raw.get("user_ids", []):
                items.append({"user_id": value})
    elif isinstance(raw, list):
        items = raw
    else:
        logger.warning("%s has unsupported format", researchers_file)
        return access

    if not isinstance(items, list):
        logger.warning("%s: 'researchers' must be a list", researchers_file)
        return access

    for item in items:
        if isinstance(item, str):
            normalized = normalize_username(item)
            if normalized:
                access.usernames.add(normalized)
            continue

        if isinstance(item, int):
            access.user_ids.add(item)
            continue

        if not isinstance(item, dict):
            continue

        username = (
            item.get("username")
            or item.get("tg")
            or item.get("telegram")
            or item.get("scientist_id")
        )
        normalized = normalize_username(username)
        if normalized:
            access.usernames.add(normalized)

        user_id = item.get("user_id", item.get("id"))
        if user_id is not None:
            try:
                access.user_ids.add(int(user_id))
            except (TypeError, ValueError):
                logger.warning("Invalid researcher user_id in %s: %r", researchers_file, user_id)

    return access


def download_yadisk_file(client: yadisk.YaDisk, path: str) -> tuple[bytes | None, str | None]:
    buffer = BytesIO()
    try:
        client.download(path, buffer)
        return buffer.getvalue(), None
    except yadisk.exceptions.PathNotFoundError:
        return None, "not_found"
    except yadisk.exceptions.UnauthorizedError:
        return None, "YADISK_TOKEN недействителен или просрочен."
    except yadisk.exceptions.ForbiddenError:
        return None, "Недостаточно прав у YADISK_TOKEN для чтения файла."
    except yadisk.exceptions.YaDiskError:
        logger.exception("Yandex.Disk download failed for %s", path)
        return None, "Не удалось скачать Excel файл с Яндекс.Диска."


def upload_yadisk_file(client: yadisk.YaDisk, path: str, data: bytes) -> tuple[bool, str]:
    try:
        max_attempts = max(1, int(os.getenv("YADISK_UPLOAD_RETRY_ATTEMPTS", "20")))
    except ValueError:
        max_attempts = 20

    try:
        retry_delay_sec = max(0.2, float(os.getenv("YADISK_UPLOAD_RETRY_DELAY_SEC", "2.0")))
    except ValueError:
        retry_delay_sec = 2.0

    for attempt in range(1, max_attempts + 1):
        try:
            client.upload(BytesIO(data), path, overwrite=True)
            return True, "ok"
        except yadisk.exceptions.UnauthorizedError:
            return False, "YADISK_TOKEN invalid or expired."
        except yadisk.exceptions.ForbiddenError:
            return False, "Insufficient permissions for YADISK_TOKEN to write file."
        except yadisk.exceptions.ParentNotFoundError:
            return False, "Parent folder for YADISK_PATH was not found on Yandex.Disk."
        except (yadisk.exceptions.ResourceIsLockedError, yadisk.exceptions.LockedError):
            if attempt < max_attempts:
                logger.warning(
                    "Yandex.Disk resource is locked for %s. Retry %s/%s in %.1fs.",
                    path,
                    attempt,
                    max_attempts,
                    retry_delay_sec,
                )
                time_module.sleep(retry_delay_sec)
                continue
            return (
                False,
                "Yandex.Disk file is temporarily locked (for example, open in editor).",
            )
        except (
            yadisk.exceptions.RetriableYaDiskError,
            yadisk.exceptions.YaDiskConnectionError,
            yadisk.exceptions.RequestTimeoutError,
            yadisk.exceptions.TooManyRequestsError,
            yadisk.exceptions.UnavailableError,
            yadisk.exceptions.GatewayTimeoutError,
            yadisk.exceptions.BadGatewayError,
        ) as exc:
            if attempt < max_attempts:
                logger.warning(
                    "Transient Yandex.Disk upload error for %s (%s). Retry %s/%s in %.1fs.",
                    path,
                    type(exc).__name__,
                    attempt,
                    max_attempts,
                    retry_delay_sec,
                )
                time_module.sleep(retry_delay_sec)
                continue
            logger.exception(
                "Yandex.Disk upload transient error after retries for %s", path
            )
            return False, "Temporary Yandex.Disk error while uploading workbook."
        except yadisk.exceptions.YaDiskError as exc:
            logger.exception(
                "Yandex.Disk upload failed for %s (%s)", path, type(exc).__name__
            )
            return (
                False,
                f"Failed to upload workbook to Yandex.Disk ({type(exc).__name__}).",
            )

    return False, "Failed to upload workbook to Yandex.Disk."


def ensure_yadisk_parent_dirs(client: yadisk.YaDisk, path: str) -> tuple[bool, str | None]:
    normalized = path.strip()
    if not normalized.startswith("disk:/"):
        return False, "YADISK_PATH must start with 'disk:/'."

    suffix = normalized[6:]
    if "/" not in suffix:
        return True, None

    parent = normalized.rsplit("/", 1)[0]
    if parent in {"disk:", "disk:/"}:
        return True, None

    parts = [part for part in parent[6:].split("/") if part]
    current = "disk:"

    for part in parts:
        current = f"{current}/{part}"
        try:
            if client.exists(current):
                continue
            client.mkdir(current)
        except (yadisk.exceptions.DirectoryExistsError, yadisk.exceptions.PathExistsError):
            continue
        except yadisk.exceptions.UnauthorizedError:
            return False, "YADISK_TOKEN РЅРµРґРµР№СЃС‚РІРёС‚РµР»РµРЅ РёР»Рё РїСЂРѕСЃСЂРѕС‡РµРЅ."
        except yadisk.exceptions.ForbiddenError:
            return False, "РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ Сѓ YADISK_TOKEN РґР»СЏ СЃРѕР·РґР°РЅРёСЏ РїР°РїРѕРє."
        except yadisk.exceptions.YaDiskError:
            logger.exception("Yandex.Disk mkdir failed for %s", current)
            return False, "РќРµ СѓРґР°Р»РѕСЃСЊ СЃРѕР·РґР°С‚СЊ РїР°РїРєСѓ РЅР° РЇРЅРґРµРєСЃ.Р”РёСЃРєРµ."

    return True, None


def ensure_storage_workbook_exists(storage: ExcelStorage) -> tuple[bool, str | None]:
    if storage.mode == "local":
        assert storage.excel_path is not None
        ensure_workbook_exists(storage.excel_path)
        return True, None

    assert storage.yadisk_client is not None
    assert storage.yadisk_path is not None

    try:
        if storage.yadisk_client.exists(storage.yadisk_path):
            return True, None
    except yadisk.exceptions.UnauthorizedError:
        return False, "YADISK_TOKEN недействителен или просрочен."
    except yadisk.exceptions.ForbiddenError:
        return False, "Недостаточно прав у YADISK_TOKEN для доступа к файлу."
    except yadisk.exceptions.YaDiskError:
        logger.exception("Yandex.Disk exists check failed for %s", storage.yadisk_path)
        return False, "Не удалось проверить наличие файла на Яндекс.Диске."

    wb = create_empty_workbook()
    try:
        data = workbook_to_bytes(wb)
    finally:
        wb.close()

    dirs_ok, dirs_error = ensure_yadisk_parent_dirs(storage.yadisk_client, storage.yadisk_path)
    if not dirs_ok:
        return False, dirs_error

    success, message = upload_yadisk_file(storage.yadisk_client, storage.yadisk_path, data)
    if not success:
        return False, message

    logger.info("Created new Excel file on Yandex.Disk at %s", storage.yadisk_path)
    return True, None


def check_storage_access(storage: ExcelStorage) -> tuple[bool, str]:
    if storage.mode == "local":
        assert storage.excel_path is not None
        success, error = ensure_storage_workbook_exists(storage)
        if not success:
            return False, error or "Не удалось инициализировать локальный Excel."

        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return False, load_error
        assert wb is not None
        try:
            ws = wb.active
            return True, f"OK: локальный файл доступен ({storage.excel_path}), строк: {ws.max_row - 1}"
        finally:
            wb.close()

    assert storage.yadisk_client is not None
    assert storage.yadisk_path is not None

    try:
        if not storage.yadisk_client.check_token():
            return False, "YADISK_TOKEN не прошел проверку (invalid token)."
    except yadisk.exceptions.YaDiskError:
        logger.exception("Token check failed")
        return False, "Не удалось проверить YADISK_TOKEN через API Яндекс.Диска."

    success, error = ensure_storage_workbook_exists(storage)
    if not success:
        return False, error or "Не удалось инициализировать файл на Яндекс.Диске."

    wb, load_error = load_workbook_from_storage(storage)
    if load_error:
        return False, load_error
    assert wb is not None
    try:
        ws = wb.active
        return True, f"OK: Яндекс.Диск доступен ({storage.yadisk_path}), строк: {ws.max_row - 1}"
    finally:
        wb.close()


def load_workbook_from_storage(storage: ExcelStorage) -> tuple[Workbook | None, str | None]:
    if storage.mode == "local":
        assert storage.excel_path is not None
        try:
            return load_workbook(storage.excel_path), None
        except PermissionError:
            logger.exception("Excel file is locked: %s", storage.excel_path)
            return None, "Файл Excel сейчас занят. Закройте его и попробуйте снова."
        except Exception:
            logger.exception("Could not read Excel file: %s", storage.excel_path)
            return None, "Не удалось прочитать Excel файл."

    assert storage.yadisk_client is not None
    assert storage.yadisk_path is not None
    data, error = download_yadisk_file(storage.yadisk_client, storage.yadisk_path)
    if error == "not_found":
        return None, "Excel файл не найден на Яндекс.Диске."
    if error:
        return None, error

    assert data is not None
    try:
        return load_workbook(BytesIO(data)), None
    except Exception:
        sanitized_data, changed = sanitize_workbook_bytes(data)
        if changed:
            try:
                wb = load_workbook(BytesIO(sanitized_data))
                logger.warning(
                    "Workbook styles were sanitized for %s due to invalid border styles.",
                    storage.yadisk_path,
                )
                upload_yadisk_file(
                    storage.yadisk_client, storage.yadisk_path, sanitized_data
                )
                return wb, None
            except Exception:
                logger.exception(
                    "Sanitized workbook could not be parsed: %s", storage.yadisk_path
                )

        logger.exception("Could not parse Excel file from Yandex.Disk: %s", storage.yadisk_path)
        return None, "Excel файл на Яндекс.Диске поврежден или в неверном формате."


def save_workbook_to_storage(storage: ExcelStorage, wb: Workbook) -> tuple[bool, str]:
    if storage.mode == "local":
        assert storage.excel_path is not None
        try:
            wb.save(storage.excel_path)
            return True, "ok"
        except PermissionError:
            logger.exception("Excel file is locked during save: %s", storage.excel_path)
            return False, "Не удалось сохранить запись: Excel файл сейчас занят."
        except Exception:
            logger.exception("Could not save Excel file: %s", storage.excel_path)
            return False, "Не удалось сохранить Excel файл."

    assert storage.yadisk_client is not None
    assert storage.yadisk_path is not None
    data = workbook_to_bytes(wb)
    return upload_yadisk_file(storage.yadisk_client, storage.yadisk_path, data)


def is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def parse_date_cell(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        full_candidates = [text]
        extracted_candidates: list[str] = []
        for pattern in (
            r"\b(\d{4}[./-]\d{1,2}[./-]\d{1,2})\b",
            r"\b(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\b",
        ):
            match = re.search(pattern, text)
            if not match:
                continue
            candidate = match.group(1)
            if candidate not in extracted_candidates:
                extracted_candidates.append(candidate)
        full_candidates = extracted_candidates + full_candidates

        for candidate in full_candidates:
            for fmt in (
                "%d.%m.%Y",
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d/%m/%Y",
                "%d.%m.%y",
                "%d-%m-%y",
                "%d/%m/%y",
            ):
                try:
                    return datetime.strptime(candidate, fmt).date()
                except ValueError:
                    continue

        # Support day-month values without explicit year, e.g. "03.03 (Вторник)".
        short_match = re.search(r"(\d{1,2})[./-](\d{1,2})(?![./-]\d)", text)
        if short_match:
            day = int(short_match.group(1))
            month = int(short_match.group(2))
            today = date.today()
            candidates: list[date] = []
            for year in (today.year, today.year + 1):
                try:
                    candidates.append(date(year, month, day))
                except ValueError:
                    continue

            if candidates:
                future = [item for item in candidates if item >= today]
                if future:
                    return min(future, key=lambda item: item.toordinal())
                return candidates[0]
    return None


def parse_time_cell(value: object) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        candidates = [text]
        match = re.search(r"(\d{1,2}:\d{2}(?::\d{2})?)", text)
        if match:
            candidates.insert(0, match.group(1))

        for candidate in candidates:
            for fmt in ("%H:%M", "%H:%M:%S"):
                try:
                    return datetime.strptime(candidate, fmt).time().replace(
                        second=0, microsecond=0
                    )
                except ValueError:
                    continue
    return None


def parse_slot_duration_hours(value: object) -> float | None:
    if not isinstance(value, str):
        return None

    tokens = re.findall(r"(\d{1,2}:\d{2}(?::\d{2})?)", value)
    if len(tokens) < 2:
        return None

    def parse_token(token: str) -> datetime | None:
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(token, fmt)
            except ValueError:
                continue
        return None

    start = parse_token(tokens[0])
    end = parse_token(tokens[1])
    if not start or not end:
        return None

    delta_hours = (end - start).total_seconds() / 3600.0
    if delta_hours <= 0:
        delta_hours += 24.0
    if delta_hours <= 0:
        return None

    return delta_hours


def parse_time_range(value: object) -> tuple[time, time] | None:
    if not isinstance(value, str):
        return None

    tokens = re.findall(r"(\d{1,2}:\d{2}(?::\d{2})?)", value)
    if len(tokens) < 2:
        return None

    parsed: list[time] = []
    for token in tokens[:2]:
        parsed_token = parse_time_cell(token)
        if not parsed_token:
            return None
        parsed.append(parsed_token)

    return parsed[0], parsed[1]


def combine_date_and_time(slot_date: date, slot_time: time) -> datetime:
    return datetime.combine(slot_date, slot_time)


def interval_from_time_cell(
    slot_date: date, time_value: object, fallback_duration_hours: float | None
) -> tuple[datetime, datetime] | None:
    time_range = parse_time_range(time_value)
    if time_range:
        start_dt = combine_date_and_time(slot_date, time_range[0])
        end_dt = combine_date_and_time(slot_date, time_range[1])
        if end_dt <= start_dt:
            end_dt += timedelta(days=1)
        return start_dt, end_dt

    start_time = parse_time_cell(time_value)
    if not start_time:
        return None

    if fallback_duration_hours is None or fallback_duration_hours <= 0:
        return None

    start_dt = combine_date_and_time(slot_date, start_time)
    end_dt = start_dt + timedelta(hours=fallback_duration_hours)
    return start_dt, end_dt


def manual_slot_key(slot_date: date, time_value: object) -> str | None:
    if isinstance(time_value, str):
        parsed_range = parse_time_range(time_value)
        if parsed_range:
            start_time, end_time = parsed_range
            return (
                f"{slot_date.strftime('%Y%m%d')}"
                f"{start_time.strftime('%H%M')}"
                f"{end_time.strftime('%H%M')}"
            )

    start_time = parse_time_cell(time_value)
    if not start_time:
        return None
    return f"{slot_date.strftime('%Y%m%d')}{start_time.strftime('%H%M')}"


def generated_slot_key(slot_date: date, start_time: time) -> str:
    return f"{slot_date.strftime('%Y%m%d')}{start_time.strftime('%H%M')}"


def parse_generated_slot_key(key: str) -> tuple[date, time] | None:
    if not re.fullmatch(r"\d{12}", key):
        return None

    date_part = key[:8]
    time_part = key[8:]
    try:
        slot_date = datetime.strptime(date_part, "%Y%m%d").date()
        slot_time = datetime.strptime(time_part, "%H%M").time()
    except ValueError:
        return None
    return slot_date, slot_time


def date_label_with_weekday(slot_date: date) -> str:
    return f"{slot_date.strftime('%d.%m')} ({WEEKDAY_NAMES_RU[slot_date.weekday()]})"


def weekday_marker(slot_date: date) -> str:
    return WEEKDAY_RAINBOW_MARKERS.get(slot_date.weekday(), "⬜")


def day_button_label(base_label: str, slot_date: date | None) -> str:
    if slot_date is None:
        return base_label
    return f"{weekday_marker(slot_date)} {base_label}"


def week_start(slot_date: date) -> date:
    return slot_date - timedelta(days=slot_date.weekday())


def format_date_display(raw: object, parsed: date | None) -> str:
    if isinstance(raw, str) and raw.strip():
        return raw.strip()
    if parsed:
        return parsed.strftime("%d.%m.%Y")
    return str(raw).strip()


def format_time_display(raw: object, parsed: time | None) -> str:
    if isinstance(raw, str) and raw.strip():
        return raw.strip()
    if parsed:
        return parsed.strftime("%H:%M")
    return str(raw).strip()


def format_date(value: object) -> str:
    parsed = parse_date_cell(value)
    return format_date_display(value, parsed)


def format_time(value: object) -> str:
    parsed = parse_time_cell(value)
    return format_time_display(value, parsed)


def get_user_handles(update: Update) -> tuple[str, list[str]]:
    user = update.effective_user
    if not user:
        return "unknown", ["unknown"]

    id_handle = f"id:{user.id}"
    username_handle = f"@{user.username}" if user.username else None
    primary = id_handle if not username_handle else f"{id_handle} {username_handle}"
    handles = [primary, id_handle]
    if user.username:
        handles.append(username_handle)

    # Keep unique values preserving order.
    unique_handles = list(dict.fromkeys(handles))
    return primary, unique_handles


def normalize_telegram_handle(value: str) -> str:
    text = (value or "").strip()
    if text.startswith("@"):
        return "@" + text[1:].lower()
    return text


def telegram_cell_matches_handles(cell_value: str, user_handles: list[str]) -> bool:
    normalized_handles = {normalize_telegram_handle(handle) for handle in user_handles}
    if not normalized_handles:
        return False

    candidates = {normalize_telegram_handle(cell_value)}
    for token in re.findall(r"id:\d+|@[A-Za-z0-9_]+", cell_value):
        candidates.add(normalize_telegram_handle(token))

    return bool(candidates & normalized_handles)


def is_labshake_busy_value(value: object) -> bool:
    if is_empty(value):
        return False
    marker = str(value).strip().lower()
    return marker in {
        LABSHAKE_BUSY_MARKER.lower(),
        RESEARCHER_BLOCK_MARKER.lower(),
    }


def is_labshake_source_busy_value(value: object) -> bool:
    if is_empty(value):
        return False
    return str(value).strip().lower() == LABSHAKE_BUSY_MARKER.lower()


def read_slot_info(ws, row: int, cols: SheetColumns) -> dict | None:
    date_value = ws.cell(row=row, column=cols.date).value
    time_value = ws.cell(row=row, column=cols.time).value

    if is_empty(date_value) or is_empty(time_value):
        return None

    slot_date = parse_date_cell(date_value)
    slot_time = parse_time_cell(time_value)
    if not slot_date or not slot_time:
        return None

    date_label = format_date_display(date_value, slot_date)
    time_label = format_time_display(time_value, slot_time)
    label = f"{date_label} {time_label}"
    duration_hours = parse_slot_duration_hours(time_value)
    slot_key = manual_slot_key(slot_date, time_value) or generated_slot_key(slot_date, slot_time)
    return {
        "row": row,
        "manual_key": slot_key,
        "label": label,
        "slot_date": slot_date,
        "slot_time": slot_time,
        "duration_hours": duration_hours,
    }


def slot_start_time_for_sort(value: object) -> time | None:
    if isinstance(value, str):
        parsed_range = parse_time_range(value)
        if parsed_range:
            return parsed_range[0]
    return parse_time_cell(value)


def slot_row_sort_key(values: list[object], cols: SheetColumns) -> tuple:
    def get_value(column_index: int) -> object:
        idx = column_index - 1
        if 0 <= idx < len(values):
            return values[idx]
        return None

    date_value = get_value(cols.date)
    time_value = get_value(cols.time)
    telegram_value = get_value(cols.telegram)

    slot_date = parse_date_cell(date_value)
    slot_time = slot_start_time_for_sort(time_value)
    date_text = str(date_value).strip().lower() if not is_empty(date_value) else ""
    time_text = str(time_value).strip().lower() if not is_empty(time_value) else ""
    booked = not is_empty(telegram_value)
    is_system_busy = is_labshake_busy_value(telegram_value)

    if is_system_busy:
        group_order = 2  # Always at the end.
    else:
        group_order = 0 if booked else 1  # booked first, free second

    return (
        group_order,
        1 if slot_date is None else 0,
        slot_date or date.max,
        1 if slot_time is None else 0,
        slot_time or time.max,
        date_text,
        time_text,
    )


def sort_slots_sheet_rows(ws, cols: SheetColumns) -> None:
    if ws.max_row < 2:
        return

    max_col = max(
        ws.max_column,
        cols.date,
        cols.time,
        cols.full_name,
        cols.phone,
        cols.telegram,
        cols.booked_at or 0,
    )
    rows: list[list[object]] = []
    for row in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row, column=col).value for col in range(1, max_col + 1)]
        if all(is_empty(cell_value) for cell_value in row_values):
            continue
        rows.append(row_values)

    rows.sort(key=lambda row_values: slot_row_sort_key(row_values, cols))

    ws.delete_rows(2, max(0, ws.max_row - 1))
    for row_values in rows:
        ws.append(row_values)

    for row_idx, row_values in enumerate(rows, start=2):
        tg_idx = cols.telegram - 1
        tg_value = row_values[tg_idx] if 0 <= tg_idx < len(row_values) else None
        set_booking_row_style(ws, row_idx, cols, booked=not is_empty(tg_value))


def resolve_slot_duration_hours(
    slot: dict, default_slot_duration_hours: float | None
) -> float:
    duration = slot.get("duration_hours")
    if isinstance(duration, (int, float)) and duration > 0:
        return float(duration)
    if default_slot_duration_hours is not None and default_slot_duration_hours > 0:
        return float(default_slot_duration_hours)
    return 0.0


def calculate_weekly_booked_hours(
    ws, cols: SheetColumns, default_slot_duration_hours: float | None
) -> dict[date, float]:
    weekly_hours: dict[date, float] = {}

    for row in range(2, ws.max_row + 1):
        tg_value = ws.cell(row=row, column=cols.telegram).value
        if is_empty(tg_value):
            continue
        if is_labshake_busy_value(tg_value):
            continue

        slot = read_slot_info(ws, row, cols)
        if not slot:
            continue

        week = week_start(slot["slot_date"])
        duration = resolve_slot_duration_hours(slot, default_slot_duration_hours)
        weekly_hours[week] = weekly_hours.get(week, 0.0) + duration

    return weekly_hours


def slot_exceeds_weekly_limit(
    slot: dict,
    weekly_booked_hours: dict[date, float],
    max_weekly_hours: float | None,
    default_slot_duration_hours: float | None,
) -> bool:
    if max_weekly_hours is None:
        return False

    booked = weekly_booked_hours.get(week_start(slot["slot_date"]), 0.0)
    if booked >= max_weekly_hours - 1e-9:
        return True

    duration = resolve_slot_duration_hours(slot, default_slot_duration_hours)
    return booked + duration > max_weekly_hours + 1e-9


def generated_slot_conflicts(
    candidate_start: datetime,
    candidate_end: datetime,
    booked_intervals: list[tuple[datetime, datetime, bool, int]],
    min_gap_hours: float,
) -> bool:
    participant_gap = timedelta(hours=min_gap_hours)
    for booked_start, booked_end, is_system_busy, _ in booked_intervals:
        # LabShake/system busy intervals are already stored with their own real boundaries.
        # Do not add extra participant gap on top of them.
        gap = timedelta(0) if is_system_busy else participant_gap
        if candidate_start < booked_end + gap and candidate_end + gap > booked_start:
            return True
    return False


def collect_day_windows_and_bookings(
    ws,
    cols: SheetColumns,
    *,
    slot_duration_hours: float | None,
    default_slot_duration_hours: float | None,
    ignore_booking_row: int | None = None,
) -> tuple[
    dict[date, list[tuple[datetime, datetime, str]]],
    dict[date, list[tuple[datetime, datetime, bool, int]]],
    dict[date, float],
]:
    day_windows: dict[date, list[tuple[datetime, datetime, str]]] = {}
    booked_by_day: dict[date, list[tuple[datetime, datetime, bool, int]]] = {}
    weekly_booked_hours: dict[date, float] = {}

    fallback_duration = (
        slot_duration_hours
        if slot_duration_hours is not None
        else default_slot_duration_hours
    )

    for row in range(2, ws.max_row + 1):
        raw_date_value = ws.cell(row=row, column=cols.date).value
        raw_time_value = ws.cell(row=row, column=cols.time).value
        if is_empty(raw_date_value) or is_empty(raw_time_value):
            continue

        slot_date = parse_date_cell(raw_date_value)
        if not slot_date:
            continue

        tg_value = ws.cell(row=row, column=cols.telegram).value
        if is_empty(tg_value):
            # In day-window mode rows without Telegram are treated as availability windows.
            window_range = parse_time_range(raw_time_value)
            if not window_range:
                continue

            start_dt = combine_date_and_time(slot_date, window_range[0])
            end_dt = combine_date_and_time(slot_date, window_range[1])
            if end_dt <= start_dt:
                end_dt += timedelta(days=1)

            day_windows.setdefault(slot_date, []).append(
                (start_dt, end_dt, date_label_with_weekday(slot_date))
            )
            continue

        if ignore_booking_row is not None and row == ignore_booking_row:
            continue

        booking_interval = interval_from_time_cell(
            slot_date, raw_time_value, fallback_duration
        )
        if not booking_interval:
            continue

        booking_start, booking_end = booking_interval
        is_system_busy = is_labshake_busy_value(tg_value)
        booked_by_day.setdefault(slot_date, []).append(
            (booking_start, booking_end, is_system_busy, row)
        )
        if not is_system_busy:
            weekly_key = week_start(slot_date)
            duration_hours = max(0.0, (booking_end - booking_start).total_seconds() / 3600.0)
            weekly_booked_hours[weekly_key] = weekly_booked_hours.get(weekly_key, 0.0) + duration_hours

    for values in booked_by_day.values():
        values.sort(key=lambda item: (item[0], item[1], item[3]))

    return day_windows, booked_by_day, weekly_booked_hours


def build_generated_available_slots(
    ws,
    cols: SheetColumns,
    *,
    working_hours: str | None,
    excluded_weekdays: set[int] | None,
    slot_duration_hours: float,
    min_gap_hours: float,
    slot_step_minutes: int,
    max_weekly_hours: float | None,
    default_slot_duration_hours: float | None,
    days_ahead: int = SLOT_HORIZON_DAYS,
    ignore_booking_row: int | None = None,
) -> list[dict]:
    day_windows, booked_by_day, weekly_booked_hours = collect_day_windows_and_bookings(
        ws,
        cols,
        slot_duration_hours=slot_duration_hours,
        default_slot_duration_hours=default_slot_duration_hours,
        ignore_booking_row=ignore_booking_row,
    )

    default_range = parse_time_range(working_hours) if working_hours else None
    horizon_days = max(1, int(days_ahead))
    if default_range:
        # If working hours are configured, generate base windows from them.
        # If explicit windows exist in Excel for a day (e.g. synchronized from LabShake),
        # restrict generation to intersection with those windows for that day.
        start_time, end_time = default_range
        source_day_windows = day_windows
        day_windows = {}
        today = date.today()
        for day_offset in range(horizon_days):
            slot_date = today + timedelta(days=day_offset)
            start_dt = combine_date_and_time(slot_date, start_time)
            end_dt = combine_date_and_time(slot_date, end_time)
            if end_dt <= start_dt:
                end_dt += timedelta(days=1)
            explicit_windows = source_day_windows.get(slot_date, [])
            if not explicit_windows:
                day_windows[slot_date] = [
                    (start_dt, end_dt, date_label_with_weekday(slot_date))
                ]
                continue

            intersections: list[tuple[datetime, datetime, str]] = []
            for window_start, window_end, date_label in explicit_windows:
                intersection_start = max(start_dt, window_start)
                intersection_end = min(end_dt, window_end)
                if intersection_end > intersection_start:
                    intersections.append((intersection_start, intersection_end, date_label))
            if intersections:
                day_windows[slot_date] = intersections

    step_delta = timedelta(minutes=slot_step_minutes)
    duration_delta = timedelta(hours=slot_duration_hours)
    today = date.today()
    horizon_end = today + timedelta(days=horizon_days - 1)
    slots: list[dict] = []
    seen_keys: set[str] = set()
    excluded = set(excluded_weekdays or set())

    for slot_date, windows in sorted(day_windows.items(), key=lambda x: x[0]):
        if slot_date < today:
            continue
        if slot_date > horizon_end:
            continue
        if slot_date.weekday() in excluded:
            continue

        booked_intervals = booked_by_day.get(slot_date, [])
        weekly_hours = weekly_booked_hours.get(week_start(slot_date), 0.0)
        if max_weekly_hours is not None and weekly_hours >= max_weekly_hours - 1e-9:
            continue

        for window_start, window_end, date_label in windows:
            start = window_start
            while start + duration_delta <= window_end:
                candidate_start = start
                candidate_end = candidate_start + duration_delta
                key = generated_slot_key(slot_date, candidate_start.time())
                start += step_delta

                if key in seen_keys:
                    continue
                if generated_slot_conflicts(
                    candidate_start, candidate_end, booked_intervals, min_gap_hours
                ):
                    continue
                if (
                    max_weekly_hours is not None
                    and weekly_hours + slot_duration_hours > max_weekly_hours + 1e-9
                ):
                    continue

                seen_keys.add(key)
                time_label = (
                    f"{candidate_start.strftime('%H:%M')}-{candidate_end.strftime('%H:%M')}"
                )
                slots.append(
                    {
                        "kind": "generated",
                        "key": key,
                        "slot_date": slot_date,
                        "slot_time": candidate_start.time(),
                        "duration_hours": slot_duration_hours,
                        "start_dt": candidate_start,
                        "end_dt": candidate_end,
                        "date_label": date_label,
                        "time_label": time_label,
                        "label": f"{date_label} {time_label}",
                    }
                )

    slots.sort(key=lambda x: (x["slot_date"], x["slot_time"], x["key"]))
    return slots


def find_generated_slot_by_key(
    ws,
    cols: SheetColumns,
    *,
    key: str,
    working_hours: str | None,
    excluded_weekdays: set[int] | None,
    slot_duration_hours: float,
    min_gap_hours: float,
    slot_step_minutes: int,
    max_weekly_hours: float | None,
    default_slot_duration_hours: float | None,
    days_ahead: int = SLOT_HORIZON_DAYS,
    ignore_booking_row: int | None = None,
) -> dict | None:
    slots = build_generated_available_slots(
        ws,
        cols,
        working_hours=working_hours,
        excluded_weekdays=excluded_weekdays,
        slot_duration_hours=slot_duration_hours,
        min_gap_hours=min_gap_hours,
        slot_step_minutes=slot_step_minutes,
        max_weekly_hours=max_weekly_hours,
        default_slot_duration_hours=default_slot_duration_hours,
        days_ahead=days_ahead,
        ignore_booking_row=ignore_booking_row,
    )
    for slot in slots:
        if slot["key"] == key:
            return slot
    return None


def find_user_booking_in_sheet(
    ws, user_handles: list[str], cols: SheetColumns
) -> dict | None:
    if not user_handles:
        return None

    candidates: list[dict] = []
    for row in range(2, ws.max_row + 1):
        tg_value = ws.cell(row=row, column=cols.telegram).value
        if is_empty(tg_value):
            continue

        tg_text = str(tg_value).strip()
        if not telegram_cell_matches_handles(tg_text, user_handles):
            continue

        slot = read_slot_info(ws, row, cols)
        if not slot:
            continue

        full_name_value = ws.cell(row=row, column=cols.full_name).value
        phone_value = ws.cell(row=row, column=cols.phone).value
        candidates.append(
            {
                **slot,
                "telegram": tg_text,
                "full_name": (
                    str(full_name_value).strip() if not is_empty(full_name_value) else ""
                ),
                "phone": str(phone_value).strip() if not is_empty(phone_value) else "",
            }
        )

    if not candidates:
        return None

    today = date.today()
    future = [item for item in candidates if item["slot_date"] >= today]
    target_pool = future if future else candidates
    target_pool.sort(key=lambda x: (x["slot_date"], x["slot_time"], x["row"]))
    return target_pool[0]


def write_booking_to_row(
    ws, row: int, telegram_handle: str, full_name: str, phone: str, cols: SheetColumns
) -> None:
    ws.cell(row=row, column=cols.telegram).value = telegram_handle
    ws.cell(row=row, column=cols.full_name).value = full_name
    ws.cell(row=row, column=cols.phone).value = phone
    if cols.booked_at:
        ws.cell(row=row, column=cols.booked_at).value = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    set_booking_row_style(ws, row, cols, booked=True)


def clear_booking_row(ws, row: int, cols: SheetColumns) -> None:
    ws.cell(row=row, column=cols.telegram).value = None
    ws.cell(row=row, column=cols.full_name).value = None
    ws.cell(row=row, column=cols.phone).value = None
    if cols.booked_at:
        ws.cell(row=row, column=cols.booked_at).value = None
    set_booking_row_style(ws, row, cols, booked=False)


def get_available_slots(
    storage: ExcelStorage,
    max_weekly_hours: float | None = None,
    default_slot_duration_hours: float | None = None,
    slot_mode: str = "manual",
    working_hours: str | None = None,
    excluded_weekdays: set[int] | None = None,
    slot_duration_hours: float | None = None,
    min_gap_hours: float = 0.0,
    slot_step_minutes: int = 60,
    days_ahead: int = SLOT_HORIZON_DAYS,
) -> tuple[list[dict], str | None]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return [], load_error
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            excluded = set(excluded_weekdays or set())
            today = date.today()
            horizon_days = max(1, int(days_ahead))
            horizon_end = today + timedelta(days=horizon_days - 1)

            if slot_mode == "day_windows":
                if slot_duration_hours is None or slot_duration_hours <= 0:
                    return [], "slot_duration_hours is required for day_windows mode."
                slots = build_generated_available_slots(
                    ws,
                    cols,
                    working_hours=working_hours,
                    excluded_weekdays=excluded,
                    slot_duration_hours=slot_duration_hours,
                    min_gap_hours=min_gap_hours,
                    slot_step_minutes=slot_step_minutes,
                    max_weekly_hours=max_weekly_hours,
                    default_slot_duration_hours=default_slot_duration_hours,
                    days_ahead=horizon_days,
                )
                return slots, None

            weekly_booked_hours = calculate_weekly_booked_hours(
                ws, cols, default_slot_duration_hours
            )
            slots: list[dict] = []
            for row in range(2, ws.max_row + 1):
                if not is_empty(ws.cell(row=row, column=cols.telegram).value):
                    continue

                slot = read_slot_info(ws, row, cols)
                if not slot:
                    continue
                if slot["slot_date"] < today:
                    continue
                if slot["slot_date"] > horizon_end:
                    continue
                if slot["slot_date"].weekday() in excluded:
                    continue
                if slot_exceeds_weekly_limit(
                    slot,
                    weekly_booked_hours,
                    max_weekly_hours,
                    default_slot_duration_hours,
                ):
                    continue
                slot["kind"] = "manual"
                slots.append(slot)

            slots.sort(key=lambda x: (x["slot_date"], x["slot_time"], x["row"]))
            return slots, None
        finally:
            wb.close()


def find_user_booking(
    storage: ExcelStorage, user_handles: list[str]
) -> tuple[dict | None, str | None]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return None, load_error
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            return find_user_booking_in_sheet(ws, user_handles, cols), None
        finally:
            wb.close()


def find_user_booking_with_retry(
    storage: ExcelStorage,
    user_handles: list[str],
    *,
    attempts: int = 3,
    delay_sec: float = 0.35,
) -> tuple[dict | None, str | None]:
    total = max(1, int(attempts))
    last_error: str | None = None
    for attempt in range(1, total + 1):
        booking, error = find_user_booking(storage, user_handles)
        if not error:
            return booking, None
        last_error = error
        if attempt < total:
            logger.warning(
                "find_user_booking failed (attempt %s/%s): %s",
                attempt,
                total,
                error,
            )
            time_module.sleep(max(0.0, float(delay_sec)))
    return None, last_error


def get_available_slots_with_retry(
    storage: ExcelStorage,
    *,
    attempts: int = 3,
    delay_sec: float = 0.35,
    **kwargs,
) -> tuple[list[dict], str | None]:
    total = max(1, int(attempts))
    last_error: str | None = None
    for attempt in range(1, total + 1):
        slots, error = get_available_slots(storage, **kwargs)
        if not error:
            return slots, None
        last_error = error
        if attempt < total:
            logger.warning(
                "get_available_slots failed (attempt %s/%s): %s",
                attempt,
                total,
                error,
            )
            time_module.sleep(max(0.0, float(delay_sec)))
    return [], last_error


def reserve_slot(
    storage: ExcelStorage,
    row: int,
    telegram_handle: str,
    full_name: str,
    phone: str,
    user_handles: list[str],
    max_weekly_hours: float | None = None,
    default_slot_duration_hours: float | None = None,
) -> tuple[bool, str]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return False, load_error
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)

            if find_user_booking_in_sheet(ws, user_handles, cols):
                return False, "У вас уже есть запись. Используйте кнопку «Перенести запись»."

            if row < 2 or row > ws.max_row:
                return False, "Слот не найден."

            slot = read_slot_info(ws, row, cols)
            if not slot:
                return False, "Слот невалидный."

            telegram_value = ws.cell(row=row, column=cols.telegram).value
            if not is_empty(telegram_value):
                return False, "Этот слот уже заняли."

            if slot_exceeds_weekly_limit(
                slot,
                calculate_weekly_booked_hours(ws, cols, default_slot_duration_hours),
                max_weekly_hours,
                default_slot_duration_hours,
            ):
                return False, "На этой неделе уже достигнут лимит часов по эксперименту."

            write_booking_to_row(ws, row, telegram_handle, full_name, phone, cols)
            sort_slots_sheet_rows(ws, cols)

            success, message = save_workbook_to_storage(storage, wb)
            if not success:
                return False, message

            return True, "ok"
        finally:
            wb.close()


def format_time_range(start_time: time, end_time: time) -> str:
    return f"{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}"


def reserve_generated_slot(
    storage: ExcelStorage,
    *,
    slot_key: str,
    working_hours: str | None,
    excluded_weekdays: set[int] | None,
    telegram_handle: str,
    full_name: str,
    phone: str,
    user_handles: list[str],
    slot_duration_hours: float,
    min_gap_hours: float,
    slot_step_minutes: int,
    max_weekly_hours: float | None = None,
    default_slot_duration_hours: float | None = None,
    days_ahead: int = SLOT_HORIZON_DAYS,
) -> tuple[bool, str]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return False, load_error
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)

            if find_user_booking_in_sheet(ws, user_handles, cols):
                return False, "У вас уже есть запись. Используйте кнопку «Перенести запись»."

            target_slot = find_generated_slot_by_key(
                ws,
                cols,
                key=slot_key,
                working_hours=working_hours,
                excluded_weekdays=excluded_weekdays,
                slot_duration_hours=slot_duration_hours,
                min_gap_hours=min_gap_hours,
                slot_step_minutes=slot_step_minutes,
                max_weekly_hours=max_weekly_hours,
                default_slot_duration_hours=default_slot_duration_hours,
                days_ahead=days_ahead,
            )
            if not target_slot:
                return False, "Этот слот недоступен. Выберите другой."

            row = ws.max_row + 1
            ws.cell(row=row, column=cols.date).value = target_slot["slot_date"].strftime("%d.%m.%Y")
            ws.cell(row=row, column=cols.time).value = format_time_range(
                target_slot["start_dt"].time(),
                target_slot["end_dt"].time(),
            )
            write_booking_to_row(ws, row, telegram_handle, full_name, phone, cols)
            sort_slots_sheet_rows(ws, cols)

            success, message = save_workbook_to_storage(storage, wb)
            if not success:
                return False, message

            return True, "ok"
        finally:
            wb.close()


def reserve_generated_slot_with_labshake(
    exp: ExperimentConfig,
    *,
    slot_key: str,
    working_hours: str | None,
    excluded_weekdays: set[int] | None,
    telegram_handle: str,
    full_name: str,
    phone: str,
    user_handles: list[str],
    slot_duration_hours: float,
    min_gap_hours: float,
    slot_step_minutes: int,
    max_weekly_hours: float | None = None,
    default_slot_duration_hours: float | None = None,
    days_ahead: int = SLOT_HORIZON_DAYS,
) -> tuple[bool, str]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(exp.storage)
        if load_error:
            return False, load_error
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)

            if find_user_booking_in_sheet(ws, user_handles, cols):
                return False, "У вас уже есть запись. Используйте кнопку «Перенести запись»."

            target_slot = find_generated_slot_by_key(
                ws,
                cols,
                key=slot_key,
                working_hours=working_hours,
                excluded_weekdays=excluded_weekdays,
                slot_duration_hours=slot_duration_hours,
                min_gap_hours=min_gap_hours,
                slot_step_minutes=slot_step_minutes,
                max_weekly_hours=max_weekly_hours,
                default_slot_duration_hours=default_slot_duration_hours,
                days_ahead=days_ahead,
            )
            if not target_slot:
                return False, "Этот слот недоступен. Выберите другой."

            reserve_ok, reserve_message = reserve_buffered_interval_in_labshake(
                exp=exp,
                slot_date=target_slot["slot_date"],
                selected_start=target_slot["start_dt"].time(),
                selected_end=target_slot["end_dt"].time(),
            )
            if not reserve_ok:
                return False, reserve_message

            row = ws.max_row + 1
            ws.cell(row=row, column=cols.date).value = target_slot["slot_date"].strftime("%d.%m.%Y")
            ws.cell(row=row, column=cols.time).value = format_time_range(
                target_slot["start_dt"].time(),
                target_slot["end_dt"].time(),
            )
            write_booking_to_row(ws, row, telegram_handle, full_name, phone, cols)
            sort_slots_sheet_rows(ws, cols)

            success, message = save_workbook_to_storage(exp.storage, wb)
            if not success:
                return False, message

            return True, "ok"
        finally:
            wb.close()


def cancel_user_booking(
    storage: ExcelStorage, user_handles: list[str], slot_mode: str = "manual"
) -> tuple[bool, str, str | None]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return False, load_error, None
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            current = find_user_booking_in_sheet(ws, user_handles, cols)
            if not current:
                return False, "У вас нет активной записи.", None

            if slot_mode == "day_windows":
                ws.delete_rows(current["row"], 1)
            else:
                clear_booking_row(ws, current["row"], cols)
            sort_slots_sheet_rows(ws, cols)
            success, message = save_workbook_to_storage(storage, wb)
            if not success:
                return False, message, None

            return True, "ok", current["label"]
        finally:
            wb.close()


def cancel_user_booking_with_labshake(
    exp: ExperimentConfig,
    user_handles: list[str],
    *,
    slot_mode: str = "manual",
    default_slot_duration_hours: float | None = None,
) -> tuple[bool, str, str | None]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(exp.storage)
        if load_error:
            return False, load_error, None
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            current = find_user_booking_in_sheet(ws, user_handles, cols)
            if not current:
                return False, "У вас нет активной записи.", None

            if slot_mode == "day_windows":
                raw_time_value = ws.cell(row=current["row"], column=cols.time).value
                fallback_duration = default_slot_duration_hours
                if (
                    fallback_duration is None or fallback_duration <= 0
                ) and exp.slot_duration_hours and exp.slot_duration_hours > 0:
                    fallback_duration = exp.slot_duration_hours
                interval = interval_from_time_cell(
                    current["slot_date"],
                    raw_time_value,
                    fallback_duration,
                )
                if not interval:
                    return (
                        False,
                        "Не удалось определить интервал текущей записи для отмены в LabShake.",
                        None,
                    )
                current_start = interval[0].time()
                current_end = interval[1].time()

                cancel_ok, cancel_message = cancel_buffered_interval_in_labshake(
                    exp=exp,
                    slot_date=current["slot_date"],
                    selected_start=current_start,
                    selected_end=current_end,
                )
                if not cancel_ok:
                    return False, cancel_message, None

            if slot_mode == "day_windows":
                ws.delete_rows(current["row"], 1)
            else:
                clear_booking_row(ws, current["row"], cols)
            sort_slots_sheet_rows(ws, cols)
            success, message = save_workbook_to_storage(exp.storage, wb)
            if not success:
                return False, message, None

            return True, "ok", current["label"]
        finally:
            wb.close()


def list_slots_for_admin(
    storage: ExcelStorage,
    *,
    only_future: bool = True,
    max_rows: int = 300,
) -> tuple[list[dict], str | None]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return [], load_error
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            today = date.today()
            slots: list[dict] = []

            for row in range(2, ws.max_row + 1):
                date_value = ws.cell(row=row, column=cols.date).value
                time_value = ws.cell(row=row, column=cols.time).value
                if is_empty(date_value) or is_empty(time_value):
                    continue

                slot_date = parse_date_cell(date_value)
                if only_future and slot_date and slot_date < today:
                    continue

                date_label = format_date_display(date_value, slot_date)
                time_label = format_time_display(time_value, parse_time_cell(time_value))
                tg_value = ws.cell(row=row, column=cols.telegram).value
                tg_text = str(tg_value).strip() if not is_empty(tg_value) else ""
                if tg_text.lower() == LABSHAKE_BUSY_MARKER.lower():
                    continue

                slots.append(
                    {
                        "row": row,
                        "label": f"{date_label} {time_label}",
                        "slot_date": slot_date,
                        "booked": bool(tg_text),
                        "telegram": tg_text,
                    }
                )

            slots.sort(
                key=lambda item: (
                    item["slot_date"] is None,
                    item["slot_date"] or date.max,
                    item["row"],
                )
            )
            return slots[:max_rows], None
        finally:
            wb.close()


def parse_admin_day_raw(day_raw: str) -> date | None:
    try:
        return datetime.strptime(day_raw, "%Y%m%d").date()
    except ValueError:
        return None


def list_days_for_admin_delete(
    exp: ExperimentConfig,
    *,
    only_future: bool = True,
) -> tuple[list[dict], str | None]:
    today = date.today()
    horizon_days = max(1, int(exp.available_days_ahead))
    horizon_end = today + timedelta(days=horizon_days - 1)
    available_by_day: dict[date, int] = {}
    booked_by_day: dict[date, int] = {}

    if exp.slot_mode == "day_windows" and exp.slot_duration_hours and exp.slot_duration_hours > 0:
        generated_slots, generated_error = get_available_slots(
            exp.storage,
            max_weekly_hours=exp.max_weekly_hours,
            default_slot_duration_hours=exp.default_slot_duration_hours,
            slot_mode=exp.slot_mode,
            working_hours=exp.working_hours,
            excluded_weekdays=exp.excluded_weekdays,
            slot_duration_hours=exp.slot_duration_hours,
            min_gap_hours=exp.min_gap_hours,
            slot_step_minutes=exp.slot_step_minutes,
            days_ahead=horizon_days,
        )
        if generated_error:
            return [], generated_error
        for slot in generated_slots:
            slot_date = slot["slot_date"]
            if only_future and slot_date < today:
                continue
            if slot_date > horizon_end:
                continue
            available_by_day[slot_date] = available_by_day.get(slot_date, 0) + 1

    with excel_lock:
        wb, load_error = load_workbook_from_storage(exp.storage)
        if load_error:
            return [], load_error
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)

            for row in range(2, ws.max_row + 1):
                slot = read_slot_info(ws, row, cols)
                if not slot:
                    continue

                slot_date = slot["slot_date"]
                if only_future and slot_date < today:
                    continue
                if slot_date > horizon_end:
                    continue

                tg_value = ws.cell(row=row, column=cols.telegram).value
                if is_empty(tg_value):
                    if exp.slot_mode == "manual":
                        available_by_day[slot_date] = available_by_day.get(slot_date, 0) + 1
                    continue

                if is_labshake_busy_value(tg_value):
                    continue
                booked_by_day[slot_date] = booked_by_day.get(slot_date, 0) + 1
        finally:
            wb.close()

    day_entries: list[dict] = []
    for slot_date in sorted(set(available_by_day.keys()) | set(booked_by_day.keys())):
        available_count = available_by_day.get(slot_date, 0)
        booked_count = booked_by_day.get(slot_date, 0)
        total_count = available_count + booked_count
        if total_count <= 0:
            continue
        day_entries.append(
            {
                "slot_date": slot_date,
                "day_raw": slot_date.strftime("%Y%m%d"),
                "label": date_label_with_weekday(slot_date),
                "available_count": available_count,
                "booked_count": booked_count,
                "total_count": total_count,
            }
        )
    return day_entries, None


def delete_slots_for_admin_interval(
    exp: ExperimentConfig,
    *,
    slot_date: date,
    start_time: time,
    end_time: time,
) -> tuple[bool, str, list[dict], list[str]]:
    target_start, target_end = interval_from_date_and_times(slot_date, start_time, end_time)
    rows_to_delete: list[int] = []
    canceled: list[dict] = []
    removed_labels: list[str] = []

    with excel_lock:
        wb, load_error = load_workbook_from_storage(exp.storage)
        if load_error:
            return False, load_error, [], []
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            fallback_duration = exp.default_slot_duration_hours or exp.slot_duration_hours

            for row in range(2, ws.max_row + 1):
                raw_date = ws.cell(row=row, column=cols.date).value
                raw_time = ws.cell(row=row, column=cols.time).value
                if is_empty(raw_date) or is_empty(raw_time):
                    continue

                row_date = parse_date_cell(raw_date)
                if row_date != slot_date:
                    continue

                row_interval = interval_from_time_cell(
                    slot_date, raw_time, fallback_duration
                )
                if not row_interval:
                    continue
                if not intervals_overlap(
                    row_interval[0], row_interval[1], target_start, target_end
                ):
                    continue

                tg_value = ws.cell(row=row, column=cols.telegram).value
                tg_text = str(tg_value).strip() if not is_empty(tg_value) else ""
                slot_label = (
                    f"{format_date_display(raw_date, row_date)} "
                    f"{format_time_display(raw_time, parse_time_cell(raw_time))}"
                )

                if exp.slot_mode == "day_windows":
                    if is_empty(tg_value):
                        # Source day-window rows are retained; the system block row below
                        # disables generated slots for this interval.
                        continue
                    if is_labshake_busy_value(tg_value):
                        rows_to_delete.append(row)
                        continue

                    rows_to_delete.append(row)
                    removed_labels.append(slot_label)
                    canceled.append(
                        {
                            "label": slot_label,
                            "telegram": tg_text,
                            "chat_id": parse_chat_id_from_telegram_cell(tg_text),
                        }
                    )
                    continue

                rows_to_delete.append(row)
                removed_labels.append(slot_label)
                if tg_text and not is_labshake_busy_value(tg_text):
                    canceled.append(
                        {
                            "label": slot_label,
                            "telegram": tg_text,
                            "chat_id": parse_chat_id_from_telegram_cell(tg_text),
                        }
                    )

            if exp.slot_mode != "day_windows" and not rows_to_delete:
                return False, "На выбранный интервал слоты не найдены.", [], []

            for row in sorted(set(rows_to_delete), reverse=True):
                ws.delete_rows(row, 1)

            if exp.slot_mode == "day_windows":
                block_row = ws.max_row + 1
                ws.cell(row=block_row, column=cols.date).value = slot_date.strftime(
                    "%d.%m.%Y"
                )
                ws.cell(row=block_row, column=cols.time).value = format_time_range(
                    start_time, end_time
                )
                ws.cell(row=block_row, column=cols.telegram).value = RESEARCHER_BLOCK_MARKER
                ws.cell(row=block_row, column=cols.full_name).value = (
                    "Заблокировано исследователем"
                )
                ws.cell(row=block_row, column=cols.phone).value = None
                if cols.booked_at:
                    ws.cell(row=block_row, column=cols.booked_at).value = None
                set_booking_row_style(ws, block_row, cols, booked=True)

            sort_slots_sheet_rows(ws, cols)
            success, message = save_workbook_to_storage(exp.storage, wb)
            if not success:
                return False, message, [], []
        finally:
            wb.close()

    return True, "ok", canceled, list(dict.fromkeys(removed_labels))


def parse_chat_id_from_telegram_cell(value: str) -> int | None:
    text = (value or "").strip()
    match = re.search(r"id:(\d+)", text)
    if match:
        try:
            return int(match.group(1))
        except ValueError:
            return None
    return None


def labshake_slot_key(slot_date: date, start_time: time, end_time: time) -> str:
    return (
        f"{slot_date.strftime('%Y-%m-%d')}"
        f"|{start_time.strftime('%H:%M')}"
        f"|{end_time.strftime('%H:%M')}"
    )


def interval_from_date_and_times(
    slot_date: date, start_time: time, end_time: time
) -> tuple[datetime, datetime]:
    start_dt = combine_date_and_time(slot_date, start_time)
    end_dt = combine_date_and_time(slot_date, end_time)
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)
    return start_dt, end_dt


def intervals_overlap(
    left_start: datetime,
    left_end: datetime,
    right_start: datetime,
    right_end: datetime,
) -> bool:
    return left_start < right_end and left_end > right_start


def pick_best_busy_interval_for_request(
    *,
    slot_date: date,
    reserve_start_dt: datetime,
    reserve_end_dt: datetime,
    busy_intervals: list[dict],
) -> tuple[datetime, datetime] | None:
    candidates: list[tuple[datetime, datetime]] = []
    for item in busy_intervals:
        if item["slot_date"] != slot_date:
            continue
        busy_start, busy_end = interval_from_date_and_times(
            slot_date,
            item["start_time"],
            item["end_time"],
        )
        if intervals_overlap(reserve_start_dt, reserve_end_dt, busy_start, busy_end):
            candidates.append((busy_start, busy_end))
    if not candidates:
        return None

    def candidate_score(interval: tuple[datetime, datetime]) -> tuple[float, float]:
        busy_start, busy_end = interval
        span_sec = max(0.0, (busy_end - busy_start).total_seconds())
        drift_sec = abs((busy_start - reserve_start_dt).total_seconds()) + abs(
            (busy_end - reserve_end_dt).total_seconds()
        )
        return span_sec, drift_sec

    candidates.sort(key=candidate_score)
    return candidates[0]


def has_busy_interval_similar_to_target(
    *,
    slot_date: date,
    reserve_start_dt: datetime,
    reserve_end_dt: datetime,
    busy_intervals: list[dict],
    max_boundary_drift_minutes: int = 20,
) -> bool:
    matched = pick_best_busy_interval_for_request(
        slot_date=slot_date,
        reserve_start_dt=reserve_start_dt,
        reserve_end_dt=reserve_end_dt,
        busy_intervals=busy_intervals,
    )
    if not matched:
        return False
    matched_start, matched_end = matched
    drift_sec = max(
        abs((matched_start - reserve_start_dt).total_seconds()),
        abs((matched_end - reserve_end_dt).total_seconds()),
    )
    return drift_sec <= max(1, int(max_boundary_drift_minutes)) * 60


def has_any_overlapping_busy_interval(
    *,
    slot_date: date,
    reserve_start_dt: datetime,
    reserve_end_dt: datetime,
    busy_intervals: list[dict],
) -> bool:
    for item in busy_intervals:
        if item["slot_date"] != slot_date:
            continue
        busy_start, busy_end = interval_from_date_and_times(
            slot_date,
            item["start_time"],
            item["end_time"],
        )
        if intervals_overlap(reserve_start_dt, reserve_end_dt, busy_start, busy_end):
            return True
    return False


def page_has_overlapping_labshake_reservation(
    page,
    *,
    reserve_start: time,
    reserve_end: time,
) -> bool:
    try:
        rows = page.locator("tr, .reservation-slot, .busy-slot")
        row_count = min(rows.count(), 420)
    except Exception:
        row_count = 0

    for index in range(row_count):
        row = rows.nth(index)
        try:
            if not row.is_visible():
                continue
        except Exception:
            continue
        try:
            row_text = re.sub(r"\s+", " ", row.inner_text() or "").strip()
        except Exception:
            continue
        if not row_text:
            continue
        lower = row_text.lower()
        if "click to reserve" in lower or "open - click" in lower:
            continue
        if "reservation" not in lower and "my reservation" not in lower:
            continue
        if row_matches_labshake_busy_interval(
            row_text,
            reserve_start=reserve_start,
            reserve_end=reserve_end,
            allow_overlap=True,
        ):
            return True
    return False


def parse_labshake_schedule(
    html_text: str, *, days_ahead: int
) -> tuple[list[dict], list[dict], str | None]:
    soup = BeautifulSoup(html_text, "html.parser")
    for node in soup(["script", "style", "noscript"]):
        node.decompose()

    today = date.today()
    horizon_end = today + timedelta(days=max(1, days_ahead) - 1)
    open_by_key: dict[str, dict] = {}
    busy_by_key: dict[str, dict] = {}
    full_text = soup.get_text("\n")

    def normalize_line(raw_line: str) -> str:
        text = html.unescape(raw_line)
        return re.sub(r"\s+", " ", text).strip()

    lines: list[str] = []
    for item in full_text.splitlines():
        normalized = normalize_line(item)
        if normalized:
            lines.append(normalized)
    if not lines:
        return [], [], "Страница LabShake пустая."

    def parse_first_date(raw_value: object) -> date | None:
        if raw_value is None:
            return None
        raw_text = str(raw_value).strip()
        if not raw_text:
            return None
        for match in LABSHAKE_DATE_TOKEN_RE.finditer(raw_text):
            parsed = parse_date_cell(match.group(1))
            if parsed:
                return parsed

        decoded = raw_text.replace("&amp;", "&")
        parsed_url = urlparse(decoded)
        raw_query = parsed_url.query or decoded.lstrip("?")
        query_items = dict(parse_qsl(raw_query, keep_blank_values=True))
        year_raw = query_items.get("y")
        month_raw = query_items.get("m")
        day_raw = query_items.get("d")
        if year_raw and month_raw and day_raw:
            try:
                parsed_from_query = date(int(year_raw), int(month_raw), int(day_raw))
            except ValueError:
                parsed_from_query = None
            if parsed_from_query:
                return parsed_from_query
        return None

    page_dates: list[date] = []
    for match in LABSHAKE_DATE_TOKEN_RE.finditer(full_text):
        parsed = parse_date_cell(match.group(1))
        if parsed and parsed not in page_dates:
            page_dates.append(parsed)
    fallback_page_date = page_dates[0] if page_dates else None

    def is_within_horizon(slot_date: date | None) -> bool:
        if slot_date is None:
            return False
        return today <= slot_date <= horizon_end

    def add_interval(
        slot_date: date | None,
        start_time: time,
        end_time: time,
        *,
        is_open: bool,
        is_busy: bool,
    ) -> None:
        if not is_within_horizon(slot_date):
            return
        assert slot_date is not None
        key = labshake_slot_key(slot_date, start_time, end_time)
        payload = {
            "key": key,
            "slot_date": slot_date,
            "start_time": start_time,
            "end_time": end_time,
        }
        if is_busy:
            busy_by_key[key] = payload
            open_by_key.pop(key, None)
            return
        if is_open and key not in busy_by_key:
            open_by_key[key] = payload

    def classify_text(text_value: str) -> tuple[bool, bool]:
        text_low = text_value.lower()
        has_open = (
            "click to reserve" in text_low
            or "open - click" in text_low
            or "available-slot" in text_low
            or "open-slot" in text_low
        )
        has_busy = (
            "my reservation" in text_low
            or "reservation-slot" in text_low
            or "busy-slot" in text_low
            or (
                "reservation" in text_low
                and "click to reserve" not in text_low
                and "open - click" not in text_low
            )
        )
        return has_open, has_busy

    def classify_node(node_text: str, class_tokens: list[str]) -> tuple[bool, bool]:
        text_open, text_busy = classify_text(node_text)
        classes_low = " ".join(token.lower() for token in class_tokens)
        class_open = "available-slot" in classes_low or "open-slot" in classes_low
        class_busy = "reservation-slot" in classes_low or "busy-slot" in classes_low
        is_open = text_open or class_open
        is_busy = text_busy or class_busy
        if is_open and is_busy:
            if "click to reserve" in node_text.lower() or class_open:
                is_busy = False
            else:
                is_open = False
        return is_open, is_busy

    def resolve_date_for_node(node: object) -> date | None:
        node_text = normalize_line(node.get_text(" ", strip=True))
        parsed_from_text = parse_first_date(node_text)
        if parsed_from_text:
            return parsed_from_text

        current = node
        depth = 0
        while current is not None and depth < 7:
            attrs = getattr(current, "attrs", {}) or {}
            for value in attrs.values():
                if isinstance(value, (list, tuple, set)):
                    values_to_check = value
                else:
                    values_to_check = [value]
                for candidate in values_to_check:
                    parsed = parse_first_date(candidate)
                    if parsed:
                        return parsed
            current = current.parent
            depth += 1

        return fallback_page_date

    def extract_slot_nodes() -> list:
        preferred = soup.select(".available-slot, .reservation-slot, .open-slot, .busy-slot")
        if preferred:
            return preferred

        nodes: list = []
        for candidate in soup.find_all(
            class_=lambda value: bool(
                value
                and any(
                    (
                        str(token).lower() == "slot"
                        or str(token).lower().endswith("-slot")
                    )
                    for token in (
                        value if isinstance(value, (list, tuple, set)) else str(value).split()
                    )
                )
            ),
        ):
            nodes.append(candidate)
        return nodes

    # Prefer structured slot blocks first: they preserve open/busy status reliably.
    for slot_node in extract_slot_nodes():
        node_text = normalize_line(slot_node.get_text(" ", strip=True))
        if not node_text:
            continue
        time_range = parse_time_range(node_text)
        if not time_range:
            continue
        slot_date = resolve_date_for_node(slot_node)
        class_tokens = [
            str(token) for token in (slot_node.get("class", []) or []) if str(token).strip()
        ]
        is_open, is_busy = classify_node(node_text, class_tokens)
        if not is_open and not is_busy:
            continue
        add_interval(
            slot_date,
            time_range[0],
            time_range[1],
            is_open=is_open,
            is_busy=is_busy,
        )

    # Fallback: nearby text context around time ranges.
    current_date = fallback_page_date
    line_count = len(lines)
    for index, line in enumerate(lines):
        parsed_date = parse_first_date(line)
        if parsed_date:
            current_date = parsed_date

        time_range = parse_time_range(line)
        if not time_range:
            continue

        left = max(0, index - 2)
        right = min(line_count, index + 5)
        context_text = " ".join(lines[left:right])
        is_open, is_busy = classify_text(context_text)
        if not is_open and not is_busy:
            continue
        if is_open and is_busy:
            if "click to reserve" in context_text.lower():
                is_busy = False
            else:
                is_open = False

        add_interval(
            current_date,
            time_range[0],
            time_range[1],
            is_open=is_open,
            is_busy=is_busy,
        )

    open_windows = sorted(
        open_by_key.values(),
        key=lambda item: (item["slot_date"], item["start_time"], item["end_time"]),
    )
    busy_intervals = sorted(
        busy_by_key.values(),
        key=lambda item: (item["slot_date"], item["start_time"], item["end_time"]),
    )
    return open_windows, busy_intervals, None


def parse_env_bool(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    text = raw.strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def get_labshake_login_credentials() -> tuple[str | None, str | None]:
    login_value = (
        os.getenv("LABSHAKE_LOGIN_EMAIL", "").strip()
        or os.getenv("LABSHAKE_LOGIN", "").strip()
    )
    password_value = (
        os.getenv("LABSHAKE_LOGIN_PASSWORD", "").strip()
        or os.getenv("LABSHAKE_PASSWORD", "").strip()
    )
    return (login_value or None), (password_value or None)


def is_labshake_auto_login_enabled() -> bool:
    raw = os.getenv("LABSHAKE_AUTO_LOGIN")
    if raw is not None and raw.strip() != "":
        return parse_env_bool("LABSHAKE_AUTO_LOGIN", default=False)
    login_value, password_value = get_labshake_login_credentials()
    return bool(login_value and password_value)


def build_cookie_header_from_browser_cookies(cookies: list[dict]) -> str:
    parts: list[str] = []
    for item in cookies:
        name = str(item.get("name", "")).strip()
        value = str(item.get("value", "")).strip()
        if not name or not value:
            continue
        parts.append(f"{name}={value}")
    return "; ".join(parts)


def refresh_labshake_cookie_via_playwright(
    *, schedule_url: str | None = None
) -> tuple[str | None, str | None]:
    login_value, password_value = get_labshake_login_credentials()
    if not login_value or not password_value:
        return (
            None,
            "Для автологина укажите LABSHAKE_LOGIN_EMAIL/LABSHAKE_LOGIN и "
            "LABSHAKE_LOGIN_PASSWORD/LABSHAKE_PASSWORD в .env.",
        )

    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return (
            None,
            "Не установлен Playwright. Установите зависимость и выполните: "
            "'python -m playwright install chromium'.",
        )

    login_url = os.getenv("LABSHAKE_LOGIN_URL", "https://labshake.com/sign-in").strip()
    if not login_url:
        login_url = "https://labshake.com/sign-in"

    timeout_raw = os.getenv("LABSHAKE_LOGIN_TIMEOUT_SEC", "45").strip()
    try:
        timeout_ms = max(10_000, int(float(timeout_raw.replace(",", ".")) * 1000))
    except ValueError:
        timeout_ms = 45_000
    headless = parse_env_bool("LABSHAKE_HEADLESS", default=True)
    browser_channel = os.getenv("LABSHAKE_BROWSER_CHANNEL", "chrome").strip().lower()
    if browser_channel == "default":
        browser_channel = ""

    def fill_first(page, selectors: list[str], value: str) -> bool:
        for selector in selectors:
            try:
                locator = page.locator(selector)
                if locator.count() > 0:
                    locator.first.fill(value)
                    return True
            except Exception:
                continue
        return False

    try:
        with sync_playwright() as p:
            launch_kwargs: dict[str, object] = {"headless": headless}
            if browser_channel:
                launch_kwargs["channel"] = browser_channel
            try:
                browser = p.chromium.launch(**launch_kwargs)
            except Exception:
                if "channel" in launch_kwargs:
                    logger.warning(
                        "Could not launch Playwright with channel '%s'. Fallback to default Chromium.",
                        browser_channel,
                    )
                    launch_kwargs.pop("channel", None)
                    browser = p.chromium.launch(**launch_kwargs)
                else:
                    raise
            context = browser.new_context()
            page = context.new_page()

            page.goto(login_url, wait_until="domcontentloaded", timeout=timeout_ms)

            email_ok = fill_first(
                page,
                [
                    "input[type='email']",
                    "input[name*='email' i]",
                    "input[name*='user' i]",
                    "input[name*='login' i]",
                    "input[id*='email' i]",
                    "input[id*='user' i]",
                ],
                login_value,
            )
            if not email_ok:
                browser.close()
                return None, "Не найдено поле логина на странице входа LabShake."

            password_ok = fill_first(
                page,
                [
                    "input[type='password']",
                    "input[name*='pass' i]",
                    "input[id*='pass' i]",
                ],
                password_value,
            )
            if not password_ok:
                browser.close()
                return None, "Не найдено поле пароля на странице входа LabShake."

            submitted = False
            for selector in [
                "button[type='submit']",
                "input[type='submit']",
                "button:has-text('Sign in')",
                "button:has-text('Log in')",
                "button:has-text('Войти')",
            ]:
                try:
                    locator = page.locator(selector)
                    if locator.count() > 0:
                        locator.first.click()
                        submitted = True
                        break
                except Exception:
                    continue
            if not submitted:
                page.keyboard.press("Enter")

            try:
                page.wait_for_load_state("networkidle", timeout=timeout_ms)
            except Exception:
                pass
            page.wait_for_timeout(1500)

            if schedule_url:
                try:
                    page.goto(
                        schedule_url,
                        wait_until="domcontentloaded",
                        timeout=timeout_ms,
                    )
                    try:
                        page.wait_for_load_state("networkidle", timeout=timeout_ms)
                    except Exception:
                        pass
                    page.wait_for_timeout(1000)
                except Exception:
                    logger.debug(
                        "Could not open schedule URL after LabShake login.",
                        exc_info=True,
                    )

                current_url = (page.url or "").lower()
                page_title = ""
                page_html = ""
                try:
                    page_title = page.title().lower()
                except Exception:
                    page_title = ""
                try:
                    page_html = page.content().lower()
                except Exception:
                    page_html = ""
                if (
                    "sign-in" in current_url
                    or "page not found" in page_title
                    or "page not found" in page_html
                    or "not found" in page_title
                ):
                    browser.close()
                    return (
                        None,
                        "Автологин выполнен, но этот аккаунт не имеет доступа к ссылке расписания.",
                    )

            cookies = context.cookies("https://labshake.com")
            cookie_header = build_cookie_header_from_browser_cookies(cookies)
            browser.close()
    except Exception as exc:
        return None, f"Ошибка автологина LabShake: {exc}"

    if not cookie_header:
        return None, "Автологин не вернул cookies LabShake."

    return cookie_header, None


def get_labshake_cookie_for_request(
    *, cookie_env: str, schedule_url: str, force_refresh: bool = False
) -> tuple[str | None, str | None]:
    cache_key = f"{cookie_env}::labshake"
    with labshake_auth_lock:
        if not force_refresh:
            cached = labshake_cookie_cache.get(cache_key)
            if cached:
                return cached, None

        env_cookie = os.getenv(cookie_env, "").strip()
        if env_cookie and not force_refresh:
            return env_cookie, None

        if not is_labshake_auto_login_enabled():
            if env_cookie:
                return env_cookie, None
            return (
                None,
                f"Не задано окружение {cookie_env} с cookie LabShake. "
                "Укажите cookie или включите автологин.",
            )

        cookie_header, login_error = refresh_labshake_cookie_via_playwright(
            schedule_url=schedule_url
        )
        if not cookie_header:
            if env_cookie:
                logger.warning(
                    "LabShake autologin failed, fallback to cookie from %s: %s",
                    cookie_env,
                    login_error,
                )
                return env_cookie, None
            return (
                None,
                "Автологин LabShake не удался. "
                + (login_error or "Не удалось получить cookie."),
            )

        labshake_cookie_cache[cache_key] = cookie_header
        os.environ[cookie_env] = cookie_header
        return cookie_header, None


def fetch_labshake_schedule_html(
    *, schedule_url: str, cookie_env: str
) -> tuple[str | None, str | None]:
    cookie, cookie_error = get_labshake_cookie_for_request(
        cookie_env=cookie_env,
        schedule_url=schedule_url,
        force_refresh=False,
    )
    if not cookie:
        return None, cookie_error or f"Не задано окружение {cookie_env} с cookie LabShake."

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/126.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }

    retry_attempts_raw = os.getenv("LABSHAKE_HTTP_RETRY_ATTEMPTS", "3").strip()
    retry_delay_raw = os.getenv("LABSHAKE_HTTP_RETRY_DELAY_SEC", "1.5").strip()
    max_retry_after_raw = os.getenv("LABSHAKE_MAX_RETRY_AFTER_SEC", "20").strip()
    try:
        retry_attempts = max(1, int(retry_attempts_raw))
    except ValueError:
        retry_attempts = 3
    try:
        retry_delay = max(0.0, float(retry_delay_raw.replace(",", ".")))
    except ValueError:
        retry_delay = 1.5
    try:
        max_retry_after = max(0.0, float(max_retry_after_raw.replace(",", ".")))
    except ValueError:
        max_retry_after = 20.0

    last_error: str | None = None
    response = None
    refreshed_auth = False
    for attempt in range(retry_attempts):
        headers["Cookie"] = cookie
        try:
            response = requests.get(schedule_url, headers=headers, timeout=30)
        except requests.RequestException as exc:
            last_error = f"Ошибка запроса к LabShake: {exc}"
            if attempt + 1 < retry_attempts:
                time_module.sleep(retry_delay * (attempt + 1))
                continue
            return None, last_error

        if response.status_code == 429:
            retry_after = response.headers.get("Retry-After", "").strip()
            delay_seconds = retry_delay * (attempt + 1)
            retry_after_value: float | None = None
            response_url = (response.url or "").lower()
            if "sign-in" in response_url or "login" in response_url:
                if not refreshed_auth:
                    refreshed_cookie, refresh_error = get_labshake_cookie_for_request(
                        cookie_env=cookie_env,
                        schedule_url=schedule_url,
                        force_refresh=True,
                    )
                    if refreshed_cookie:
                        cookie = refreshed_cookie
                        refreshed_auth = True
                        continue
                    return None, (
                        "LabShake вернул страницу входа (HTTP 429). "
                        "Похоже, cookie устарел или не подходит. "
                        + (refresh_error or "Обновите LABSHAKE_COOKIE и повторите синхронизацию.")
                    )
            try:
                if retry_after:
                    retry_after_value = float(retry_after)
            except ValueError:
                retry_after_value = None
            if retry_after_value is not None:
                delay_seconds = max(delay_seconds, retry_after_value)
            if max_retry_after > 0:
                delay_seconds = min(delay_seconds, max_retry_after)
            last_error = (
                "LabShake временно ограничил запросы (HTTP 429). "
                "Подождите немного и повторите."
            )
            if attempt + 1 < retry_attempts:
                time_module.sleep(delay_seconds)
                continue
            return None, last_error

        if response.status_code >= 500:
            last_error = f"LabShake временно недоступен (HTTP {response.status_code})."
            if attempt + 1 < retry_attempts:
                time_module.sleep(retry_delay * (attempt + 1))
                continue
            return None, last_error

        if response.status_code in {401, 403, 404}:
            path_lower = urlparse(schedule_url).path.lower()
            if "/shared-resource/" in path_lower and not refreshed_auth:
                refreshed_cookie, refresh_error = get_labshake_cookie_for_request(
                    cookie_env=cookie_env,
                    schedule_url=schedule_url,
                    force_refresh=True,
                )
                if refreshed_cookie:
                    cookie = refreshed_cookie
                    refreshed_auth = True
                    continue
                if response.status_code in {401, 403}:
                    return None, (
                        "LabShake отклонил доступ к расписанию. "
                        + (refresh_error or "Проверьте доступ к лаборатории.")
                    )
                # For 404 continue to generic HTTP handling below.

        final_url = (response.url or "").lower()
        body_lower = response.text.lower()
        if "login" in final_url or ("sign in" in body_lower and "password" in body_lower):
            if not refreshed_auth:
                refreshed_cookie, refresh_error = get_labshake_cookie_for_request(
                    cookie_env=cookie_env,
                    schedule_url=schedule_url,
                    force_refresh=True,
                )
                if refreshed_cookie:
                    cookie = refreshed_cookie
                    refreshed_auth = True
                    continue
                return None, (
                    "LabShake вернул страницу входа. "
                    + (
                        refresh_error
                        or "Проверьте cookie в .env (обычно cookie sessionid)."
                    )
                )
            return None, (
                "LabShake вернул страницу входа. "
                "Автовход уже выполнен, но доступ не получен."
            )

        break

    if response is None:
        return None, last_error or "Не удалось выполнить запрос к LabShake."

    if response.status_code >= 400:
        if response.status_code == 404 and "/shared-resource/" in urlparse(schedule_url).path.lower():
            return None, (
                "LabShake вернул HTTP 404 для этой ссылки расписания. "
                "Скорее всего, у текущего аккаунта нет доступа к этой lab "
                "или ссылка введена неверно."
            )
        return None, f"LabShake вернул HTTP {response.status_code}."

    return response.text, None


def build_labshake_day_url(schedule_url: str, slot_date: date) -> str:
    parsed = urlparse(schedule_url)
    query_items = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query_items["y"] = str(slot_date.year)
    query_items["m"] = str(slot_date.month)
    query_items["d"] = str(slot_date.day)
    new_query = urlencode(query_items, doseq=True)
    return urlunparse(parsed._replace(query=new_query))


def fetch_labshake_schedule_range_daily(
    *,
    schedule_url: str,
    cookie_env: str,
    days_ahead: int,
) -> tuple[list[dict], list[dict], set[date], list[date], str | None]:
    total_days = max(1, days_ahead)
    today = date.today()
    open_windows: list[dict] = []
    busy_intervals: list[dict] = []
    covered_days: set[date] = set()
    failed_days: list[date] = []
    failed_reasons: list[str] = []
    per_day_delay_raw = os.getenv("LABSHAKE_PER_DAY_DELAY_SEC", "0.35").strip()
    try:
        per_day_delay = max(0.0, float(per_day_delay_raw.replace(",", ".")))
    except ValueError:
        per_day_delay = 0.35

    for day_offset in range(total_days):
        target_day = today + timedelta(days=day_offset)
        day_url = build_labshake_day_url(schedule_url, target_day)
        html_text, fetch_error = fetch_labshake_schedule_html(
            schedule_url=day_url, cookie_env=cookie_env
        )
        if fetch_error:
            logger.warning("LabShake day fetch failed for %s: %s", target_day, fetch_error)
            failed_days.append(target_day)
            failed_reasons.append(fetch_error)
            if "HTTP 429" in fetch_error:
                break
            if per_day_delay > 0:
                time_module.sleep(per_day_delay)
            continue

        assert html_text is not None
        day_open, day_busy, parse_error = parse_labshake_schedule(
            html_text, days_ahead=total_days
        )
        if parse_error:
            logger.warning(
                "LabShake day parse failed for %s: %s", target_day, parse_error
            )
            failed_days.append(target_day)
            failed_reasons.append(parse_error)
            if per_day_delay > 0:
                time_module.sleep(per_day_delay)
            continue

        filtered_open = [item for item in day_open if item["slot_date"] == target_day]
        filtered_busy = [item for item in day_busy if item["slot_date"] == target_day]

        if not filtered_open and not filtered_busy and (day_open or day_busy):
            logger.warning(
                "LabShake day mismatch for %s (URL returned different date content).",
                target_day,
            )
            failed_days.append(target_day)
            failed_reasons.append("URL вернул расписание другого дня.")
            if per_day_delay > 0:
                time_module.sleep(per_day_delay)
            continue

        covered_days.add(target_day)
        open_windows.extend(filtered_open)
        busy_intervals.extend(filtered_busy)
        if per_day_delay > 0:
            time_module.sleep(per_day_delay)

    if not covered_days:
        reason_text = ""
        if failed_reasons:
            reason_text = f" Причина: {failed_reasons[0]}"
        return [], [], set(), failed_days, (
            "Не удалось загрузить данные LabShake ни за один день выбранного периода."
            + reason_text
        )

    unique_open: dict[str, dict] = {}
    unique_busy: dict[str, dict] = {}
    for item in open_windows:
        unique_open[item["key"]] = item
    for item in busy_intervals:
        unique_busy[item["key"]] = item

    merged_open = sorted(
        unique_open.values(),
        key=lambda item: (item["slot_date"], item["start_time"], item["end_time"]),
    )
    merged_busy = sorted(
        unique_busy.values(),
        key=lambda item: (item["slot_date"], item["start_time"], item["end_time"]),
    )
    return merged_open, merged_busy, covered_days, failed_days, None


def fetch_labshake_schedule_range_weekly(
    *,
    schedule_url: str,
    cookie_env: str,
    days_ahead: int,
) -> tuple[list[dict], list[dict], set[date], list[date], str | None]:
    total_days = max(1, days_ahead)
    today = date.today()
    horizon_end = today + timedelta(days=total_days - 1)
    open_windows: list[dict] = []
    busy_intervals: list[dict] = []
    covered_days: set[date] = set()
    failed_days: list[date] = []
    failed_reasons: list[str] = []
    per_request_delay_raw = os.getenv("LABSHAKE_PER_DAY_DELAY_SEC", "0.35").strip()
    try:
        per_request_delay = max(0.0, float(per_request_delay_raw.replace(",", ".")))
    except ValueError:
        per_request_delay = 0.35

    week_count = (total_days + 6) // 7
    for week_index in range(week_count):
        week_start = today + timedelta(days=week_index * 7)
        week_end = min(horizon_end, week_start + timedelta(days=6))
        week_days = {
            week_start + timedelta(days=i)
            for i in range((week_end - week_start).days + 1)
        }

        week_url = build_labshake_day_url(schedule_url, week_start)
        html_text, fetch_error = fetch_labshake_schedule_html(
            schedule_url=week_url,
            cookie_env=cookie_env,
        )
        if fetch_error:
            logger.warning(
                "LabShake weekly fetch failed for %s..%s: %s",
                week_start,
                week_end,
                fetch_error,
            )
            failed_days.extend(sorted(week_days))
            failed_reasons.append(fetch_error)
            if "HTTP 429" in fetch_error:
                break
            if per_request_delay > 0:
                time_module.sleep(per_request_delay)
            continue

        assert html_text is not None
        week_open, week_busy, parse_error = parse_labshake_schedule(
            html_text, days_ahead=total_days
        )
        if parse_error:
            logger.warning(
                "LabShake weekly parse failed for %s..%s: %s",
                week_start,
                week_end,
                parse_error,
            )
            failed_days.extend(sorted(week_days))
            failed_reasons.append(parse_error)
            if per_request_delay > 0:
                time_module.sleep(per_request_delay)
            continue

        filtered_open = [item for item in week_open if item["slot_date"] in week_days]
        filtered_busy = [item for item in week_busy if item["slot_date"] in week_days]

        if not filtered_open and not filtered_busy and (week_open or week_busy):
            logger.warning(
                "LabShake weekly mismatch for %s..%s (unexpected dates in payload).",
                week_start,
                week_end,
            )
            failed_days.extend(sorted(week_days))
            failed_reasons.append("Weekly URL вернул расписание другого диапазона дат.")
            if per_request_delay > 0:
                time_module.sleep(per_request_delay)
            continue

        covered_days.update(week_days)
        open_windows.extend(filtered_open)
        busy_intervals.extend(filtered_busy)
        if per_request_delay > 0:
            time_module.sleep(per_request_delay)

    if not covered_days:
        reason_text = ""
        if failed_reasons:
            reason_text = f" Причина: {failed_reasons[0]}"
        return [], [], set(), failed_days, (
            "Не удалось загрузить данные LabShake ни за одну неделю выбранного периода."
            + reason_text
        )

    unique_open: dict[str, dict] = {}
    unique_busy: dict[str, dict] = {}
    for item in open_windows:
        unique_open[item["key"]] = item
    for item in busy_intervals:
        unique_busy[item["key"]] = item

    merged_open = sorted(
        unique_open.values(),
        key=lambda item: (item["slot_date"], item["start_time"], item["end_time"]),
    )
    merged_busy = sorted(
        unique_busy.values(),
        key=lambda item: (item["slot_date"], item["start_time"], item["end_time"]),
    )
    return merged_open, merged_busy, covered_days, failed_days, None


def fetch_labshake_schedule_range(
    *,
    schedule_url: str,
    cookie_env: str,
    days_ahead: int,
) -> tuple[list[dict], list[dict], set[date], list[date], str | None]:
    path = urlparse(schedule_url).path.lower()
    if "/weekly" in path:
        weekly_open, weekly_busy, weekly_days, weekly_failed, weekly_error = (
            fetch_labshake_schedule_range_weekly(
                schedule_url=schedule_url,
                cookie_env=cookie_env,
                days_ahead=days_ahead,
            )
        )
        if weekly_error is not None:
            return weekly_open, weekly_busy, weekly_days, weekly_failed, weekly_error

        if weekly_open or weekly_busy:
            return weekly_open, weekly_busy, weekly_days, weekly_failed, None

        # Weekly layout can omit parsable interval text. Fallback to daily endpoint.
        daily_url = schedule_url.replace("/reservation/weekly", "/reservation")
        if daily_url == schedule_url:
            return weekly_open, weekly_busy, weekly_days, weekly_failed, weekly_error

        logger.info(
            "Weekly LabShake parsing returned no intervals, fallback to daily endpoint."
        )
        daily_open, daily_busy, daily_days, daily_failed, daily_error = (
            fetch_labshake_schedule_range_daily(
                schedule_url=daily_url,
                cookie_env=cookie_env,
                days_ahead=days_ahead,
            )
        )
        if daily_error is None:
            return daily_open, daily_busy, daily_days, daily_failed, None

        return daily_open, daily_busy, daily_days, daily_failed, daily_error

    return fetch_labshake_schedule_range_daily(
        schedule_url=schedule_url,
        cookie_env=cookie_env,
        days_ahead=days_ahead,
    )


def normalize_labshake_daily_url(schedule_url: str) -> str:
    normalized = schedule_url.strip()
    if "/reservation/weekly" in normalized:
        normalized = normalized.replace("/reservation/weekly", "/reservation")
    return normalized


def days_ahead_for_target_date(target_day: date) -> int:
    today = date.today()
    if target_day <= today:
        return 1
    return (target_day - today).days + 1


def fetch_labshake_schedule_for_day(
    *,
    schedule_url: str,
    cookie_env: str,
    target_day: date,
) -> tuple[list[dict], list[dict], str | None]:
    day_url = build_labshake_day_url(
        normalize_labshake_daily_url(schedule_url), target_day
    )
    html_text, fetch_error = fetch_labshake_schedule_html(
        schedule_url=day_url,
        cookie_env=cookie_env,
    )
    if fetch_error:
        return [], [], fetch_error
    assert html_text is not None
    open_windows, busy_intervals, parse_error = parse_labshake_schedule(
        html_text,
        days_ahead=days_ahead_for_target_date(target_day),
    )
    if parse_error:
        return [], [], parse_error

    open_for_day = [item for item in open_windows if item["slot_date"] == target_day]
    busy_for_day = [item for item in busy_intervals if item["slot_date"] == target_day]
    return open_for_day, busy_for_day, None


def parse_cookie_header(cookie_header: str) -> list[dict]:
    cookie = SimpleCookie()
    try:
        cookie.load(cookie_header)
    except Exception:
        return []

    items: list[dict] = []
    for key, morsel in cookie.items():
        value = str(morsel.value or "").strip()
        if not value:
            continue
        items.append(
            {
                "name": key,
                "value": value,
                "domain": "labshake.com",
                "path": "/",
                "httpOnly": False,
                "secure": True,
            }
        )
    return items


def apply_cookie_header_to_browser_context(context, cookie_header: str) -> None:
    cookies = parse_cookie_header(cookie_header)
    if cookies:
        context.add_cookies(cookies)


def is_labshake_login_page(page) -> bool:
    try:
        current_url = (page.url or "").lower()
    except Exception:
        current_url = ""
    if "sign-in" in current_url or "login" in current_url:
        return True
    try:
        page_text = page.content().lower()
    except Exception:
        return False
    return "sign in" in page_text and "password" in page_text


def labshake_fill_first(page, selectors: list[str], value: str) -> bool:
    for selector in selectors:
        try:
            locator = page.locator(selector)
            if locator.count() <= 0:
                continue
            target = locator.first
            if not target.is_visible():
                continue
            target.fill(value)
            return True
        except Exception:
            continue
    return False


def labshake_submit_login(page) -> bool:
    for selector in [
        "button[type='submit']",
        "input[type='submit']",
        "button:has-text('Sign in')",
        "button:has-text('Log in')",
        "button:has-text('Войти')",
    ]:
        try:
            locator = page.locator(selector)
            if locator.count() <= 0:
                continue
            target = locator.first
            if target.is_visible():
                target.click()
                return True
        except Exception:
            continue
    try:
        page.keyboard.press("Enter")
        return True
    except Exception:
        return False


def labshake_login_in_browser(page, *, timeout_ms: int) -> tuple[bool, str | None]:
    login_value, password_value = get_labshake_login_credentials()
    if not login_value or not password_value:
        return (
            False,
            "Для записи в LabShake настройте логин/пароль в .env: "
            "LABSHAKE_LOGIN_EMAIL и LABSHAKE_LOGIN_PASSWORD.",
        )

    login_url = os.getenv("LABSHAKE_LOGIN_URL", "https://labshake.com/sign-in").strip()
    if not login_url:
        login_url = "https://labshake.com/sign-in"

    page.goto(login_url, wait_until="domcontentloaded", timeout=timeout_ms)

    email_ok = labshake_fill_first(
        page,
        [
            "input[type='email']",
            "input[name*='email' i]",
            "input[name*='user' i]",
            "input[name*='login' i]",
            "input[id*='email' i]",
            "input[id*='user' i]",
        ],
        login_value,
    )
    if not email_ok:
        return False, "Не найдено поле логина на странице LabShake."

    password_ok = labshake_fill_first(
        page,
        [
            "input[type='password']",
            "input[name*='pass' i]",
            "input[id*='pass' i]",
        ],
        password_value,
    )
    if not password_ok:
        return False, "Не найдено поле пароля на странице LabShake."

    if not labshake_submit_login(page):
        return False, "Не удалось отправить форму входа LabShake."

    try:
        page.wait_for_load_state("networkidle", timeout=timeout_ms)
    except Exception:
        pass
    page.wait_for_timeout(1000)
    if is_labshake_login_page(page):
        return False, "Не удалось войти в LabShake. Проверьте логин/пароль."
    return True, None


def labshake_select_option(locator, value: int) -> bool:
    candidates = [f"{value:02d}", str(value)]
    for candidate in candidates:
        try:
            locator.select_option(value=candidate)
            return True
        except Exception:
            continue
    for candidate in candidates:
        try:
            locator.select_option(label=candidate)
            return True
        except Exception:
            continue
    try:
        options = locator.locator("option")
        option_count = min(60, options.count())
    except Exception:
        option_count = 0
    for index in range(option_count):
        try:
            option = options.nth(index)
            raw_value = str(option.get_attribute("value") or "").strip()
            raw_text = str(option.inner_text() or "").strip()
            if raw_value in candidates:
                locator.select_option(value=raw_value)
                return True
            if raw_text in candidates:
                locator.select_option(label=raw_text)
                return True
        except Exception:
            continue
    return False


def get_visible_labshake_time_selects(page, *, limit: int = 40) -> list:
    try:
        all_selects = page.locator("select")
        select_count = min(max(4, limit), all_selects.count())
    except Exception:
        return []

    visible_selects = []
    for index in range(select_count):
        locator = all_selects.nth(index)
        try:
            if locator.is_visible():
                visible_selects.append(locator)
        except Exception:
            continue
    return visible_selects


def disable_labshake_all_day_if_enabled(page) -> None:
    candidate_selectors = [
        "input[type='checkbox'][name*='all' i]",
        "input[type='checkbox'][id*='all' i]",
        "input[type='checkbox'][name*='full' i]",
        "input[type='checkbox'][id*='full' i]",
        "label:has-text('all day') input[type='checkbox']",
        "label:has-text('Reserve all day') input[type='checkbox']",
    ]
    for selector in candidate_selectors:
        try:
            locator = page.locator(selector)
            count = min(locator.count(), 16)
        except Exception:
            count = 0
        for index in range(count):
            checkbox = locator.nth(index)
            try:
                if not checkbox.is_visible():
                    continue
            except Exception:
                continue

            try:
                attr_text = " ".join(
                    [
                        str(checkbox.get_attribute("name") or ""),
                        str(checkbox.get_attribute("id") or ""),
                        str(checkbox.get_attribute("value") or ""),
                        str(checkbox.get_attribute("aria-label") or ""),
                        str(checkbox.get_attribute("title") or ""),
                    ]
                ).lower()
            except Exception:
                attr_text = ""

            if attr_text and (
                ("all" not in attr_text and "full" not in attr_text)
                or "day" not in attr_text
            ):
                continue

            try:
                if not checkbox.is_checked():
                    continue
            except Exception:
                continue

            try:
                checkbox.click(timeout=1800)
            except Exception:
                try:
                    checkbox.click(timeout=1800, force=True)
                except Exception:
                    continue
            page.wait_for_timeout(180)
            try:
                if checkbox.is_checked():
                    checkbox.evaluate(
                        "el => {el.checked = false; el.dispatchEvent(new Event('change', {bubbles: true}));}"
                    )
            except Exception:
                pass


def has_visible_labshake_reserve_button(page, *, max_count: int = 8) -> bool:
    for selector in [
        "button:has-text('Reserve')",
        "input[type='submit'][value*='Reserve' i]",
        "a:has-text('Reserve')",
    ]:
        try:
            locator = page.locator(selector)
            count = min(max_count, locator.count())
        except Exception:
            count = 0
        for index in range(count):
            try:
                if locator.nth(index).is_visible():
                    return True
            except Exception:
                continue
    return False


def is_labshake_reserve_form_visible(page) -> bool:
    visible_selects = get_visible_labshake_time_selects(page)
    if len(visible_selects) >= 4:
        return True

    if not has_visible_labshake_reserve_button(page):
        return False

    try:
        time_inputs = page.locator(
            "input[type='time'], input[name*='time' i], input[id*='time' i]"
        )
        input_count = min(time_inputs.count(), 12)
    except Exception:
        input_count = 0
    visible_inputs = 0
    for index in range(input_count):
        try:
            if time_inputs.nth(index).is_visible():
                visible_inputs += 1
        except Exception:
            continue
    return visible_inputs >= 2


def click_labshake_reserve_submit_near_time_selects(page, time_selects: list) -> bool:
    if not time_selects:
        return False

    reserve_selectors = [
        "button:has-text('Reserve')",
        "input[type='submit'][value*='Reserve' i]",
        "a:has-text('Reserve')",
    ]

    anchor = time_selects[0]
    for ancestor_xpath in [
        "xpath=ancestor::form[1]",
        "xpath=ancestor::tr[1]",
        "xpath=ancestor::table[1]",
        "xpath=ancestor::div[1]",
        "xpath=ancestor::td[1]",
    ]:
        try:
            container = anchor.locator(ancestor_xpath)
            if container.count() <= 0:
                continue
            scope = container.first
        except Exception:
            continue

        for selector in reserve_selectors:
            try:
                locator = scope.locator(selector)
                count = min(locator.count(), 8)
            except Exception:
                count = 0
            for index in range(count):
                target = locator.nth(index)
                try:
                    if not target.is_visible():
                        continue
                except Exception:
                    continue
                try:
                    target.click(timeout=3000)
                except Exception:
                    try:
                        target.click(timeout=3000, force=True)
                    except Exception:
                        continue
                return True

    for selector in reserve_selectors:
        try:
            locator = page.locator(selector)
            count = min(locator.count(), 16)
        except Exception:
            count = 0
        for index in range(count - 1, -1, -1):
            target = locator.nth(index)
            try:
                if not target.is_visible():
                    continue
            except Exception:
                continue
            try:
                target.click(timeout=3000)
            except Exception:
                try:
                    target.click(timeout=3000, force=True)
                except Exception:
                    continue
            return True
    return False


def fill_labshake_comment_near_time_selects(
    page,
    *,
    time_selects: list,
    comment_text: str,
) -> bool:
    comment = (comment_text or "").strip()
    if not comment or not time_selects:
        return False

    comment_selectors = [
        "textarea[name*='comment' i]",
        "textarea[placeholder*='comment' i]",
        "textarea[id*='comment' i]",
        "input[name*='comment' i]",
        "input[placeholder*='comment' i]",
        "input[id*='comment' i]",
    ]

    anchor = time_selects[0]
    for ancestor_xpath in [
        "xpath=ancestor::form[1]",
        "xpath=ancestor::tr[1]",
        "xpath=ancestor::table[1]",
        "xpath=ancestor::div[1]",
        "xpath=ancestor::td[1]",
    ]:
        try:
            container = anchor.locator(ancestor_xpath)
            if container.count() <= 0:
                continue
            scope = container.first
        except Exception:
            continue

        for selector in comment_selectors:
            try:
                locator = scope.locator(selector)
                count = min(locator.count(), 8)
            except Exception:
                count = 0
            for index in range(count - 1, -1, -1):
                target = locator.nth(index)
                try:
                    if not target.is_visible():
                        continue
                except Exception:
                    continue
                try:
                    target.fill(comment)
                except Exception:
                    try:
                        target.click(timeout=800, force=True)
                        target.fill(comment)
                    except Exception:
                        continue
                try:
                    target.dispatch_event("input")
                    target.dispatch_event("change")
                except Exception:
                    pass
                return True

    for selector in comment_selectors:
        try:
            locator = page.locator(selector)
            count = min(locator.count(), 24)
        except Exception:
            count = 0
        for index in range(count - 1, -1, -1):
            target = locator.nth(index)
            try:
                if not target.is_visible():
                    continue
            except Exception:
                continue
            try:
                target.fill(comment)
            except Exception:
                continue
            try:
                target.dispatch_event("input")
                target.dispatch_event("change")
            except Exception:
                pass
            return True

    return False


def wait_for_labshake_schedule_ready(page, *, timeout_ms: int) -> None:
    try:
        page.wait_for_load_state("networkidle", timeout=timeout_ms)
    except Exception:
        pass

    slot_hint_selectors = [
        ".available-slot",
        ".open-slot",
        "text=/click to reserve/i",
        "text=/open\\s*-\\s*click/i",
        "button:has-text('Reserve')",
    ]
    for selector in slot_hint_selectors:
        try:
            page.wait_for_selector(selector, timeout=2200)
            return
        except Exception:
            continue
    page.wait_for_timeout(600)


def extract_open_range_for_click_item(item) -> tuple[time, time] | None:
    text_candidates: list[str] = []
    try:
        raw_text = re.sub(r"\s+", " ", item.inner_text() or "").strip()
        if raw_text:
            text_candidates.append(raw_text)
    except Exception:
        pass

    for ancestor_xpath in ("xpath=ancestor::tr[1]", "xpath=ancestor::td[1]"):
        try:
            ancestor = item.locator(ancestor_xpath)
            if ancestor.count() <= 0:
                continue
            raw_text = re.sub(r"\s+", " ", ancestor.first.inner_text() or "").strip()
            if raw_text:
                text_candidates.append(raw_text)
        except Exception:
            continue

    for candidate in text_candidates:
        parsed_range = parse_time_range(candidate)
        if parsed_range:
            return parsed_range
    return None


def click_labshake_open_window(
    page,
    *,
    reserve_start: time,
    reserve_end: time,
) -> tuple[bool, str | None]:
    selectors = [
        ".available-slot",
        ".open-slot",
        "a:has-text('click to reserve')",
        "td:has-text('click to reserve')",
        "tr:has-text('click to reserve')",
        "text=/open\\s*-\\s*click/i",
    ]

    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = min(locator.count(), 200)
        except Exception:
            count = 0
        for index in range(count):
            item = locator.nth(index)
            try:
                if not item.is_visible():
                    continue
            except Exception:
                continue

            item_text = ""
            try:
                item_text = re.sub(r"\s+", " ", item.inner_text() or "").strip()
            except Exception:
                item_text = ""
            item_text_lower = item_text.lower() if item_text else ""
            if "click to reserve" not in item_text_lower and "open" not in item_text_lower:
                try:
                    row_text = re.sub(
                        r"\s+", " ", (item.locator("xpath=ancestor::tr[1]").first.inner_text() or "")
                    ).strip().lower()
                except Exception:
                    row_text = ""
                if "click to reserve" not in row_text and "open" not in row_text:
                    continue

            parsed_range = extract_open_range_for_click_item(item)
            if not parsed_range:
                continue
            open_start, open_end = parsed_range
            if reserve_start < open_start or reserve_end > open_end:
                continue

            for click_target in (
                item,
                item.locator("a").first,
                item.locator("button").first,
            ):
                try:
                    click_target.click(timeout=2500)
                    page.wait_for_timeout(500)
                    if is_labshake_reserve_form_visible(page):
                        return True, None
                except Exception:
                    continue

    return False, "В LabShake не найдено открытое окно для выбранного времени."


def fill_labshake_reserve_form(
    page,
    *,
    reserve_start: time,
    reserve_end: time,
    comment_text: str | None,
) -> tuple[bool, str | None]:
    disable_labshake_all_day_if_enabled(page)

    visible_selects = get_visible_labshake_time_selects(page)

    if len(visible_selects) < 4:
        return False, "Не удалось найти форму выбора времени в LabShake."

    time_selects = visible_selects[-4:]
    if not labshake_select_option(time_selects[0], reserve_start.hour):
        return False, "Не удалось выбрать час начала в LabShake."
    if not labshake_select_option(time_selects[1], reserve_start.minute):
        return False, "Не удалось выбрать минуты начала в LabShake."
    if not labshake_select_option(time_selects[2], reserve_end.hour):
        return False, "Не удалось выбрать час окончания в LabShake."
    if not labshake_select_option(time_selects[3], reserve_end.minute):
        return False, "Не удалось выбрать минуты окончания в LabShake."

    comment = (comment_text or "").strip()
    if comment:
        filled = fill_labshake_comment_near_time_selects(
            page,
            time_selects=time_selects,
            comment_text=comment,
        )
        if not filled:
            logger.warning("LabShake comment field not found; reservation will be sent without comment.")

    if click_labshake_reserve_submit_near_time_selects(page, time_selects):
        return True, None

    return False, "Не удалось нажать кнопку Reserve в LabShake."


def row_matches_labshake_busy_interval(
    row_text: str,
    *,
    reserve_start: time,
    reserve_end: time,
    allow_overlap: bool = False,
) -> bool:
    normalized = re.sub(r"\s+", " ", row_text or "").strip()
    if not normalized:
        return False
    lower = normalized.lower()
    if "click to reserve" in lower or "open - click" in lower:
        return False
    parsed_range = parse_time_range(normalized)
    if not parsed_range:
        return False
    row_start, row_end = parsed_range
    if row_start == reserve_start and row_end == reserve_end:
        return True
    if not allow_overlap:
        return False

    probe_day = date(2000, 1, 1)
    req_start_dt, req_end_dt = interval_from_date_and_times(
        probe_day, reserve_start, reserve_end
    )
    row_start_dt, row_end_dt = interval_from_date_and_times(
        probe_day, row_start, row_end
    )
    return intervals_overlap(req_start_dt, req_end_dt, row_start_dt, row_end_dt)


def confirm_labshake_modal_if_present(page) -> bool:
    positive_tokens = ("ok", "yes", "confirm", "delete", "удал", "подтверд", "да")
    negative_tokens = ("cancel", "close", "no", "нет", "отмена", "закрыть")
    selectors = [
        ".modal.show button, .modal.in button, [role='dialog'] button, .bootbox button, .swal2-container button",
        ".modal.show a, .modal.in a, [role='dialog'] a, .bootbox a, .swal2-container a",
    ]
    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = min(locator.count(), 32)
        except Exception:
            count = 0
        for index in range(count):
            btn = locator.nth(index)
            try:
                if not btn.is_visible():
                    continue
            except Exception:
                continue
            try:
                text = re.sub(r"\s+", " ", btn.inner_text() or "").strip().lower()
            except Exception:
                text = ""
            if not text:
                try:
                    text = " ".join(
                        [
                            str(btn.get_attribute("value") or ""),
                            str(btn.get_attribute("title") or ""),
                            str(btn.get_attribute("aria-label") or ""),
                        ]
                    ).strip().lower()
                except Exception:
                    text = ""

            if any(token in text for token in negative_tokens):
                continue
            if not any(token in text for token in positive_tokens):
                continue

            try:
                btn.click(timeout=2500)
            except Exception:
                try:
                    btn.click(timeout=2500, force=True)
                except Exception:
                    continue
            page.wait_for_timeout(250)
            return True
    return False


def click_visible_labshake_cancel_action(page, *, row=None, control=None) -> bool:
    selectors: list[tuple[object, str]] = []
    if row is not None:
        selectors.extend(
            [
                (row, ".dropdown-menu a:has-text('Cancel')"),
                (row, ".dropdown-menu button:has-text('Cancel')"),
                (row, "a:has-text('Cancel')"),
                (row, "button:has-text('Cancel')"),
            ]
        )
    if control is not None:
        try:
            control_id = str(control.get_attribute("id") or "").strip()
        except Exception:
            control_id = ""
        if control_id:
            selectors.extend(
                [
                    (page, f".dropdown-menu[aria-labelledby='{control_id}'] a:has-text('Cancel')"),
                    (page, f".dropdown-menu[aria-labelledby='{control_id}'] button:has-text('Cancel')"),
                ]
            )
    selectors.extend(
        [
            (page, ".dropdown-menu a:has-text('Cancel')"),
            (page, ".dropdown-menu button:has-text('Cancel')"),
            (page, "a:has-text('Cancel')"),
            (page, "button:has-text('Cancel')"),
            (page, "text=/^\\s*cancel\\s*$/i"),
        ]
    )

    for scope, selector in selectors:
        try:
            locator = scope.locator(selector)
            count = min(locator.count(), 24)
        except Exception:
            count = 0
        for index in range(count):
            target = locator.nth(index)
            try:
                if not target.is_visible():
                    continue
            except Exception:
                continue
            try:
                raw_text = re.sub(r"\s+", " ", target.inner_text() or "").strip().lower()
                if raw_text and "cancel" not in raw_text:
                    continue
            except Exception:
                pass
            try:
                page.once("dialog", lambda dialog: dialog.accept())
            except Exception:
                pass
            try:
                target.click(timeout=3000)
            except Exception:
                try:
                    target.click(timeout=3000, force=True)
                except Exception:
                    continue
            page.wait_for_timeout(180)
            confirm_labshake_modal_if_present(page)
            return True
    return False


def open_labshake_row_actions_menu(page, row) -> bool:
    selectors = [
        "a[data-toggle='dropdown']",
        "button[data-toggle='dropdown']",
        "a.dropdown-toggle",
        "button.dropdown-toggle",
        "a:has(i.fa-pencil)",
        "a:has(i.fa-edit)",
        "a:has(i[class*='pencil'])",
        "a:has(i[class*='edit'])",
        "button:has(i[class*='pencil'])",
        "button:has(i[class*='edit'])",
        "a:has-text('Edit')",
        "button:has-text('Edit')",
        "[aria-label*='edit' i]",
        "[title*='edit' i]",
    ]
    for selector in selectors:
        try:
            controls = row.locator(selector)
            count = min(controls.count(), 16)
        except Exception:
            count = 0
        for idx in range(count):
            control = controls.nth(idx)
            try:
                if not control.is_visible():
                    continue
            except Exception:
                continue
            try:
                control.click(timeout=2200)
            except Exception:
                try:
                    control.click(timeout=2200, force=True)
                except Exception:
                    continue
            page.wait_for_timeout(220)
            if click_visible_labshake_cancel_action(page, row=row, control=control):
                return True

    try:
        clicked = bool(
            row.evaluate(
                """
                (el) => {
                  const toText = (v) => (v || '').toString().toLowerCase();
                  const nodes = Array.from(
                    el.querySelectorAll('a,button,[role="button"],i,span,svg')
                  );
                  const score = (node) => {
                    const cls = toText(node.className);
                    const title = toText(node.getAttribute && node.getAttribute('title'));
                    const aria = toText(node.getAttribute && node.getAttribute('aria-label'));
                    const txt = toText(node.textContent);
                    const meta = [cls, title, aria, txt].join(' ');
                    if (/(cancel|delete|remove)/.test(meta)) return -1000;
                    if (/(pencil|edit|dropdown|toggle|menu)/.test(meta)) return 10;
                    if (node.tagName === 'A' || node.tagName === 'BUTTON') return 3;
                    return 0;
                  };
                  nodes.sort((a, b) => score(b) - score(a));
                  for (const node of nodes) {
                    if (score(node) <= 0) continue;
                    try {
                      node.click();
                      return true;
                    } catch (e) {}
                  }
                  const firstCell = el.querySelector('td,th,div');
                  if (firstCell) {
                    try {
                      firstCell.click();
                      return true;
                    } catch (e) {}
                  }
                  return false;
                }
                """
            )
        )
    except Exception:
        clicked = False
    if clicked:
        page.wait_for_timeout(280)
        if click_visible_labshake_cancel_action(page, row=row):
            return True

    return False


def force_click_cancel_for_row_via_js(page, row) -> bool:
    try:
        clicked = bool(
            row.evaluate(
                """
                (row) => {
                  const norm = (v) => (v || '').toString().replace(/\\s+/g, ' ').trim().toLowerCase();
                  const isCancelNode = (node) => {
                    if (!node) return false;
                    const txt = norm(node.textContent);
                    const cls = norm(node.className);
                    const title = norm(node.getAttribute && node.getAttribute('title'));
                    const aria = norm(node.getAttribute && node.getAttribute('aria-label'));
                    const meta = [txt, cls, title, aria].join(' ');
                    return /(cancel|delete|remove|fa-trash)/.test(meta);
                  };
                  const isToggleNode = (node) => {
                    if (!node) return false;
                    const txt = norm(node.textContent);
                    const cls = norm(node.className);
                    const title = norm(node.getAttribute && node.getAttribute('title'));
                    const aria = norm(node.getAttribute && node.getAttribute('aria-label'));
                    const dataToggle = norm(node.getAttribute && node.getAttribute('data-toggle'));
                    const meta = [txt, cls, title, aria, dataToggle].join(' ');
                    return /(dropdown|toggle|edit|pencil|fa-pencil|fa-edit)/.test(meta);
                  };

                  const candidates = Array.from(
                    row.querySelectorAll('a,button,[role="button"],i,span,svg')
                  );
                  for (const node of candidates) {
                    if (!isToggleNode(node)) continue;
                    try { node.click(); } catch (e) {}
                  }

                  const menus = new Set();
                  const inRowMenus = row.querySelectorAll('.dropdown-menu, [role="menu"], ul, div');
                  inRowMenus.forEach((m) => menus.add(m));

                  const controls = row.querySelectorAll('a,button,[role="button"]');
                  controls.forEach((control) => {
                    const id = control.getAttribute && control.getAttribute('id');
                    if (id) {
                      document
                        .querySelectorAll(`.dropdown-menu[aria-labelledby="${id}"], [role="menu"][aria-labelledby="${id}"]`)
                        .forEach((m) => menus.add(m));
                    }
                    let sib = control.nextElementSibling;
                    while (sib) {
                      menus.add(sib);
                      sib = sib.nextElementSibling;
                    }
                    const parent = control.parentElement;
                    if (parent) {
                      parent.querySelectorAll('.dropdown-menu, [role="menu"], ul').forEach((m) => menus.add(m));
                    }
                  });

                  for (const menu of menus) {
                    if (!menu) continue;
                    const items = menu.querySelectorAll('a,button,[role="menuitem"],li');
                    for (const item of items) {
                      if (!isCancelNode(item)) continue;
                      try {
                        item.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
                        return true;
                      } catch (e) {}
                      try {
                        item.click();
                        return true;
                      } catch (e) {}
                    }
                  }

                  for (const item of row.querySelectorAll('a,button,[role="menuitem"],li,span')) {
                    if (!isCancelNode(item)) continue;
                    try {
                      item.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
                      return true;
                    } catch (e) {}
                    try {
                      item.click();
                      return true;
                    } catch (e) {}
                  }
                  return false;
                }
                """
            )
        )
    except Exception:
        return False
    return clicked


def click_labshake_cancel_window(
    page,
    *,
    reserve_start: time,
    reserve_end: time,
) -> tuple[bool, str | None]:
    target_found = False

    for overlap_mode in (False, True):
        toggle_selectors = [
            "tr a:has(i.fa-pencil)",
            "tr button:has(i.fa-pencil)",
            "tr [data-toggle='dropdown']",
            "tr a.dropdown-toggle",
            "tr button.dropdown-toggle",
        ]
        for selector in toggle_selectors:
            try:
                locator = page.locator(selector)
                count = min(locator.count(), 220)
            except Exception:
                count = 0
            for index in range(count):
                toggle = locator.nth(index)
                try:
                    if not toggle.is_visible():
                        continue
                except Exception:
                    continue

                row_text = ""
                try:
                    row = toggle.locator("xpath=ancestor::tr[1]")
                    if row.count() <= 0:
                        continue
                    row_text = re.sub(r"\s+", " ", row.first.inner_text() or "").strip()
                except Exception:
                    continue

                if not row_matches_labshake_busy_interval(
                    row_text,
                    reserve_start=reserve_start,
                    reserve_end=reserve_end,
                    allow_overlap=overlap_mode,
                ):
                    continue

                target_found = True
                try:
                    toggle.click(timeout=2500)
                except Exception:
                    try:
                        toggle.click(timeout=2500, force=True)
                    except Exception:
                        continue
                page.wait_for_timeout(350)
                if click_visible_labshake_cancel_action(page, row=row, control=toggle):
                    return True, None
                try:
                    page.once("dialog", lambda dialog: dialog.accept())
                except Exception:
                    pass
                if force_click_cancel_for_row_via_js(page, row):
                    page.wait_for_timeout(350)
                    return True, None

        try:
            rows = page.locator("tr, .reservation-slot, .busy-slot")
            row_count = min(rows.count(), 320)
        except Exception:
            row_count = 0
        for index in range(row_count):
            row = rows.nth(index)
            row_text = ""
            try:
                row_text = re.sub(r"\s+", " ", row.inner_text() or "").strip()
            except Exception:
                continue
            if not row_matches_labshake_busy_interval(
                row_text,
                reserve_start=reserve_start,
                reserve_end=reserve_end,
                allow_overlap=overlap_mode,
            ):
                continue

            target_found = True
            if open_labshake_row_actions_menu(page, row):
                return True, None
            try:
                page.once("dialog", lambda dialog: dialog.accept())
            except Exception:
                pass
            if force_click_cancel_for_row_via_js(page, row):
                page.wait_for_timeout(320)
                return True, None

            try:
                row.click(timeout=1500)
            except Exception:
                pass
            page.wait_for_timeout(250)
            if click_visible_labshake_cancel_action(page):
                return True, None

    if not target_found:
        return False, "В LabShake не найдена бронь для отмены на выбранное время."
    return False, "Не удалось нажать кнопку Cancel в LabShake."


def is_buffered_interval_available_in_labshake(
    *,
    slot_date: date,
    reserve_start_dt: datetime,
    reserve_end_dt: datetime,
    open_windows: list[dict],
    busy_intervals: list[dict],
) -> tuple[bool, str | None]:
    open_ranges: list[tuple[datetime, datetime]] = []
    for item in open_windows:
        if item["slot_date"] != slot_date:
            continue
        start_dt, end_dt = interval_from_date_and_times(
            slot_date,
            item["start_time"],
            item["end_time"],
        )
        open_ranges.append((start_dt, end_dt))

    if not open_ranges:
        return False, "В LabShake нет открытых окон на выбранный день."

    for item in busy_intervals:
        if item["slot_date"] != slot_date:
            continue
        busy_start, busy_end = interval_from_date_and_times(
            slot_date,
            item["start_time"],
            item["end_time"],
        )
        if intervals_overlap(reserve_start_dt, reserve_end_dt, busy_start, busy_end):
            return False, "Слот уже заняли в LabShake. Выберите другое время."

    covered = any(
        reserve_start_dt >= open_start and reserve_end_dt <= open_end
        for open_start, open_end in open_ranges
    )
    if not covered:
        return (
            False,
            "В LabShake выбранное окно уже недоступно с учетом буфера. Выберите другой слот.",
        )

    return True, None


def reserve_buffered_interval_in_labshake(
    *,
    exp: ExperimentConfig,
    slot_date: date,
    selected_start: time,
    selected_end: time,
) -> tuple[bool, str]:
    if not exp.labshake_schedule_url:
        return True, "ok"

    buffer_minutes = max(0, int(exp.slot_step_minutes))
    selected_start_dt, selected_end_dt = interval_from_date_and_times(
        slot_date,
        selected_start,
        selected_end,
    )
    reserve_start_dt = selected_start_dt - timedelta(minutes=buffer_minutes)
    reserve_end_dt = selected_end_dt + timedelta(minutes=buffer_minutes)

    if reserve_start_dt.date() != slot_date or reserve_end_dt.date() != slot_date:
        return (
            False,
            "Невозможно забронировать LabShake: буфер выводит интервал за пределы выбранного дня.",
        )

    open_windows, busy_intervals, fetch_error = fetch_labshake_schedule_for_day(
        schedule_url=exp.labshake_schedule_url,
        cookie_env=exp.labshake_cookie_env,
        target_day=slot_date,
    )
    if fetch_error:
        return False, fetch_error

    available, availability_error = is_buffered_interval_available_in_labshake(
        slot_date=slot_date,
        reserve_start_dt=reserve_start_dt,
        reserve_end_dt=reserve_end_dt,
        open_windows=open_windows,
        busy_intervals=busy_intervals,
    )
    if not available:
        return False, availability_error or "Слот недоступен в LabShake."

    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return (
            False,
            "Не установлен Playwright. Выполните: python -m playwright install chromium",
        )

    timeout_raw = os.getenv("LABSHAKE_LOGIN_TIMEOUT_SEC", "45").strip()
    try:
        timeout_ms = max(10_000, int(float(timeout_raw.replace(",", ".")) * 1000))
    except ValueError:
        timeout_ms = 45_000
    headless = parse_env_bool("LABSHAKE_HEADLESS", default=True)
    browser_channel = os.getenv("LABSHAKE_BROWSER_CHANNEL", "chrome").strip().lower()
    if browser_channel == "default":
        browser_channel = ""

    day_url = build_labshake_day_url(
        normalize_labshake_daily_url(exp.labshake_schedule_url),
        slot_date,
    )

    browser = None
    context = None
    try:
        with sync_playwright() as playwright:
            launch_kwargs: dict[str, object] = {"headless": headless}
            if browser_channel:
                launch_kwargs["channel"] = browser_channel
            try:
                browser = playwright.chromium.launch(**launch_kwargs)
            except Exception:
                if "channel" in launch_kwargs:
                    launch_kwargs.pop("channel", None)
                    browser = playwright.chromium.launch(**launch_kwargs)
                else:
                    raise

            context = browser.new_context()
            env_cookie = os.getenv(exp.labshake_cookie_env, "").strip()
            if env_cookie:
                apply_cookie_header_to_browser_context(context, env_cookie)
            page = context.new_page()
            page.goto(day_url, wait_until="domcontentloaded", timeout=timeout_ms)
            wait_for_labshake_schedule_ready(page, timeout_ms=timeout_ms)

            if is_labshake_login_page(page):
                login_ok, login_error = labshake_login_in_browser(
                    page, timeout_ms=timeout_ms
                )
                if not login_ok:
                    return False, login_error or "Не удалось войти в LabShake."
                page.goto(day_url, wait_until="domcontentloaded", timeout=timeout_ms)
                wait_for_labshake_schedule_ready(page, timeout_ms=timeout_ms)
                if is_labshake_login_page(page):
                    return False, "После входа LabShake все еще запрашивает авторизацию."

            if not is_labshake_reserve_form_visible(page):
                clicked, click_error = click_labshake_open_window(
                    page,
                    reserve_start=reserve_start_dt.time(),
                    reserve_end=reserve_end_dt.time(),
                )
                if not clicked and not is_labshake_reserve_form_visible(page):
                    return False, (
                        click_error
                        or (
                            "Не удалось открыть форму резервирования LabShake "
                            f"для интервала {reserve_start_dt.strftime('%H:%M')}-"
                            f"{reserve_end_dt.strftime('%H:%M')}."
                        )
                    )

            form_ok, form_error = fill_labshake_reserve_form(
                page,
                reserve_start=reserve_start_dt.time(),
                reserve_end=reserve_end_dt.time(),
                comment_text=exp.labshake_booking_comment,
            )
            if (
                not form_ok
                and form_error
                and "форму выбора времени" in form_error.lower()
            ):
                reclicked, _ = click_labshake_open_window(
                    page,
                    reserve_start=reserve_start_dt.time(),
                    reserve_end=reserve_end_dt.time(),
                )
                if reclicked:
                    page.wait_for_timeout(450)
                    form_ok, form_error = fill_labshake_reserve_form(
                        page,
                        reserve_start=reserve_start_dt.time(),
                        reserve_end=reserve_end_dt.time(),
                        comment_text=exp.labshake_booking_comment,
                    )
            if not form_ok:
                return False, form_error or "Не удалось заполнить форму LabShake."

            try:
                page.wait_for_load_state("networkidle", timeout=timeout_ms)
            except Exception:
                pass
            page.wait_for_timeout(1200)

            html_after = page.content()
            _, busy_after, parse_error = parse_labshake_schedule(
                html_after,
                days_ahead=days_ahead_for_target_date(slot_date),
            )
            if parse_error:
                return False, (
                    "Не удалось подтвердить бронь в LabShake после нажатия Reserve. "
                    "Повторите попытку."
                )

            matched_interval = pick_best_busy_interval_for_request(
                slot_date=slot_date,
                reserve_start_dt=reserve_start_dt,
                reserve_end_dt=reserve_end_dt,
                busy_intervals=busy_after,
            )
            if not matched_interval:
                return (
                    False,
                    "LabShake не подтвердил бронь выбранного окна. Попробуйте еще раз.",
                )
            matched_start, matched_end = matched_interval

            boundary_drift_sec = max(
                abs((matched_start - reserve_start_dt).total_seconds()),
                abs((matched_end - reserve_end_dt).total_seconds()),
            )
            if boundary_drift_sec > 15 * 60:
                rollback_ok, rollback_error = click_labshake_cancel_window(
                    page,
                    reserve_start=matched_start.time(),
                    reserve_end=matched_end.time(),
                )
                if rollback_ok:
                    return (
                        False,
                        "LabShake создал запись на неверный интервал "
                        f"{matched_start.strftime('%H:%M')}-{matched_end.strftime('%H:%M')}. "
                        "Запись автоматически отменена. Попробуйте снова.",
                    )
                return (
                    False,
                    "LabShake создал запись на неверный интервал "
                    f"{matched_start.strftime('%H:%M')}-{matched_end.strftime('%H:%M')}. "
                    f"Не удалось автоматически отменить: {rollback_error}. Проверьте LabShake вручную.",
                )

            cookies = context.cookies("https://labshake.com")
            cookie_header = build_cookie_header_from_browser_cookies(cookies)
            if cookie_header:
                os.environ[exp.labshake_cookie_env] = cookie_header
                cache_key = f"{exp.labshake_cookie_env}::labshake"
                with labshake_auth_lock:
                    labshake_cookie_cache[cache_key] = cookie_header
    except Exception as exc:
        return False, f"Ошибка резервирования LabShake: {exc}"
    finally:
        if context is not None:
            try:
                context.close()
            except Exception:
                pass
        if browser is not None:
            try:
                browser.close()
            except Exception:
                pass

    return True, "ok"


def cancel_buffered_interval_in_labshake(
    *,
    exp: ExperimentConfig,
    slot_date: date,
    selected_start: time,
    selected_end: time,
) -> tuple[bool, str]:
    if not exp.labshake_schedule_url:
        return True, "ok"

    buffer_minutes = max(0, int(exp.slot_step_minutes))
    selected_start_dt, selected_end_dt = interval_from_date_and_times(
        slot_date,
        selected_start,
        selected_end,
    )
    reserve_start_dt = selected_start_dt - timedelta(minutes=buffer_minutes)
    reserve_end_dt = selected_end_dt + timedelta(minutes=buffer_minutes)

    if reserve_start_dt.date() != slot_date or reserve_end_dt.date() != slot_date:
        return (
            False,
            "Невозможно отменить бронь LabShake: буфер выводит интервал за пределы выбранного дня.",
        )

    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return (
            False,
            "Не установлен Playwright. Выполните: python -m playwright install chromium",
        )

    timeout_raw = os.getenv("LABSHAKE_LOGIN_TIMEOUT_SEC", "45").strip()
    try:
        timeout_ms = max(10_000, int(float(timeout_raw.replace(",", ".")) * 1000))
    except ValueError:
        timeout_ms = 45_000
    headless = parse_env_bool("LABSHAKE_HEADLESS", default=True)
    browser_channel = os.getenv("LABSHAKE_BROWSER_CHANNEL", "chrome").strip().lower()
    if browser_channel == "default":
        browser_channel = ""

    day_url = build_labshake_day_url(
        normalize_labshake_daily_url(exp.labshake_schedule_url),
        slot_date,
    )

    browser = None
    context = None
    try:
        with sync_playwright() as playwright:
            launch_kwargs: dict[str, object] = {"headless": headless}
            if browser_channel:
                launch_kwargs["channel"] = browser_channel
            try:
                browser = playwright.chromium.launch(**launch_kwargs)
            except Exception:
                if "channel" in launch_kwargs:
                    launch_kwargs.pop("channel", None)
                    browser = playwright.chromium.launch(**launch_kwargs)
                else:
                    raise

            context = browser.new_context()
            env_cookie = os.getenv(exp.labshake_cookie_env, "").strip()
            if env_cookie:
                apply_cookie_header_to_browser_context(context, env_cookie)
            page = context.new_page()
            page.goto(day_url, wait_until="domcontentloaded", timeout=timeout_ms)
            wait_for_labshake_schedule_ready(page, timeout_ms=timeout_ms)

            if is_labshake_login_page(page):
                login_ok, login_error = labshake_login_in_browser(
                    page, timeout_ms=timeout_ms
                )
                if not login_ok:
                    return False, login_error or "Не удалось войти в LabShake."
                page.goto(day_url, wait_until="domcontentloaded", timeout=timeout_ms)
                wait_for_labshake_schedule_ready(page, timeout_ms=timeout_ms)
                if is_labshake_login_page(page):
                    return False, "После входа LabShake все еще запрашивает авторизацию."

            canceled, cancel_error = click_labshake_cancel_window(
                page,
                reserve_start=reserve_start_dt.time(),
                reserve_end=reserve_end_dt.time(),
            )
            if not canceled:
                html_current = page.content()
                _, busy_current, parse_current_error = parse_labshake_schedule(
                    html_current,
                    days_ahead=days_ahead_for_target_date(slot_date),
                )
                still_busy_dom = page_has_overlapping_labshake_reservation(
                    page,
                    reserve_start=reserve_start_dt.time(),
                    reserve_end=reserve_end_dt.time(),
                )
                if not parse_current_error:
                    still_busy = has_any_overlapping_busy_interval(
                        slot_date=slot_date,
                        reserve_start_dt=reserve_start_dt,
                        reserve_end_dt=reserve_end_dt,
                        busy_intervals=busy_current,
                    )
                    if not still_busy and not still_busy_dom:
                        canceled = True
                if not canceled:
                    return False, cancel_error or "Не удалось отменить бронь в LabShake."

            confirmed_canceled = False
            parse_error_seen: str | None = None
            for attempt in range(1, 5):
                if attempt > 1:
                    try:
                        page.reload(wait_until="domcontentloaded", timeout=timeout_ms)
                    except Exception:
                        pass
                    wait_for_labshake_schedule_ready(page, timeout_ms=timeout_ms)
                    page.wait_for_timeout(450 + attempt * 200)

                try:
                    page.wait_for_load_state("networkidle", timeout=timeout_ms)
                except Exception:
                    pass
                page.wait_for_timeout(600)

                html_after = page.content()
                _, busy_after, parse_error = parse_labshake_schedule(
                    html_after,
                    days_ahead=days_ahead_for_target_date(slot_date),
                )
                if parse_error:
                    parse_error_seen = parse_error
                    continue

                still_busy = has_any_overlapping_busy_interval(
                    slot_date=slot_date,
                    reserve_start_dt=reserve_start_dt,
                    reserve_end_dt=reserve_end_dt,
                    busy_intervals=busy_after,
                )
                still_busy_dom = page_has_overlapping_labshake_reservation(
                    page,
                    reserve_start=reserve_start_dt.time(),
                    reserve_end=reserve_end_dt.time(),
                )
                if not still_busy and not still_busy_dom:
                    confirmed_canceled = True
                    break

            if not confirmed_canceled:
                second_try_ok, _ = click_labshake_cancel_window(
                    page,
                    reserve_start=reserve_start_dt.time(),
                    reserve_end=reserve_end_dt.time(),
                )
                if second_try_ok:
                    for attempt in range(1, 4):
                        if attempt > 1:
                            try:
                                page.reload(wait_until="domcontentloaded", timeout=timeout_ms)
                            except Exception:
                                pass
                            wait_for_labshake_schedule_ready(page, timeout_ms=timeout_ms)
                            page.wait_for_timeout(350 + attempt * 180)

                        try:
                            page.wait_for_load_state("networkidle", timeout=timeout_ms)
                        except Exception:
                            pass
                        page.wait_for_timeout(420)

                        html_after_retry = page.content()
                        _, busy_after_retry, parse_retry_error = parse_labshake_schedule(
                            html_after_retry,
                            days_ahead=days_ahead_for_target_date(slot_date),
                        )
                        if parse_retry_error:
                            continue
                        still_busy_retry = has_any_overlapping_busy_interval(
                            slot_date=slot_date,
                            reserve_start_dt=reserve_start_dt,
                            reserve_end_dt=reserve_end_dt,
                            busy_intervals=busy_after_retry,
                        )
                        still_busy_retry_dom = page_has_overlapping_labshake_reservation(
                            page,
                            reserve_start=reserve_start_dt.time(),
                            reserve_end=reserve_end_dt.time(),
                        )
                        if not still_busy_retry and not still_busy_retry_dom:
                            confirmed_canceled = True
                            break

            if not confirmed_canceled:
                if parse_error_seen:
                    return False, (
                        "Не удалось подтвердить отмену в LabShake после нажатия Cancel. "
                        f"Причина: {parse_error_seen}"
                    )
                return False, "LabShake не подтвердил отмену брони. Проверьте расписание вручную."

            cookies = context.cookies("https://labshake.com")
            cookie_header = build_cookie_header_from_browser_cookies(cookies)
            if cookie_header:
                os.environ[exp.labshake_cookie_env] = cookie_header
                cache_key = f"{exp.labshake_cookie_env}::labshake"
                with labshake_auth_lock:
                    labshake_cookie_cache[cache_key] = cookie_header
    except Exception as exc:
        return False, f"Ошибка отмены брони LabShake: {exc}"
    finally:
        if context is not None:
            try:
                context.close()
            except Exception:
                pass
        if browser is not None:
            try:
                browser.close()
            except Exception:
                pass

    return True, "ok"


def sync_day_windows_from_labshake(
    exp: ExperimentConfig,
) -> tuple[bool, str, list[dict], list[str], list[str]]:
    if exp.slot_mode != "day_windows":
        return (
            False,
            "Синхронизация LabShake сейчас поддерживается только для slot_mode=day_windows.",
            [],
            [],
            [],
        )
    if not exp.labshake_schedule_url:
        return False, "В experiments.json не указан labshake_schedule_url.", [], [], []

    open_windows, busy_intervals, source_days, failed_days, fetch_error = fetch_labshake_schedule_range(
        schedule_url=exp.labshake_schedule_url,
        cookie_env=exp.labshake_cookie_env,
        days_ahead=exp.available_days_ahead,
    )
    if fetch_error:
        return False, fetch_error, [], [], []
    if not source_days:
        return False, "LabShake не вернул дней для синхронизации.", [], [], []

    busy_by_day: dict[date, list[tuple[datetime, datetime]]] = {}
    for item in busy_intervals:
        start_dt, end_dt = interval_from_date_and_times(
            item["slot_date"], item["start_time"], item["end_time"]
        )
        busy_by_day.setdefault(item["slot_date"], []).append((start_dt, end_dt))
    for values in busy_by_day.values():
        values.sort(key=lambda x: (x[0], x[1]))

    canceled: list[dict] = []
    removed_mismatch_labels: list[str] = []
    removed_slot_labels: list[str] = []

    with excel_lock:
        wb, load_error = load_workbook_from_storage(exp.storage)
        if load_error:
            return False, load_error, [], [], []
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            before_generated: dict[str, dict] = {}
            if exp.slot_duration_hours and exp.slot_duration_hours > 0:
                slots_before = build_generated_available_slots(
                    ws,
                    cols,
                    working_hours=exp.working_hours,
                    excluded_weekdays=exp.excluded_weekdays,
                    slot_duration_hours=exp.slot_duration_hours,
                    min_gap_hours=exp.min_gap_hours,
                    slot_step_minutes=exp.slot_step_minutes,
                    max_weekly_hours=exp.max_weekly_hours,
                    default_slot_duration_hours=exp.default_slot_duration_hours,
                    days_ahead=exp.available_days_ahead,
                )
                before_generated = {
                    slot["key"]: slot
                    for slot in slots_before
                    if slot["slot_date"] in source_days
                }

            rows_to_delete: list[int] = []
            removed_windows = 0
            removed_bookings = 0

            for row in range(2, ws.max_row + 1):
                raw_date = ws.cell(row=row, column=cols.date).value
                raw_time = ws.cell(row=row, column=cols.time).value
                if is_empty(raw_date) or is_empty(raw_time):
                    continue

                slot_date = parse_date_cell(raw_date)
                if not slot_date:
                    continue
                if slot_date not in source_days:
                    continue

                parsed_range = parse_time_range(
                    raw_time if isinstance(raw_time, str) else str(raw_time)
                )
                if not parsed_range:
                    continue

                tg_value = ws.cell(row=row, column=cols.telegram).value
                tg_text = str(tg_value).strip() if not is_empty(tg_value) else ""

                if not tg_text:
                    rows_to_delete.append(row)
                    removed_windows += 1
                    continue

                if tg_text == LABSHAKE_BUSY_MARKER:
                    rows_to_delete.append(row)
                    continue

                # Participant booking rows are preserved during LabShake sync.
                # Cancellations should happen only via explicit researcher action.
                continue

            for row in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row, 1)

            added_windows = 0
            for window in open_windows:
                row = ws.max_row + 1
                ws.cell(row=row, column=cols.date).value = window["slot_date"].strftime(
                    "%d.%m.%Y"
                )
                ws.cell(row=row, column=cols.time).value = format_time_range(
                    window["start_time"], window["end_time"]
                )
                clear_booking_row(ws, row, cols)
                added_windows += 1

            added_busy = 0
            for interval in busy_intervals:
                row = ws.max_row + 1
                ws.cell(row=row, column=cols.date).value = interval["slot_date"].strftime(
                    "%d.%m.%Y"
                )
                ws.cell(row=row, column=cols.time).value = format_time_range(
                    interval["start_time"], interval["end_time"]
                )
                ws.cell(row=row, column=cols.telegram).value = LABSHAKE_BUSY_MARKER
                ws.cell(row=row, column=cols.full_name).value = "Занято в LabShake"
                ws.cell(row=row, column=cols.phone).value = None
                if cols.booked_at:
                    ws.cell(row=row, column=cols.booked_at).value = None
                set_booking_row_style(ws, row, cols, booked=True)
                added_busy += 1

            after_generated: dict[str, dict] = {}
            if exp.slot_duration_hours and exp.slot_duration_hours > 0:
                slots_after = build_generated_available_slots(
                    ws,
                    cols,
                    working_hours=exp.working_hours,
                    excluded_weekdays=exp.excluded_weekdays,
                    slot_duration_hours=exp.slot_duration_hours,
                    min_gap_hours=exp.min_gap_hours,
                    slot_step_minutes=exp.slot_step_minutes,
                    max_weekly_hours=exp.max_weekly_hours,
                    default_slot_duration_hours=exp.default_slot_duration_hours,
                    days_ahead=exp.available_days_ahead,
                )
                after_generated = {
                    slot["key"]: slot
                    for slot in slots_after
                    if slot["slot_date"] in source_days
                }

            removed_keys = set(before_generated.keys()) - set(after_generated.keys())
            for key in sorted(
                removed_keys,
                key=lambda item: (
                    before_generated[item]["slot_date"],
                    before_generated[item]["slot_time"],
                ),
            ):
                slot = before_generated[key]
                slot_label = slot["label"]
                removed_slot_labels.append(slot_label)
                for busy_start, busy_end in busy_by_day.get(slot["slot_date"], []):
                    if intervals_overlap(slot["start_dt"], slot["end_dt"], busy_start, busy_end):
                        removed_mismatch_labels.append(slot_label)
                        break

            sort_slots_sheet_rows(ws, cols)
            success, message = save_workbook_to_storage(exp.storage, wb)
            if not success:
                return False, message, [], [], []
        finally:
            wb.close()

    removed_slot_labels = list(dict.fromkeys(removed_slot_labels))
    removed_mismatch_labels = list(dict.fromkeys(removed_mismatch_labels))

    summary = (
        "Синхронизация LabShake завершена.\n"
        f"Дней синхронизировано: {len(source_days)}\n"
        f"Не удалось загрузить дней: {len(failed_days)}\n"
        f"Открытых окон в LabShake: {len(open_windows)}\n"
        f"Занятых интервалов из LabShake: {len(busy_intervals)}\n"
        f"Добавлено окон: {added_windows}\n"
        f"Удалено устаревших окон: {removed_windows}\n"
        f"Отменено записей: {removed_bookings}\n"
        f"Удалено доступных слотов: {len(removed_slot_labels)}\n"
        f"Из них по несовмещению времени: {len(removed_mismatch_labels)}"
    )
    return True, summary, canceled, removed_slot_labels, removed_mismatch_labels


def sync_day_windows_from_labshake_with_retry(
    exp: ExperimentConfig,
    *,
    attempts: int = 2,
    delay_sec: float = 0.9,
) -> tuple[bool, str, list[dict], list[str], list[str]]:
    total = max(1, int(attempts))
    last_message = "Ошибка синхронизации LabShake."
    last_canceled: list[dict] = []
    last_removed: list[str] = []
    last_mismatch: list[str] = []
    for attempt in range(1, total + 1):
        ok, message, canceled, removed, mismatch = sync_day_windows_from_labshake(exp)
        if ok:
            return True, message, canceled, removed, mismatch

        last_message = message
        last_canceled = canceled
        last_removed = removed
        last_mismatch = mismatch

        if attempt < total:
            logger.warning(
                "sync_day_windows_from_labshake failed (attempt %s/%s) for %s: %s",
                attempt,
                total,
                exp.experiment_id,
                message,
            )
            time_module.sleep(max(0.0, float(delay_sec)))

    return False, last_message, last_canceled, last_removed, last_mismatch


def delete_slot_row_for_admin(
    storage: ExcelStorage,
    *,
    row: int,
) -> tuple[bool, str, str | None, str | None, int | None]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return False, load_error, None, None, None
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)

            if row < 2 or row > ws.max_row:
                return False, "Слот не найден.", None, None, None

            date_value = ws.cell(row=row, column=cols.date).value
            time_value = ws.cell(row=row, column=cols.time).value
            if is_empty(date_value) or is_empty(time_value):
                return False, "Слот невалидный.", None, None, None

            slot_date = parse_date_cell(date_value)
            slot_label = (
                f"{format_date_display(date_value, slot_date)} "
                f"{format_time_display(time_value, parse_time_cell(time_value))}"
            )

            tg_value = ws.cell(row=row, column=cols.telegram).value
            tg_text = str(tg_value).strip() if not is_empty(tg_value) else None
            chat_id = parse_chat_id_from_telegram_cell(tg_text or "")

            ws.delete_rows(row, 1)
            sort_slots_sheet_rows(ws, cols)
            success, message = save_workbook_to_storage(storage, wb)
            if not success:
                return False, message, None, None, None

            return True, "ok", slot_label, tg_text, chat_id
        finally:
            wb.close()


def move_user_booking(
    storage: ExcelStorage,
    user_handles: list[str],
    new_row: int,
    max_weekly_hours: float | None = None,
    default_slot_duration_hours: float | None = None,
) -> tuple[bool, str, str | None]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return False, load_error, None
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            current = find_user_booking_in_sheet(ws, user_handles, cols)
            if not current:
                return False, "У вас нет активной записи.", None

            if new_row == current["row"]:
                return False, "Вы уже записаны на этот слот.", None

            if new_row < 2 or new_row > ws.max_row:
                return False, "Слот не найден.", None

            new_slot = read_slot_info(ws, new_row, cols)
            if not new_slot:
                return False, "Новый слот невалидный.", None

            if not is_empty(ws.cell(row=new_row, column=cols.telegram).value):
                return False, "Этот слот уже заняли.", None

            if max_weekly_hours is not None:
                weekly_booked_hours = calculate_weekly_booked_hours(
                    ws, cols, default_slot_duration_hours
                )
                old_week = week_start(current["slot_date"])
                new_week = week_start(new_slot["slot_date"])
                old_duration = resolve_slot_duration_hours(
                    current, default_slot_duration_hours
                )
                new_duration = resolve_slot_duration_hours(
                    new_slot, default_slot_duration_hours
                )
                projected_hours = weekly_booked_hours.get(new_week, 0.0) + new_duration
                if old_week == new_week:
                    projected_hours -= old_duration

                if projected_hours > max_weekly_hours + 1e-9:
                    return (
                        False,
                        "На этой неделе уже достигнут лимит часов по эксперименту.",
                        None,
                    )

            clear_booking_row(ws, current["row"], cols)
            write_booking_to_row(
                ws,
                new_row,
                current["telegram"],
                current["full_name"],
                current["phone"],
                cols,
            )
            sort_slots_sheet_rows(ws, cols)

            success, message = save_workbook_to_storage(storage, wb)
            if not success:
                return False, message, None

            return True, "ok", new_slot["label"]
        finally:
            wb.close()


def move_user_booking_generated(
    storage: ExcelStorage,
    user_handles: list[str],
    *,
    slot_key: str,
    working_hours: str | None,
    excluded_weekdays: set[int] | None,
    slot_duration_hours: float,
    min_gap_hours: float,
    slot_step_minutes: int,
    max_weekly_hours: float | None = None,
    default_slot_duration_hours: float | None = None,
    days_ahead: int = SLOT_HORIZON_DAYS,
) -> tuple[bool, str, str | None]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(storage)
        if load_error:
            return False, load_error, None
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            current = find_user_booking_in_sheet(ws, user_handles, cols)
            if not current:
                return False, "У вас нет активной записи.", None

            if not parse_generated_slot_key(slot_key):
                return False, "Некорректный слот.", None

            current_key = generated_slot_key(current["slot_date"], current["slot_time"])
            if current_key == slot_key:
                return False, "Вы уже записаны на этот слот.", None

            target_slot = find_generated_slot_by_key(
                ws,
                cols,
                key=slot_key,
                working_hours=working_hours,
                excluded_weekdays=excluded_weekdays,
                slot_duration_hours=slot_duration_hours,
                min_gap_hours=min_gap_hours,
                slot_step_minutes=slot_step_minutes,
                max_weekly_hours=max_weekly_hours,
                default_slot_duration_hours=default_slot_duration_hours,
                days_ahead=days_ahead,
                ignore_booking_row=current["row"],
            )
            if not target_slot:
                return False, "Этот слот недоступен. Выберите другой.", None

            ws.cell(row=current["row"], column=cols.date).value = target_slot["slot_date"].strftime(
                "%d.%m.%Y"
            )
            ws.cell(row=current["row"], column=cols.time).value = format_time_range(
                target_slot["start_dt"].time(),
                target_slot["end_dt"].time(),
            )
            write_booking_to_row(
                ws,
                current["row"],
                current["telegram"],
                current["full_name"],
                current["phone"],
                cols,
            )
            sort_slots_sheet_rows(ws, cols)

            success, message = save_workbook_to_storage(storage, wb)
            if not success:
                return False, message, None

            return True, "ok", target_slot["label"]
        finally:
            wb.close()


def move_user_booking_generated_with_labshake(
    exp: ExperimentConfig,
    user_handles: list[str],
    *,
    slot_key: str,
    working_hours: str | None,
    excluded_weekdays: set[int] | None,
    slot_duration_hours: float,
    min_gap_hours: float,
    slot_step_minutes: int,
    max_weekly_hours: float | None = None,
    default_slot_duration_hours: float | None = None,
    days_ahead: int = SLOT_HORIZON_DAYS,
) -> tuple[bool, str, str | None]:
    with excel_lock:
        wb, load_error = load_workbook_from_storage(exp.storage)
        if load_error:
            return False, load_error, None
        assert wb is not None

        try:
            ws = wb.active
            cols = detect_sheet_columns(ws)
            current = find_user_booking_in_sheet(ws, user_handles, cols)
            if not current:
                return False, "У вас нет активной записи.", None

            if not parse_generated_slot_key(slot_key):
                return False, "Некорректный слот.", None

            current_key = generated_slot_key(current["slot_date"], current["slot_time"])
            if current_key == slot_key:
                return False, "Вы уже записаны на этот слот.", None

            target_slot = find_generated_slot_by_key(
                ws,
                cols,
                key=slot_key,
                working_hours=working_hours,
                excluded_weekdays=excluded_weekdays,
                slot_duration_hours=slot_duration_hours,
                min_gap_hours=min_gap_hours,
                slot_step_minutes=slot_step_minutes,
                max_weekly_hours=max_weekly_hours,
                default_slot_duration_hours=default_slot_duration_hours,
                days_ahead=days_ahead,
                ignore_booking_row=current["row"],
            )
            if not target_slot:
                return False, "Этот слот недоступен. Выберите другой.", None

            current_time_value = ws.cell(row=current["row"], column=cols.time).value
            fallback_duration = default_slot_duration_hours
            if (fallback_duration is None or fallback_duration <= 0) and slot_duration_hours > 0:
                fallback_duration = slot_duration_hours
            current_interval = interval_from_time_cell(
                current["slot_date"],
                current_time_value,
                fallback_duration,
            )
            if not current_interval:
                return (
                    False,
                    "Не удалось определить текущий слот для отмены в LabShake.",
                    None,
                )
            current_start = current_interval[0].time()
            current_end = current_interval[1].time()

            reserve_ok, reserve_message = reserve_buffered_interval_in_labshake(
                exp=exp,
                slot_date=target_slot["slot_date"],
                selected_start=target_slot["start_dt"].time(),
                selected_end=target_slot["end_dt"].time(),
            )
            if not reserve_ok:
                return False, reserve_message, None

            cancel_old_ok, cancel_old_message = cancel_buffered_interval_in_labshake(
                exp=exp,
                slot_date=current["slot_date"],
                selected_start=current_start,
                selected_end=current_end,
            )
            if not cancel_old_ok:
                rollback_ok, rollback_message = cancel_buffered_interval_in_labshake(
                    exp=exp,
                    slot_date=target_slot["slot_date"],
                    selected_start=target_slot["start_dt"].time(),
                    selected_end=target_slot["end_dt"].time(),
                )
                if rollback_ok:
                    return (
                        False,
                        "Не удалось отменить предыдущую запись в LabShake. "
                        f"Причина: {cancel_old_message}. "
                        "Перенос не выполнен, новый слот отменен. Попробуйте снова.",
                        None,
                    )
                return (
                    False,
                    "Не удалось отменить предыдущую запись в LabShake. "
                    f"Также не удалось откатить новый слот: {rollback_message}. "
                    "Проверьте LabShake вручную и свяжитесь с экспериментатором.",
                    None,
                )

            ws.cell(row=current["row"], column=cols.date).value = target_slot[
                "slot_date"
            ].strftime("%d.%m.%Y")
            ws.cell(row=current["row"], column=cols.time).value = format_time_range(
                target_slot["start_dt"].time(),
                target_slot["end_dt"].time(),
            )
            write_booking_to_row(
                ws,
                current["row"],
                current["telegram"],
                current["full_name"],
                current["phone"],
                cols,
            )
            sort_slots_sheet_rows(ws, cols)

            success, message = save_workbook_to_storage(exp.storage, wb)
            if not success:
                return False, message, None

            return True, "ok", target_slot["label"]
        finally:
            wb.close()


def clear_booking_context(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.pop("booking_step", None)
    context.user_data.pop("selected_row", None)
    context.user_data.pop("selected_generated_key", None)
    context.user_data.pop("selected_slot_kind", None)
    context.user_data.pop("selected_label", None)
    context.user_data.pop("full_name", None)
    context.user_data.pop("selected_mode", None)


def clear_admin_context(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.pop("admin_step", None)
    context.user_data.pop("admin_experiment_id", None)
    context.user_data.pop("admin_field", None)
    context.user_data.pop("admin_pending_sync_after_link", None)
    context.user_data.pop("admin_delete_day_raw", None)
    context.user_data.pop("admin_delete_day_page", None)


def get_target_message(update: Update):
    if update.message:
        return update.message
    if update.callback_query and update.callback_query.message:
        return update.callback_query.message
    return None


def get_experiments(context: ContextTypes.DEFAULT_TYPE) -> list[ExperimentConfig]:
    return context.bot_data.get("experiments", [])


def get_participant_experiments(
    context: ContextTypes.DEFAULT_TYPE,
) -> list[ExperimentConfig]:
    return [exp for exp in get_experiments(context) if exp.participant_visible]


def get_experiment_by_id(
    context: ContextTypes.DEFAULT_TYPE, experiment_id: str
) -> ExperimentConfig | None:
    for exp in get_experiments(context):
        if exp.experiment_id == experiment_id:
            return exp
    return None


def get_current_experiment(context: ContextTypes.DEFAULT_TYPE) -> ExperimentConfig | None:
    current_id = context.user_data.get("experiment_id")
    if not current_id:
        return None

    experiments = get_experiments(context)
    for exp in experiments:
        if exp.experiment_id == current_id:
            return exp
    return None


def clear_hidden_selected_experiment(context: ContextTypes.DEFAULT_TYPE) -> bool:
    exp = get_current_experiment(context)
    if not exp or exp.participant_visible:
        return False
    context.user_data.pop("experiment_id", None)
    clear_booking_context(context)
    return True


def current_storage(context: ContextTypes.DEFAULT_TYPE) -> ExcelStorage | None:
    exp = get_current_experiment(context)
    if not exp:
        return None
    return exp.storage


def current_role(context: ContextTypes.DEFAULT_TYPE) -> str | None:
    role = context.user_data.get("role")
    if role in {"participant", "researcher"}:
        return role
    return None


def get_researcher_access(context: ContextTypes.DEFAULT_TYPE) -> ResearcherAccess:
    access = context.bot_data.get("researchers_access")
    if isinstance(access, ResearcherAccess):
        return access
    return ResearcherAccess()


def is_researcher_allowed(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> bool:
    user = update.effective_user
    if not user:
        return False

    access = get_researcher_access(context)
    if user.id in access.user_ids:
        return True

    normalized = normalize_username(user.username)
    if normalized and normalized in access.usernames:
        return True

    return False


def experiment_belongs_to_user(
    exp: ExperimentConfig, user_id: int, username: str | None
) -> bool:
    scientist_raw = (exp.scientist_id or "").strip()
    if not scientist_raw:
        return False

    if scientist_raw.startswith("id:"):
        try:
            return int(scientist_raw[3:]) == user_id
        except ValueError:
            return False

    if scientist_raw.isdigit():
        return int(scientist_raw) == user_id

    scientist_username = normalize_username(scientist_raw)
    current_username = normalize_username(username)
    if scientist_username and current_username:
        return scientist_username == current_username
    return False


def get_researcher_experiments(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> list[ExperimentConfig]:
    user = update.effective_user
    if not user:
        return []

    own = [
        exp
        for exp in get_experiments(context)
        if experiment_belongs_to_user(exp, user.id, user.username)
    ]
    own.sort(key=lambda item: item.title.lower())
    return own


def parse_admin_numeric_value(
    raw_text: str, field_name: str
) -> float | int | None:
    text = raw_text.strip().lower()
    if text in {"none", "null", "пусто", "нет"}:
        if field_name == "available_days_ahead":
            raise ValueError("Введите количество дней целым числом, например 14.")
        return None

    if field_name in {"slot_step_minutes", "available_days_ahead"}:
        try:
            value = int(text)
        except ValueError:
            example = "14" if field_name == "available_days_ahead" else "60"
            raise ValueError(f"Введите целое число, например {example}.") from None
        if value <= 0:
            raise ValueError("Значение должно быть больше 0.")
        if field_name == "available_days_ahead" and value > 365:
            raise ValueError("Слишком большое значение. Допустимо до 365 дней.")
        return value

    try:
        value = float(text.replace(",", "."))
    except ValueError:
        raise ValueError("Введите число, например 3 или 1.5.") from None

    if field_name == "min_gap_hours":
        if value < 0:
            raise ValueError("Значение должно быть >= 0.")
    else:
        if value <= 0:
            raise ValueError("Значение должно быть больше 0.")
    return value


def normalize_labshake_schedule_url(raw_value: str) -> tuple[str | None, str | None]:
    text = (raw_value or "").strip()
    if not text:
        return None, "Ссылка пустая."

    parsed = urlparse(text)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return None, "Нужна полная ссылка, начинающаяся с http:// или https://"

    host = parsed.netloc.lower()
    if "labshake.com" not in host:
        return None, "Это не ссылка LabShake. Нужна ссылка вида https://labshake.com/..."

    path = parsed.path.lower()
    if "reservation" not in path and "schedule" not in path:
        return None, (
            "Нужна ссылка на страницу расписания ресурса (обычно с /reservation в пути)."
        )

    return text, None


def update_experiment_config_value(
    context: ContextTypes.DEFAULT_TYPE,
    *,
    experiment_id: str,
    field_name: str,
    value: object,
) -> tuple[bool, str]:
    experiments_file = context.bot_data.get("experiments_file", "experiments.json")
    default_storage = context.bot_data.get("default_storage")
    if not isinstance(default_storage, ExcelStorage):
        return False, "Не найдена базовая конфигурация хранилища."

    if not os.path.exists(experiments_file):
        return False, f"Файл {experiments_file} не найден."

    try:
        with open(experiments_file, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception as exc:
        return False, f"Не удалось прочитать {experiments_file}: {exc}"

    if isinstance(raw, dict):
        items = raw.get("experiments")
    elif isinstance(raw, list):
        items = raw
    else:
        items = None

    if not isinstance(items, list):
        return False, "Неверный формат experiments.json."

    original_text = json.dumps(raw, ensure_ascii=False, indent=2)

    target: dict | None = None
    for idx, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            continue
        candidate_id = str(item.get("id", "")).strip()
        candidate_slug = make_experiment_id(candidate_id or str(item.get("title", "")), idx)
        if candidate_slug == experiment_id:
            target = item
            break

    if target is None:
        return False, "Эксперимент не найден в experiments.json."

    # Keep stable id even if title changes for records without explicit id.
    if not str(target.get("id", "")).strip():
        target["id"] = experiment_id

    extra_params = target.get("extra_params")
    if extra_params is None or not isinstance(extra_params, dict):
        extra_params = {}
        target["extra_params"] = extra_params

    if field_name in {
        "working_hours",
        "excluded_days",
        "max_weekly_hours",
        "available_days_ahead",
        "slot_duration_hours",
        "min_gap_hours",
        "slot_step_minutes",
        "labshake_booking_comment",
        "slot_mode",
    }:
        if value is None:
            extra_params.pop(field_name, None)
        else:
            extra_params[field_name] = value

        if field_name == "slot_mode" and value == "day_windows":
            current_duration = target.get("slot_duration_hours", extra_params.get("slot_duration_hours"))
            if current_duration in {None, ""}:
                extra_params["slot_duration_hours"] = 3
        if field_name == "slot_step_minutes":
            extra_params.pop("min_gap_hours", None)
        if field_name == "available_days_ahead":
            if value is None:
                extra_params.pop("labshake_days_ahead", None)
            else:
                extra_params["labshake_days_ahead"] = value
    elif field_name == "participant_visible":
        target["participant_visible"] = bool(value)
        extra_params.pop("participant_visible", None)
    else:
        target[field_name] = value

    updated_text = json.dumps(raw, ensure_ascii=False, indent=2)

    try:
        with open(experiments_file, "w", encoding="utf-8") as f:
            f.write(updated_text)
        updated_experiments = load_experiments_config(default_storage)
    except Exception as exc:
        try:
            with open(experiments_file, "w", encoding="utf-8") as f:
                f.write(original_text)
        except Exception:
            logger.exception("Could not rollback %s", experiments_file)
        return False, str(exc)

    context.bot_data["experiments"] = updated_experiments
    return True, "ok"


def current_experiment_limits(context: ContextTypes.DEFAULT_TYPE) -> tuple[float | None, float | None]:
    exp = get_current_experiment(context)
    if not exp:
        return None, None
    return exp.max_weekly_hours, exp.default_slot_duration_hours


def current_slot_generation_config(
    context: ContextTypes.DEFAULT_TYPE,
) -> tuple[str, str | None, float | None, float, int]:
    exp = get_current_experiment(context)
    if not exp:
        return "manual", None, None, 0.0, 60
    return (
        exp.slot_mode,
        exp.working_hours,
        exp.slot_duration_hours,
        exp.min_gap_hours,
        exp.slot_step_minutes,
    )


def current_excluded_weekdays(context: ContextTypes.DEFAULT_TYPE) -> set[int]:
    exp = get_current_experiment(context)
    if not exp:
        return set()
    return set(exp.excluded_weekdays)


def current_available_days_ahead(context: ContextTypes.DEFAULT_TYPE) -> int:
    exp = get_current_experiment(context)
    if not exp:
        return SLOT_HORIZON_DAYS
    try:
        value = int(exp.available_days_ahead)
    except (TypeError, ValueError):
        return SLOT_HORIZON_DAYS
    return max(1, value)


def current_scientist_id(context: ContextTypes.DEFAULT_TYPE) -> str:
    exp = get_current_experiment(context)
    if exp and exp.scientist_id:
        return exp.scientist_id

    experiments = get_experiments(context)
    if experiments and experiments[0].scientist_id:
        return experiments[0].scientist_id

    return "scientist_id"


def should_use_labshake_booking(exp: ExperimentConfig | None) -> bool:
    if not exp:
        return False
    return exp.slot_mode == "day_windows" and bool(exp.labshake_schedule_url)


def should_use_labshake_writeback(exp: ExperimentConfig | None) -> bool:
    # Temporarily keep LabShake in read-only mode for participant flows.
    # Sync (read) remains enabled, but reserve/cancel/move in LabShake are off
    # unless LABSHAKE_WRITEBACK_ENABLED=1 is explicitly set.
    if not should_use_labshake_booking(exp):
        return False
    return parse_env_bool("LABSHAKE_WRITEBACK_ENABLED", default=False)


def support_error_message(context: ContextTypes.DEFAULT_TYPE) -> str:
    scientist_id = current_scientist_id(context)
    return f"Попробуйте позже, либо напишите лично экспериментатору: {scientist_id}"


def mark_terms_accepted(context: ContextTypes.DEFAULT_TYPE) -> None:
    exp = get_current_experiment(context)
    if not exp:
        return
    accepted = set(context.user_data.get("terms_accepted_for", []))
    accepted.add(exp.experiment_id)
    context.user_data["terms_accepted_for"] = list(accepted)


def experiments_keyboard(experiments: list[ExperimentConfig]) -> InlineKeyboardMarkup:
    keyboard = [
        [
            InlineKeyboardButton(
                exp.title, callback_data=f"exp_select:{exp.experiment_id}"
            )
        ]
        for exp in experiments
    ]
    return InlineKeyboardMarkup(keyboard)


def terms_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Я прочитал и подхожу требованиям",
                    callback_data="accept_terms",
                )
            ]
        ]
    )


def edit_booking_data_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Исправить ФИО", callback_data="edit_booking_data"
                )
            ]
        ]
    )


def main_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("📝 Записаться", callback_data="menu_book")],
            [InlineKeyboardButton("🔁 Перенести запись", callback_data="menu_move")],
            [InlineKeyboardButton("🧪 Другие эксперименты", callback_data="menu_experiments")],
        ]
    )


def post_booking_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("🏠 Меню", callback_data="menu_main")],
            [InlineKeyboardButton("🧪 Другие эксперименты", callback_data="menu_experiments")],
        ]
    )


def entry_keyboard(include_experiment_selector: bool = False) -> ReplyKeyboardMarkup:
    rows: list[list[str]] = [[ROLE_PARTICIPANT_BUTTON_TEXT, ROLE_RESEARCHER_BUTTON_TEXT]]
    if include_experiment_selector:
        rows.append([EXPERIMENT_SELECTOR_BUTTON_TEXT])

    return ReplyKeyboardMarkup(
        rows,
        resize_keyboard=True,
        one_time_keyboard=False,
        is_persistent=True,
        input_field_placeholder="Выберите роль",
    )


def researcher_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("🧪 Мои эксперименты", callback_data="admin_edit")],
            [InlineKeyboardButton("👤 Режим участника", callback_data="admin_to_participant")],
        ]
    )


def researcher_experiments_keyboard(experiments: list[ExperimentConfig]) -> InlineKeyboardMarkup:
    keyboard = [
        [InlineKeyboardButton(exp.title, callback_data=f"admin_exp:{exp.experiment_id}")]
        for exp in experiments
    ]
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_menu")])
    return InlineKeyboardMarkup(keyboard)


def researcher_experiment_actions_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("Изменить название", callback_data="admin_action:edit_title")],
            [InlineKeyboardButton("Изменить условия", callback_data="admin_action:edit_terms")],
            [InlineKeyboardButton("Ссылка на таблицу", callback_data="admin_action:table_link")],
            [InlineKeyboardButton("Синхронизировать LabShake", callback_data="admin_action:sync_labshake")],
            [InlineKeyboardButton("Параметры эксперимента", callback_data="admin_action:params")],
            [InlineKeyboardButton("Удалить слот", callback_data="admin_action:delete_slot:0")],
            [InlineKeyboardButton("К списку экспериментов", callback_data="admin_edit")],
            [InlineKeyboardButton("Меню исследователя", callback_data="admin_menu")],
        ]
    )


def labshake_link_choice_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Использовать текущую",
                    callback_data="admin_labshake_link:use_current",
                )
            ],
            [
                InlineKeyboardButton(
                    "Изменить ссылку",
                    callback_data="admin_labshake_link:change",
                )
            ],
            [InlineKeyboardButton("К действиям эксперимента", callback_data="admin_back_actions")],
        ]
    )


def format_admin_field_value(exp: ExperimentConfig, field_name: str) -> str:
    if field_name == "participant_visible":
        return "Опубликован" if exp.participant_visible else "Невидимка"
    if field_name == "working_hours":
        value = exp.working_hours
    elif field_name == "excluded_days":
        return format_weekday_set(exp.excluded_weekdays)
    elif field_name == "max_weekly_hours":
        value = exp.max_weekly_hours
    elif field_name == "slot_mode":
        value = exp.slot_mode
    elif field_name == "available_days_ahead":
        value = exp.available_days_ahead
    elif field_name == "slot_duration_hours":
        value = exp.slot_duration_hours
    elif field_name == "min_gap_hours":
        value = exp.min_gap_hours
    elif field_name == "slot_step_minutes":
        value = exp.slot_step_minutes
    elif field_name == "labshake_booking_comment":
        value = exp.labshake_booking_comment
    elif field_name == "scientist_id":
        value = exp.scientist_id
    else:
        value = None

    if value is None:
        return "не задано"
    if field_name == "labshake_booking_comment":
        text = str(value).strip()
        if len(text) > 80:
            return text[:77] + "..."
        return text
    return str(value)


def researcher_visibility_publish_keyboard(
    is_visible: bool,
) -> InlineKeyboardMarkup:
    publish_label = "Опубликовать"
    hide_label = "Сделать невидимым"
    if is_visible:
        publish_label = "Опубликован (оставить)"
    else:
        hide_label = "Невидимка (оставить)"
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton(publish_label, callback_data="admin_visibility:publish")],
            [InlineKeyboardButton(hide_label, callback_data="admin_visibility:hide")],
            [InlineKeyboardButton("Отмена", callback_data="admin_visibility:cancel")],
        ]
    )


def build_visibility_publish_confirmation_text(exp: ExperimentConfig) -> str:
    working_hours = exp.working_hours or "не задано"
    excluded_days = format_weekday_set(exp.excluded_weekdays)
    max_weekly_hours = (
        str(exp.max_weekly_hours) if exp.max_weekly_hours is not None else "не задано"
    )
    slot_duration_hours = (
        str(exp.slot_duration_hours) if exp.slot_duration_hours is not None else "не задано"
    )
    labshake_schedule_url = exp.labshake_schedule_url or "не задано"
    labshake_booking_comment = exp.labshake_booking_comment or "не задано"
    return (
        f"Ваш id: {exp.scientist_id}.\n"
        f"Название вашего эксперимента: {exp.title}.\n"
        f"Бронирование осуществляется через режим {exp.slot_mode}.\n"
        f"Запись доступна в интервале {working_hours}.\n"
        f"Длительность одной записи {slot_duration_hours}.\n"
        f"Между участниками перерыв в {exp.slot_step_minutes} (это окно учитывается при бронировании).\n"
        f"Следующие дни недели недоступны для бронирования: {excluded_days}.\n"
        f"Максимальное количество часов в неделю, доступное для бронирования: {max_weekly_hours}.\n"
        f"Запись доступна на {exp.available_days_ahead} дней заранее.\n"
        "Вы синхронизируете LabShake с правильной ссылкой "
        "(это нужная вам лаборатория и вы скопировали ссылку по инструкции): "
        f"{labshake_schedule_url}.\n"
        "Если вы не уверены, проверьте ссылку еще раз или обратитесь в поддержку (@LeylaGkk).\n"
        f"Ваш комментарий к записям: {labshake_booking_comment}.\n"
        "Если есть что-то, что нужно изменить, нажмите кнопку \"Отмена\" "
        "и отредактируйте эксперимент.\n"
        "Если все верно, нажмите \"Опубликовать\", и эксперимент станет доступным для участников.\n"
        "Не забудьте изменить параметр видимости, если ваш эксперимент больше не нуждается в участниках."
    )


def researcher_params_keyboard() -> InlineKeyboardMarkup:
    keyboard: list[list[InlineKeyboardButton]] = []
    for field_name in ADMIN_FIELD_ORDER:
        keyboard.append(
            [
                InlineKeyboardButton(
                    ADMIN_FIELD_LABELS[field_name],
                    callback_data=f"admin_field:{field_name}",
                )
            ]
        )
    keyboard.append([InlineKeyboardButton("Назад", callback_data="admin_back_actions")])
    keyboard.append([InlineKeyboardButton("Меню исследователя", callback_data="admin_menu")])
    return InlineKeyboardMarkup(keyboard)


def researcher_delete_slots_keyboard(
    days: list[dict], page: int, experiment_id: str, page_size: int = 14
) -> InlineKeyboardMarkup:
    keyboard: list[list[InlineKeyboardButton]] = []
    total_pages = max(1, (len(days) + page_size - 1) // page_size)
    page = max(0, min(page, total_pages - 1))
    start = page * page_size
    end = min(len(days), start + page_size)

    for day_item in days[start:end]:
        count = day_item.get("total_count", 0)
        slot_date = parse_admin_day_raw(str(day_item.get("day_raw", "")))
        label = day_button_label(str(day_item["label"]), slot_date)
        keyboard.append(
            [
                InlineKeyboardButton(
                    f"{label} ({count})",
                    callback_data=(
                        f"admin_delete_day:{experiment_id}:{day_item['day_raw']}:{page}"
                    ),
                )
            ]
        )

    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(
            InlineKeyboardButton(
                "Назад", callback_data=f"admin_action:delete_slot:{page - 1}"
            )
        )
    if page + 1 < total_pages:
        nav.append(
            InlineKeyboardButton(
                "Дальше", callback_data=f"admin_action:delete_slot:{page + 1}"
            )
        )
    if nav:
        keyboard.append(nav)

    keyboard.append([InlineKeyboardButton("К действиям эксперимента", callback_data="admin_back_actions")])
    keyboard.append([InlineKeyboardButton("Меню исследователя", callback_data="admin_menu")])
    return InlineKeyboardMarkup(keyboard)


def researcher_delete_scope_keyboard(day_raw: str, page: int, experiment_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Удалить целый день",
                    callback_data=(
                        f"admin_delete_scope:day:{experiment_id}:{day_raw}:{page}"
                    ),
                )
            ],
            [
                InlineKeyboardButton(
                    "Удалить интервал",
                    callback_data=(
                        f"admin_delete_scope:interval:{experiment_id}:{day_raw}:{page}"
                    ),
                )
            ],
            [
                InlineKeyboardButton(
                    "К выбору дня",
                    callback_data=f"admin_action:delete_slot:{page}",
                )
            ],
            [InlineKeyboardButton("К действиям эксперимента", callback_data="admin_back_actions")],
            [InlineKeyboardButton("Меню исследователя", callback_data="admin_menu")],
        ]
    )


def storage_reference_text(storage: ExcelStorage) -> str:
    if storage.mode == "yadisk" and storage.yadisk_path:
        raw_path = storage.yadisk_path
        if raw_path.startswith("disk:/"):
            encoded = raw_path[len("disk:/") :]
            encoded = f"disk/{encoded}"
            from urllib.parse import quote

            return f"https://disk.360.yandex.com/edit/disk/{quote(encoded, safe='')}"
        return raw_path
    if storage.mode == "local":
        return f"Локальный файл: {storage.excel_path}"
    return "Ссылка недоступна."


def researcher_fields_keyboard() -> InlineKeyboardMarkup:
    return researcher_params_keyboard()


def researcher_slot_mode_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("manual", callback_data="admin_set_mode:manual")],
            [InlineKeyboardButton("day_windows", callback_data="admin_set_mode:day_windows")],
            [InlineKeyboardButton("Назад", callback_data="admin_back_fields")],
        ]
    )


def terms_accepted(context: ContextTypes.DEFAULT_TYPE) -> bool:
    exp = get_current_experiment(context)
    if not exp:
        return False
    accepted = context.user_data.get("terms_accepted_for", [])
    return exp.experiment_id in accepted


async def send_role_prompt(
    update: Update, context: ContextTypes.DEFAULT_TYPE, prefix: str | None = None
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return

    text = "Кто вы?"
    if prefix:
        text = f"{prefix}\n\n{text}"

    await target_message.reply_text(text, reply_markup=entry_keyboard())


async def clear_callback_inline_keyboard(update: Update) -> None:
    query = update.callback_query
    if not query:
        return
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        logger.debug("Could not clear previous inline keyboard.", exc_info=True)


async def send_experiments_menu(
    update: Update, context: ContextTypes.DEFAULT_TYPE, prefix: str | None = None
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return
    await clear_callback_inline_keyboard(update)

    if current_role(context) == "researcher":
        experiments = get_experiments(context)
    else:
        experiments = get_participant_experiments(context)
    if not experiments:
        empty_text = "Сейчас нет опубликованных экспериментов. Попробуйте позже."
        if prefix:
            empty_text = f"{prefix}\n\n{empty_text}"
        await target_message.reply_text(empty_text)
        return

    text = "Выберите эксперимент:"
    if prefix:
        text = f"{prefix}\n\n{text}"

    await target_message.reply_text(text, reply_markup=experiments_keyboard(experiments))


async def send_researcher_menu(
    update: Update, context: ContextTypes.DEFAULT_TYPE, prefix: str | None = None
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return
    await clear_callback_inline_keyboard(update)

    text = "Меню исследователя:"
    if prefix:
        text = f"{prefix}\n\n{text}"
    await target_message.reply_text(text, reply_markup=researcher_menu_keyboard())


async def send_researcher_experiment_picker(
    update: Update, context: ContextTypes.DEFAULT_TYPE, prefix: str | None = None
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return
    await clear_callback_inline_keyboard(update)

    experiments = get_researcher_experiments(update, context)
    if not experiments:
        await target_message.reply_text(
            "Для вашего аккаунта нет привязанных экспериментов по scientist_id.",
            reply_markup=researcher_menu_keyboard(),
        )
        return

    text = "Выберите эксперимент для настройки:"
    if prefix:
        text = f"{prefix}\n\n{text}"
    await target_message.reply_text(text, reply_markup=researcher_experiments_keyboard(experiments))


async def send_researcher_experiment_actions(
    update: Update, context: ContextTypes.DEFAULT_TYPE, prefix: str | None = None
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return
    await clear_callback_inline_keyboard(update)

    experiment_id = context.user_data.get("admin_experiment_id")
    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return

    exp = get_experiment_by_id(context, experiment_id)
    if not exp:
        await send_researcher_experiment_picker(
            update,
            context,
            prefix="Эксперимент не найден. Выберите снова.",
        )
        return

    user = update.effective_user
    if not user or not experiment_belongs_to_user(exp, user.id, user.username):
        await send_researcher_experiment_picker(
            update,
            context,
            prefix="У вас нет доступа к этому эксперименту.",
        )
        return

    text = f"Эксперимент: {exp.title}\n\nВыберите действие:"
    if prefix:
        text = f"{prefix}\n\n{text}"
    await target_message.reply_text(text, reply_markup=researcher_experiment_actions_keyboard())


async def send_researcher_fields_menu(
    update: Update, context: ContextTypes.DEFAULT_TYPE, prefix: str | None = None
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return
    await clear_callback_inline_keyboard(update)

    experiment_id = context.user_data.get("admin_experiment_id")
    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return

    exp = get_experiment_by_id(context, experiment_id)
    if not exp:
        await send_researcher_experiment_picker(
            update,
            context,
            prefix="Эксперимент не найден. Выберите снова.",
        )
        return

    user = update.effective_user
    if not user or not experiment_belongs_to_user(exp, user.id, user.username):
        await send_researcher_experiment_picker(
            update,
            context,
            prefix="У вас нет доступа к этому эксперименту.",
        )
        return

    details = [
        f"{ADMIN_FIELD_LABELS[field_name]}: {format_admin_field_value(exp, field_name)}"
        for field_name in ADMIN_FIELD_ORDER
    ]
    text = f"Эксперимент: {exp.title}\n\n" + "\n".join(details) + "\n\nВыберите параметр:"
    if prefix:
        text = f"{prefix}\n\n{text}"
    await target_message.reply_text(text, reply_markup=researcher_params_keyboard())


async def send_researcher_delete_slots(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    *,
    page: int = 0,
    prefix: str | None = None,
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return
    await clear_callback_inline_keyboard(update)

    experiment_id = context.user_data.get("admin_experiment_id")
    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return

    exp = get_experiment_by_id(context, experiment_id)
    if not exp:
        await send_researcher_experiment_picker(
            update,
            context,
            prefix="Эксперимент не найден. Выберите снова.",
        )
        return

    user = update.effective_user
    if not user or not experiment_belongs_to_user(exp, user.id, user.username):
        await send_researcher_experiment_picker(
            update,
            context,
            prefix="У вас нет доступа к этому эксперименту.",
        )
        return

    days, error = list_days_for_admin_delete(exp, only_future=True)
    if error:
        await target_message.reply_text(support_error_message(context))
        return

    if not days:
        await target_message.reply_text(
            "Нет будущих слотов для удаления.",
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("К действиям эксперимента", callback_data="admin_back_actions")]]
            ),
        )
        return

    text = "Сначала выберите день для удаления слотов:"
    if prefix:
        text = f"{prefix}\n\n{text}"
    await target_message.reply_text(
        text,
        reply_markup=researcher_delete_slots_keyboard(
            days, page=page, experiment_id=exp.experiment_id
        ),
    )


async def ensure_researcher_access(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> bool:
    if is_researcher_allowed(update, context):
        return True

    context.user_data["role"] = "participant"
    clear_admin_context(context)

    target_message = get_target_message(update)
    if target_message:
        await target_message.reply_text("У вас нет доступа к этому режиму.")
        await send_experiments_menu(update, context)
    return False


async def ensure_participant_access(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> bool:
    role = current_role(context)
    if role == "participant":
        clear_hidden_selected_experiment(context)
        return True

    if role == "researcher":
        # Participant actions should always be one tap away even after researcher mode.
        context.user_data["role"] = "participant"
        clear_admin_context(context)
        clear_hidden_selected_experiment(context)
        return True

    await send_role_prompt(update, context, prefix="Сначала выберите роль.")
    return False


async def send_terms_prompt(
    update: Update, context: ContextTypes.DEFAULT_TYPE, prefix: str | None = None
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return
    await clear_callback_inline_keyboard(update)

    experiment = get_current_experiment(context)
    if not experiment or not experiment.participant_visible:
        if experiment and not experiment.participant_visible:
            clear_hidden_selected_experiment(context)
        await send_experiments_menu(
            update,
            context,
            prefix="Сначала выберите эксперимент.",
        )
        return

    terms_text = experiment.terms_text or DEFAULT_TERMS_TEXT
    text = terms_text if not prefix else f"{prefix}\n\n{terms_text}"
    await target_message.reply_text(text, reply_markup=terms_keyboard())


async def send_main_menu(
    update: Update, context: ContextTypes.DEFAULT_TYPE, prefix: str | None = None
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return
    await clear_callback_inline_keyboard(update)

    experiment = get_current_experiment(context)
    if not experiment or not experiment.participant_visible:
        if experiment and not experiment.participant_visible:
            clear_hidden_selected_experiment(context)
        await send_experiments_menu(
            update,
            context,
            prefix="Сначала выберите эксперимент.",
        )
        return

    text = f"Эксперимент: {experiment.title}\n\nВыберите действие:"
    if prefix:
        text = f"{prefix}\n\n{text}"

    await target_message.reply_text(text, reply_markup=main_menu_keyboard())


def get_window_bounds(mode: str, offset: int, window_days: int = SLOT_WINDOW_DAYS) -> tuple[date, date]:
    offset = max(0, offset)
    window_days = max(1, int(window_days))
    today = date.today()

    if mode == "move":
        base = today - timedelta(days=today.weekday())
    else:
        base = today

    start = base + timedelta(days=offset * window_days)
    end = start + timedelta(days=window_days)
    return start, end


def first_non_empty_offset_for_book(slots: list[dict]) -> int:
    if not slots:
        return 0

    today = date.today()
    future_dates = [slot["slot_date"] for slot in slots if slot["slot_date"] >= today]
    if not future_dates:
        return 0

    first_date = min(future_dates)
    return max(0, (first_date - today).days // SLOT_WINDOW_DAYS)


def build_slot_window_keyboard(
    mode: str,
    window_slots: list[dict],
    offset: int,
    has_prev: bool,
    has_next: bool,
) -> InlineKeyboardMarkup:
    keyboard: list[list[InlineKeyboardButton]] = []

    for slot in window_slots:
        if slot.get("kind") == "generated":
            callback_data = f"slotg:{mode}:{slot['key']}"
        else:
            manual_key_value = slot.get("manual_key")
            if isinstance(manual_key_value, str) and re.fullmatch(r"\d{12}|\d{16}", manual_key_value):
                callback_data = f"slotm:{mode}:{manual_key_value}"
            else:
                callback_data = f"slot:{mode}:{slot['row']}"
        keyboard.append(
            [
                InlineKeyboardButton(
                    slot["label"], callback_data=callback_data
                )
            ]
        )

    nav_buttons: list[InlineKeyboardButton] = []
    if has_prev:
        nav_buttons.append(
            InlineKeyboardButton("Назад", callback_data=f"page:{mode}:{offset - 1}")
        )
    if has_next:
        nav_buttons.append(
            InlineKeyboardButton("Дальше", callback_data=f"page:{mode}:{offset + 1}")
        )
    if nav_buttons:
        keyboard.append(nav_buttons)

    if mode == "move":
        keyboard.append(
            [InlineKeyboardButton("Отменить запись", callback_data="cancel_my_booking")]
        )

    keyboard.append([InlineKeyboardButton("Меню", callback_data="menu_main")])
    return InlineKeyboardMarkup(keyboard)


def build_day_selection_keyboard(
    mode: str,
    day_slots: list[dict],
    offset: int,
    has_prev: bool,
    has_next: bool,
) -> InlineKeyboardMarkup:
    keyboard: list[list[InlineKeyboardButton]] = []
    by_day: dict[date, list[dict]] = {}
    for slot in day_slots:
        by_day.setdefault(slot["slot_date"], []).append(slot)

    for day in sorted(by_day.keys()):
        day_label = by_day[day][0].get("date_label") or date_label_with_weekday(day)
        day_label = day_button_label(day_label, day)
        keyboard.append(
            [
                InlineKeyboardButton(
                    f"{day_label} ({len(by_day[day])})",
                    callback_data=f"day:{mode}:{day.strftime('%Y%m%d')}:{offset}",
                )
            ]
        )

    nav_buttons: list[InlineKeyboardButton] = []
    if has_prev:
        nav_buttons.append(
            InlineKeyboardButton("Назад", callback_data=f"page:{mode}:{offset - 1}")
        )
    if has_next:
        nav_buttons.append(
            InlineKeyboardButton("Дальше", callback_data=f"page:{mode}:{offset + 1}")
        )
    if nav_buttons:
        keyboard.append(nav_buttons)

    if mode == "move":
        keyboard.append(
            [InlineKeyboardButton("Отменить запись", callback_data="cancel_my_booking")]
        )

    keyboard.append([InlineKeyboardButton("Меню", callback_data="menu_main")])
    return InlineKeyboardMarkup(keyboard)


def build_day_slots_keyboard(
    mode: str,
    day_slots: list[dict],
    offset: int,
) -> InlineKeyboardMarkup:
    keyboard: list[list[InlineKeyboardButton]] = []
    for slot in day_slots:
        callback_data = (
            f"slotg:{mode}:{slot['key']}"
            if slot.get("kind") == "generated"
            else f"slot:{mode}:{slot['row']}"
        )
        label = slot.get("time_label")
        if not label:
            label = slot["label"]
        keyboard.append([InlineKeyboardButton(label, callback_data=callback_data)])

    keyboard.append(
        [InlineKeyboardButton("К выбору дня", callback_data=f"page:{mode}:{offset}")]
    )
    if mode == "move":
        keyboard.append(
            [InlineKeyboardButton("Отменить запись", callback_data="cancel_my_booking")]
        )
    keyboard.append([InlineKeyboardButton("Меню", callback_data="menu_main")])
    return InlineKeyboardMarkup(keyboard)


async def show_slot_window(
    update: Update, context: ContextTypes.DEFAULT_TYPE, mode: str, offset: int
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return

    storage = current_storage(context)
    if not storage:
        await send_experiments_menu(
            update,
            context,
            prefix="Сначала выберите эксперимент.",
        )
        return

    max_weekly_hours, default_slot_duration_hours = current_experiment_limits(context)
    (
        slot_mode,
        working_hours,
        slot_duration_hours,
        min_gap_hours,
        slot_step_minutes,
    ) = current_slot_generation_config(context)
    excluded_weekdays = current_excluded_weekdays(context)
    available_days_ahead = current_available_days_ahead(context)
    slots, error = await asyncio.to_thread(
        get_available_slots_with_retry,
        storage,
        max_weekly_hours=max_weekly_hours,
        default_slot_duration_hours=default_slot_duration_hours,
        slot_mode=slot_mode,
        working_hours=working_hours,
        excluded_weekdays=excluded_weekdays,
        slot_duration_hours=slot_duration_hours,
        min_gap_hours=min_gap_hours,
        slot_step_minutes=slot_step_minutes,
        days_ahead=available_days_ahead,
    )
    if error:
        exp = get_current_experiment(context)
        logger.warning(
            "show_slot_window get_available_slots failed: experiment=%s mode=%s offset=%s slot_mode=%s error=%s",
            exp.experiment_id if exp else "none",
            mode,
            offset,
            slot_mode,
            error,
        )
        await target_message.reply_text(support_error_message(context))
        return

    offset = max(0, offset)
    window_days = max(1, min(SLOT_WINDOW_DAYS, available_days_ahead))
    start, end = get_window_bounds(mode, offset, window_days)
    window_slots = [
        slot for slot in slots if start <= slot["slot_date"] < end
    ]
    has_prev = offset > 0
    has_next = any(slot["slot_date"] >= end for slot in slots)

    if mode == "book":
        context.user_data["book_offset"] = offset
    else:
        context.user_data["move_offset"] = offset

    end_inclusive = end - timedelta(days=1)
    period_text = f"{start.strftime('%d.%m.%Y')} - {end_inclusive.strftime('%d.%m.%Y')}"

    if not slots:
        await target_message.reply_text(
            "Сейчас все даты заняты. Попробуйте записаться позже.",
            reply_markup=build_slot_window_keyboard(mode, [], offset, has_prev, has_next),
        )
        return

    if not window_slots:
        await target_message.reply_text(
            f"В период {period_text} свободных слотов нет. "
            "Нажмите «Дальше», чтобы посмотреть следующий период.",
            reply_markup=build_slot_window_keyboard(
                mode, window_slots, offset, has_prev, has_next
            ),
        )
        return

    if slot_mode == "day_windows":
        title = "Дни для записи" if mode == "book" else "Дни для переноса"
        await target_message.reply_text(
            f"{title} на период {period_text}:\nСначала выберите день.",
            reply_markup=build_day_selection_keyboard(
                mode, window_slots, offset, has_prev, has_next
            ),
        )
        return

    title = "Свободные слоты" if mode == "book" else "Слоты для переноса"
    await target_message.reply_text(
        f"{title} на период {period_text}:\nВыберите удобные дату и время.",
        reply_markup=build_slot_window_keyboard(
            mode, window_slots, offset, has_prev, has_next
        ),
    )


async def start_book_flow(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await ensure_participant_access(update, context):
        return
    clear_admin_context(context)
    storage = current_storage(context)
    if not storage:
        await send_experiments_menu(
            update,
            context,
            prefix="Сначала выберите эксперимент.",
        )
        return

    if not terms_accepted(context):
        await send_terms_prompt(
            update,
            context,
            prefix="Перед записью нужно подтвердить, что вы прочитали условия.",
        )
        return

    clear_booking_context(context)
    _, handles = get_user_handles(update)
    booking, error = await asyncio.to_thread(
        find_user_booking_with_retry, storage, handles
    )
    if error:
        target_message = get_target_message(update)
        if target_message:
            await target_message.reply_text(support_error_message(context))
        return

    if booking:
        await send_main_menu(
            update,
            context,
            prefix=(
                f"У вас уже есть запись на {booking['label']}.\n"
                "Для изменения нажмите «Перенести запись»."
            ),
        )
        return

    exp = get_current_experiment(context)
    if should_use_labshake_booking(exp):
        target_message = get_target_message(update)
        if target_message:
            await target_message.reply_text(
                "Пожалуйста, подождите минутку. Проверяю актуальность расписания, "
                "ничего не нажимайте."
            )

        assert exp is not None
        sync_ok, sync_message, _, _, _ = await asyncio.to_thread(
            sync_day_windows_from_labshake_with_retry, exp
        )
        if not sync_ok:
            logger.warning(
                "Participant auto-sync with LabShake failed for experiment %s: %s",
                exp.experiment_id,
                sync_message,
            )
            if target_message:
                await target_message.reply_text(support_error_message(context))
            return

    max_weekly_hours, default_slot_duration_hours = current_experiment_limits(context)
    (
        slot_mode,
        working_hours,
        slot_duration_hours,
        min_gap_hours,
        slot_step_minutes,
    ) = current_slot_generation_config(context)
    excluded_weekdays = current_excluded_weekdays(context)
    available_days_ahead = current_available_days_ahead(context)
    slots, error = await asyncio.to_thread(
        get_available_slots_with_retry,
        storage,
        max_weekly_hours=max_weekly_hours,
        default_slot_duration_hours=default_slot_duration_hours,
        slot_mode=slot_mode,
        working_hours=working_hours,
        excluded_weekdays=excluded_weekdays,
        slot_duration_hours=slot_duration_hours,
        min_gap_hours=min_gap_hours,
        slot_step_minutes=slot_step_minutes,
        days_ahead=available_days_ahead,
    )
    if error:
        logger.warning("start_book_flow get_available_slots failed: %s", error)
        target_message = get_target_message(update)
        if target_message:
            await target_message.reply_text(support_error_message(context))
        return

    initial_offset = first_non_empty_offset_for_book(slots)
    await show_slot_window(update, context, mode="book", offset=initial_offset)


async def start_move_flow(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await ensure_participant_access(update, context):
        return
    clear_admin_context(context)
    storage = current_storage(context)
    if not storage:
        await send_experiments_menu(
            update,
            context,
            prefix="Сначала выберите эксперимент.",
        )
        return

    if not terms_accepted(context):
        await send_terms_prompt(
            update,
            context,
            prefix="Перед переносом нужно подтвердить, что вы прочитали условия.",
        )
        return

    clear_booking_context(context)
    _, handles = get_user_handles(update)
    booking, error = await asyncio.to_thread(
        find_user_booking_with_retry, storage, handles
    )
    if error:
        target_message = get_target_message(update)
        if target_message:
            await target_message.reply_text(support_error_message(context))
        return

    if not booking:
        await send_main_menu(update, context, prefix="У вас пока нет активной записи.")
        return

    exp = get_current_experiment(context)
    if should_use_labshake_booking(exp):
        target_message = get_target_message(update)
        if target_message:
            await target_message.reply_text(
                "Пожалуйста, подождите минутку. Проверяю актуальность расписания, "
                "ничего не нажимайте."
            )

        assert exp is not None
        sync_ok, sync_message, _, _, _ = await asyncio.to_thread(
            sync_day_windows_from_labshake_with_retry, exp
        )
        if not sync_ok:
            logger.warning(
                "Participant move auto-sync with LabShake failed for experiment %s: %s",
                exp.experiment_id,
                sync_message,
            )
            if target_message:
                await target_message.reply_text(support_error_message(context))
            return

    target_message = get_target_message(update)
    if target_message:
        await target_message.reply_text(
            f"Вы записаны на {booking['label']}.\n"
            "Выберите новый слот или нажмите «Отменить запись»."
        )

    await show_slot_window(update, context, mode="move", offset=0)


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    clear_booking_context(context)
    clear_admin_context(context)
    context.user_data.pop("role", None)
    await send_role_prompt(update, context)


async def experiments_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    clear_booking_context(context)
    clear_admin_context(context)
    role = current_role(context)
    if not role:
        await send_role_prompt(update, context, prefix="Сначала выберите роль.")
        return
    if role == "researcher":
        if not await ensure_researcher_access(update, context):
            return
        await send_researcher_experiment_picker(update, context)
        return
    await send_experiments_menu(update, context)


async def book_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    role = current_role(context)
    if not role:
        await send_role_prompt(update, context, prefix="Сначала выберите роль.")
        return
    await start_book_flow(update, context)


async def move_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    role = current_role(context)
    if not role:
        await send_role_prompt(update, context, prefix="Сначала выберите роль.")
        return
    await start_move_flow(update, context)


async def menu_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    role = current_role(context)
    if not role:
        await send_role_prompt(update, context, prefix="Сначала выберите роль.")
        return
    if role == "researcher":
        clear_booking_context(context)
        clear_admin_context(context)
        if not await ensure_researcher_access(update, context):
            return
        await send_researcher_experiment_picker(update, context)
        return

    if not get_current_experiment(context):
        await send_experiments_menu(update, context)
        return
    clear_booking_context(context)
    await send_main_menu(update, context)


async def researcher_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    clear_booking_context(context)
    clear_admin_context(context)
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return
    await send_researcher_experiment_picker(update, context)


async def admin_menu_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    clear_booking_context(context)
    clear_admin_context(context)
    if not await ensure_researcher_access(update, context):
        return
    await send_researcher_menu(update, context)


async def admin_edit_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return
    clear_admin_context(context)
    await send_researcher_experiment_picker(update, context)


async def admin_to_participant_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "participant"
    clear_admin_context(context)
    await send_experiments_menu(update, context, prefix="Режим участника включен.")


async def admin_exp_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return

    data = query.data or ""
    match = re.match(r"^admin_exp:([a-z0-9_]+)$", data)
    if not match:
        await query.message.reply_text("Не удалось определить эксперимент.")
        return

    experiment_id = match.group(1)
    allowed_ids = {
        exp.experiment_id for exp in get_researcher_experiments(update, context)
    }
    if experiment_id not in allowed_ids:
        await query.message.reply_text(
            "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
        )
        await send_researcher_experiment_picker(update, context)
        return

    if not get_experiment_by_id(context, experiment_id):
        await send_researcher_experiment_picker(
            update,
            context,
            prefix="Эксперимент не найден. Выберите снова.",
        )
        return

    context.user_data["admin_experiment_id"] = experiment_id
    context.user_data.pop("admin_step", None)
    context.user_data.pop("admin_field", None)
    context.user_data.pop("admin_pending_sync_after_link", None)
    context.user_data.pop("admin_delete_day_raw", None)
    context.user_data.pop("admin_delete_day_page", None)
    await send_researcher_experiment_actions(update, context)


async def admin_back_actions_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data.pop("admin_step", None)
    context.user_data.pop("admin_field", None)
    context.user_data.pop("admin_pending_sync_after_link", None)
    context.user_data.pop("admin_delete_day_raw", None)
    context.user_data.pop("admin_delete_day_page", None)
    if not await ensure_researcher_access(update, context):
        return
    await send_researcher_experiment_actions(update, context)


async def admin_back_fields_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data.pop("admin_step", None)
    context.user_data.pop("admin_field", None)
    context.user_data.pop("admin_pending_sync_after_link", None)
    context.user_data.pop("admin_delete_day_raw", None)
    context.user_data.pop("admin_delete_day_page", None)
    if not await ensure_researcher_access(update, context):
        return
    await send_researcher_fields_menu(update, context)


async def run_labshake_sync_and_report(
    update: Update, context: ContextTypes.DEFAULT_TYPE, exp: ExperimentConfig
) -> None:
    target_message = get_target_message(update)
    if not target_message:
        return

    if not exp.labshake_schedule_url:
        context.user_data["admin_step"] = "await_labshake_url_for_sync"
        context.user_data["admin_pending_sync_after_link"] = True
        await target_message.reply_text(
            "Ссылка LabShake пока не задана.\n"
            "Отправьте ссылку на любой день расписания лаборатории "
            "(например .../reservation?y=2026&m=3&d=3)."
        )
        return

    status_every_raw = os.getenv("LABSHAKE_SYNC_STATUS_EVERY_SEC", "30").strip()
    try:
        status_every_sec = max(5.0, float(status_every_raw.replace(",", ".")))
    except ValueError:
        status_every_sec = 30.0

    progress_message = await target_message.reply_text(
        "Синхронизация LabShake запущена. Это может занять до 1-3 минут."
    )

    sync_started_at = time_module.monotonic()
    sync_task = asyncio.create_task(asyncio.to_thread(sync_day_windows_from_labshake, exp))
    while True:
        try:
            (
                success,
                sync_message,
                canceled,
                removed_slot_labels,
                removed_mismatch_labels,
            ) = await asyncio.wait_for(asyncio.shield(sync_task), timeout=status_every_sec)
            break
        except asyncio.TimeoutError:
            elapsed_seconds = int(max(0.0, time_module.monotonic() - sync_started_at))
            elapsed_min, elapsed_sec = divmod(elapsed_seconds, 60)
            try:
                await progress_message.edit_text(
                    "Синхронизация LabShake все еще выполняется "
                    f"({elapsed_min:02d}:{elapsed_sec:02d})."
                )
            except Exception:
                logger.debug("Could not update sync progress message.", exc_info=True)
                # Fallback when message editing is unavailable in this chat.
                await target_message.reply_text(
                    "Синхронизация LabShake все еще выполняется "
                    f"({elapsed_min:02d}:{elapsed_sec:02d})."
                )

    if not success:
        try:
            await progress_message.edit_text(
                "Синхронизация LabShake завершилась с ошибкой. Формирую сообщение..."
            )
        except Exception:
            logger.debug("Could not update final sync progress message.", exc_info=True)
        await target_message.reply_text(
            f"Не удалось синхронизировать LabShake: {sync_message}\n{support_error_message(context)}"
        )
        await send_researcher_experiment_actions(update, context)
        return

    try:
        await progress_message.edit_text(
            "Синхронизация LabShake завершена. Формирую итоговый отчет..."
        )
    except Exception:
        logger.debug("Could not update final sync progress message.", exc_info=True)

    notified = 0
    notify_failed = 0
    unresolved = 0
    for item in canceled:
        tg_text = str(item.get("telegram") or "").strip()
        chat_id = item.get("chat_id")
        slot_label = item.get("label") or "ваш слот"

        if chat_id is None and tg_text.startswith("@"):
            try:
                chat = await context.bot.get_chat(tg_text)
                chat_id = int(chat.id)
            except Exception:
                logger.debug("Could not resolve chat id for %s", tg_text, exc_info=True)

        if chat_id is None:
            unresolved += 1
            continue

        notify_text = (
            f"Ваш слот на эксперимент «{exp.title}» ({slot_label}) больше недоступен, "
            "потому что расписание LabShake обновилось. Пожалуйста, запишитесь заново."
        )
        try:
            await context.bot.send_message(chat_id=chat_id, text=notify_text)
            notified += 1
        except Exception:
            notify_failed += 1
            logger.exception("Could not notify participant %s", chat_id)

    notify_summary = ""
    if canceled:
        notify_summary = (
            "\n\nУведомления участникам:\n"
            f"Отправлено: {notified}\n"
            f"Не удалось отправить: {notify_failed}\n"
            f"Нет chat_id для автоуведомления: {unresolved}"
        )
    removed_summary = ""
    if removed_slot_labels:
        shown = removed_slot_labels[:SYNC_REMOVED_SLOTS_LIMIT]
        removed_summary = (
            "\n\nИз вашего эксперимента были удалены следующие слоты:\n"
            + "\n".join(f"- {item}" for item in shown)
        )
        if len(removed_slot_labels) > SYNC_REMOVED_SLOTS_LIMIT:
            removed_summary += (
                f"\n... и еще {len(removed_slot_labels) - SYNC_REMOVED_SLOTS_LIMIT}"
            )

    mismatch_summary = ""
    if removed_mismatch_labels:
        shown_mismatch = removed_mismatch_labels[:SYNC_REMOVED_SLOTS_LIMIT]
        mismatch_summary = (
            "\n\nУдалены по несовмещению времени с занятостью LabShake:\n"
            + "\n".join(f"- {item}" for item in shown_mismatch)
        )
        if len(removed_mismatch_labels) > SYNC_REMOVED_SLOTS_LIMIT:
            mismatch_summary += (
                f"\n... и еще {len(removed_mismatch_labels) - SYNC_REMOVED_SLOTS_LIMIT}"
            )

    await target_message.reply_text(
        sync_message + removed_summary + mismatch_summary + notify_summary
    )
    await send_researcher_experiment_actions(update, context)


async def admin_action_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return

    experiment_id = context.user_data.get("admin_experiment_id")
    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return

    exp = get_experiment_by_id(context, experiment_id)
    if not exp:
        await send_researcher_experiment_picker(update, context)
        return
    if experiment_id not in {
        item.experiment_id for item in get_researcher_experiments(update, context)
    }:
        await query.message.reply_text(
            "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
        )
        await send_researcher_experiment_picker(update, context)
        return

    data = query.data or ""
    match = re.match(
        r"^admin_action:(edit_title|edit_terms|table_link|sync_labshake|params|delete_slot(?::\d+)?)$",
        data,
    )
    if not match:
        await query.message.reply_text("Не удалось определить действие.")
        return

    action = match.group(1)
    if action == "edit_title":
        context.user_data["admin_field"] = "title"
        context.user_data["admin_step"] = "await_admin_value"
        await query.message.reply_text("Введите новое название эксперимента:")
        return

    if action == "edit_terms":
        context.user_data["admin_field"] = "default_terms_text"
        context.user_data["admin_step"] = "await_admin_value"
        await query.message.reply_text(
            "Введите новый текст условий эксперимента (можно несколькими строками)."
        )
        return

    if action == "table_link":
        link = storage_reference_text(exp.storage)
        await query.message.reply_text(f"Таблица эксперимента:\n{link}")
        await send_researcher_experiment_actions(update, context)
        return

    if action == "sync_labshake":
        context.user_data["admin_pending_sync_after_link"] = True
        if not exp.labshake_schedule_url:
            context.user_data["admin_step"] = "await_labshake_url_for_sync"
            await query.message.reply_text(
                "Ссылка LabShake пока не задана.\n"
                "Отправьте ссылку на любой день расписания лаборатории "
                "(например .../reservation?y=2026&m=3&d=3)."
            )
            return

        context.user_data.pop("admin_step", None)
        await query.message.reply_text(
            (
                "Текущая ссылка LabShake:\n"
                f"{exp.labshake_schedule_url}\n\n"
                "Изменить ссылку перед синхронизацией?"
            ),
            reply_markup=labshake_link_choice_keyboard(),
        )
        return

    if action == "params":
        await send_researcher_fields_menu(update, context)
        return

    if action.startswith("delete_slot"):
        page = 0
        parts = action.split(":")
        if len(parts) == 2 and parts[1].isdigit():
            page = int(parts[1])
        context.user_data.pop("admin_step", None)
        context.user_data.pop("admin_delete_day_raw", None)
        context.user_data.pop("admin_delete_day_page", None)
        await send_researcher_delete_slots(update, context, page=page)
        return


async def admin_labshake_link_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return

    experiment_id = context.user_data.get("admin_experiment_id")
    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return

    exp = get_experiment_by_id(context, experiment_id)
    if not exp:
        await send_researcher_experiment_picker(update, context)
        return
    if experiment_id not in {
        item.experiment_id for item in get_researcher_experiments(update, context)
    }:
        await query.message.reply_text(
            "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
        )
        await send_researcher_experiment_picker(update, context)
        return

    data = query.data or ""
    match = re.match(r"^admin_labshake_link:(use_current|change)$", data)
    if not match:
        await query.message.reply_text("Не удалось определить действие для ссылки LabShake.")
        return

    action = match.group(1)
    if action == "change":
        context.user_data["admin_step"] = "await_labshake_url_for_sync"
        context.user_data["admin_pending_sync_after_link"] = True
        await query.message.reply_text(
            "Отправьте новую ссылку на любой день расписания LabShake."
        )
        return

    if not exp.labshake_schedule_url:
        context.user_data["admin_step"] = "await_labshake_url_for_sync"
        context.user_data["admin_pending_sync_after_link"] = True
        await query.message.reply_text(
            "Текущая ссылка не задана. Отправьте ссылку на расписание LabShake."
        )
        return

    context.user_data.pop("admin_pending_sync_after_link", None)
    context.user_data.pop("admin_step", None)
    await run_labshake_sync_and_report(update, context, exp)


async def admin_delete_row_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return

    experiment_id = context.user_data.get("admin_experiment_id")
    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return

    exp = get_experiment_by_id(context, experiment_id)
    if not exp:
        await send_researcher_experiment_picker(update, context)
        return

    data = query.data or ""
    match = re.match(r"^admin_delete_row:(\d+):(\d+)$", data)
    if not match:
        await query.message.reply_text("Не удалось определить слот.")
        return

    row = int(match.group(1))
    page = int(match.group(2))
    success, message, slot_label, tg_text, chat_id = delete_slot_row_for_admin(
        exp.storage, row=row
    )
    if not success:
        await query.message.reply_text(support_error_message(context))
        await send_researcher_delete_slots(update, context, page=page)
        return

    notify_status = ""
    if tg_text:
        notify_text = (
            f"Ваш слот на эксперимент «{exp.title}» ({slot_label}) был отменен "
            "экспериментатором. Пожалуйста, запишитесь заново."
        )
        if chat_id is None and tg_text.startswith("@"):
            try:
                chat = await context.bot.get_chat(tg_text)
                chat_id = int(chat.id)
            except Exception:
                logger.debug("Could not resolve chat id for %s", tg_text, exc_info=True)

        if chat_id is not None:
            try:
                await context.bot.send_message(chat_id=chat_id, text=notify_text)
                notify_status = "\nУчастник уведомлен в Telegram."
            except Exception:
                logger.exception("Could not notify participant %s", chat_id)
                notify_status = (
                    "\nНе удалось уведомить участника автоматически."
                )
        else:
            notify_status = (
                "\nВ таблице нет Telegram id участника (только username), "
                "поэтому автоуведомление недоступно."
            )

    await send_researcher_delete_slots(
        update,
        context,
        page=page,
        prefix=f"Слот удален: {slot_label}.{notify_status}",
    )


async def notify_canceled_participants(
    context: ContextTypes.DEFAULT_TYPE,
    exp: ExperimentConfig,
    canceled: list[dict],
) -> str:
    if not canceled:
        return ""

    notified = 0
    failed = 0
    unresolved = 0
    for item in canceled:
        tg_text = str(item.get("telegram") or "").strip()
        chat_id = item.get("chat_id")
        slot_label = item.get("label") or "ваш слот"

        if chat_id is None and tg_text.startswith("@"):
            try:
                chat = await context.bot.get_chat(tg_text)
                chat_id = int(chat.id)
            except Exception:
                logger.debug("Could not resolve chat id for %s", tg_text, exc_info=True)

        if chat_id is None:
            unresolved += 1
            continue

        notify_text = (
            f"Ваш слот на эксперимент «{exp.title}» ({slot_label}) был отменен "
            "экспериментатором. Пожалуйста, запишитесь заново."
        )
        try:
            await context.bot.send_message(chat_id=chat_id, text=notify_text)
            notified += 1
        except Exception:
            failed += 1
            logger.exception("Could not notify participant %s", chat_id)

    return (
        "\nУведомления участникам:\n"
        f"Отправлено: {notified}\n"
        f"Не удалось отправить: {failed}\n"
        f"Нет chat_id для автоуведомления: {unresolved}"
    )


async def admin_delete_day_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return

    data = query.data or ""
    match_new = re.match(r"^admin_delete_day:([a-z0-9_]+):(\d{8}):(\d+)$", data)
    match_old = re.match(r"^admin_delete_day:(\d{8}):(\d+)$", data)
    if not match_new and not match_old:
        await query.message.reply_text("Не удалось определить день.")
        return

    if match_new:
        experiment_id = match_new.group(1)
        day_raw = match_new.group(2)
        page = int(match_new.group(3))
        context.user_data["admin_experiment_id"] = experiment_id
    else:
        experiment_id = context.user_data.get("admin_experiment_id")
        day_raw = match_old.group(1)
        page = int(match_old.group(2))

    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return
    if experiment_id not in {
        exp.experiment_id for exp in get_researcher_experiments(update, context)
    }:
        await query.message.reply_text(
            "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
        )
        await send_researcher_experiment_picker(update, context)
        return

    day_value = parse_admin_day_raw(day_raw)
    if not day_value:
        await query.message.reply_text("Некорректная дата дня.")
        return

    context.user_data["admin_delete_day_raw"] = day_raw
    context.user_data["admin_delete_day_page"] = page
    context.user_data.pop("admin_step", None)

    await query.message.reply_text(
        (
            f"Выбран день: {date_label_with_weekday(day_value)}.\n"
            "Удалить целый день или только определенный интервал?"
        ),
        reply_markup=researcher_delete_scope_keyboard(day_raw, page, experiment_id),
    )


async def admin_delete_scope_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return

    data = query.data or ""
    match_new = re.match(
        r"^admin_delete_scope:(day|interval):([a-z0-9_]+):(\d{8}):(\d+)$", data
    )
    match_old = re.match(r"^admin_delete_scope:(day|interval):(\d{8}):(\d+)$", data)
    if not match_new and not match_old:
        await query.message.reply_text("Не удалось определить действие удаления.")
        return

    if match_new:
        mode = match_new.group(1)
        experiment_id = match_new.group(2)
        day_raw = match_new.group(3)
        page = int(match_new.group(4))
        context.user_data["admin_experiment_id"] = experiment_id
    else:
        mode = match_old.group(1)
        experiment_id = context.user_data.get("admin_experiment_id")
        day_raw = match_old.group(2)
        page = int(match_old.group(3))

    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return

    exp = get_experiment_by_id(context, experiment_id)
    if not exp:
        await send_researcher_experiment_picker(update, context)
        return
    if experiment_id not in {
        item.experiment_id for item in get_researcher_experiments(update, context)
    }:
        await query.message.reply_text(
            "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
        )
        await send_researcher_experiment_picker(update, context)
        return

    day_value = parse_admin_day_raw(day_raw)
    if not day_value:
        await query.message.reply_text("Некорректная дата дня.")
        return

    if mode == "interval":
        context.user_data["admin_step"] = "await_admin_delete_interval"
        context.user_data["admin_delete_day_raw"] = day_raw
        context.user_data["admin_delete_day_page"] = page
        await query.message.reply_text(
            (
                f"Введите интервал для {date_label_with_weekday(day_value)} "
                "(например 13:00-16:00)."
            )
        )
        return

    success, message, canceled, removed_labels = delete_slots_for_admin_interval(
        exp,
        slot_date=day_value,
        start_time=time(0, 0),
        end_time=time(23, 59),
    )
    if not success:
        await query.message.reply_text(
            message if message == "На выбранный интервал слоты не найдены." else support_error_message(context)
        )
        await send_researcher_delete_slots(update, context, page=page)
        return

    notify_status = await notify_canceled_participants(context, exp, canceled)
    summary = (
        f"Готово. День {date_label_with_weekday(day_value)} закрыт для записи."
    )
    if removed_labels:
        summary += f"\nУдалено занятых слотов: {len(removed_labels)}."
    if notify_status:
        summary += f"\n{notify_status}"

    context.user_data.pop("admin_step", None)
    await send_researcher_delete_slots(update, context, page=page, prefix=summary)


async def admin_field_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return

    data = query.data or ""
    match = re.match(r"^admin_field:([a-z_]+)$", data)
    if not match:
        await query.message.reply_text("Не удалось определить параметр.")
        return

    field_name = match.group(1)
    if field_name not in ADMIN_FIELD_LABELS:
        await query.message.reply_text("Этот параметр пока не поддерживается.")
        return

    if not context.user_data.get("admin_experiment_id"):
        await send_researcher_experiment_picker(update, context)
        return
    if context.user_data.get("admin_experiment_id") not in {
        exp.experiment_id for exp in get_researcher_experiments(update, context)
    }:
        await query.message.reply_text(
            "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
        )
        await send_researcher_experiment_picker(update, context)
        return

    context.user_data["admin_field"] = field_name

    if field_name == "participant_visible":
        context.user_data.pop("admin_step", None)
        exp = get_experiment_by_id(context, context.user_data.get("admin_experiment_id", ""))
        if not exp:
            await send_researcher_experiment_picker(
                update,
                context,
                prefix="Эксперимент не найден. Выберите снова.",
            )
            return
        await query.message.reply_text(
            build_visibility_publish_confirmation_text(exp),
            reply_markup=researcher_visibility_publish_keyboard(exp.participant_visible),
        )
        return

    if field_name == "slot_mode":
        context.user_data.pop("admin_step", None)
        await query.message.reply_text(
            "Не меняйте данный параметр без необходимости!" \
            "manual - берет даты и время, что Вы вручную прописали в Excel файле. Не синхронизируется в LabShake - бронирование вручную." \
            "day_windows - стандартный режим. Синхронизирует слоты с LabShake и предоставляет все доступные. Бронирование автоматическое."              \
            "Выберите режим слотов:",
            reply_markup=researcher_slot_mode_keyboard(),
        )
        return

    context.user_data["admin_step"] = "await_admin_value"

    if field_name == "working_hours":
        prompt = (
            "Рабочее окно позволяет задать интервал в котором доступна запись\n"
            "Введите рабочее окно для записи (например 10:00-17:00).\n"
            "Напишите none, чтобы взять весь доступный интервал из LabJournal (только с разрешения лаборантов)."
        )
    elif field_name == "excluded_days":
        prompt = (
            "Введите дни недели, которые будут недоступны для записи, через запятую.\n"
            "Пример: Суббота, Воскресенье.\n"
            "Можно также: Пн, Вт, Ср... Напишите none, чтобы убрать исключения."
        )
    elif field_name == "max_weekly_hours":
        prompt = (
            "Введите лимит часов записей в неделю (например 16).\n"
            "Напишите none, чтобы убрать лимит."
        )
    elif field_name == "available_days_ahead":
        prompt = (
            "Введите количество дней, на которое показывать доступные слоты "
            "(например 7 или 14).\n"
            "Это же значение будет использоваться для синхронизации LabShake."
        )
    elif field_name == "slot_duration_hours":
        prompt = (
            "Введите длительность записи в часах (например 3).\n"
        )
    elif field_name == "slot_step_minutes":
        prompt = (
            "Введите перерыв между слотами в минутах (например 60).\n"
            "Только целое число > 0."
        )
    elif field_name == "labshake_booking_comment":
        prompt = (
            "Введите комментарий, который будет автоматически подставляться в поле "
            "\"comment\" при бронировании LabShake.\n\n"
            "Рекомендация для исследователя:\n"
            "Enter the experiment code (e.g. 6.13; 2.28), your First and Last Name, "
            "and contact in the \"comment\" field (tg username for example).\n"
            "Example: 1.1 Ivan Ivanov @ivanivanov\n\n"
            "Напишите none, чтобы убрать комментарий."
        )
    elif field_name == "scientist_id":
        prompt = "Введите Telegram исследователя, например @ivanov (лишь в случае, если Вы передаете администрирование другому человеку)"
    else:
        prompt = "Введите новое значение."

    await query.message.reply_text(prompt)


async def admin_set_mode_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return

    experiment_id = context.user_data.get("admin_experiment_id")
    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return
    if experiment_id not in {
        exp.experiment_id for exp in get_researcher_experiments(update, context)
    }:
        await query.message.reply_text(
            "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
        )
        await send_researcher_experiment_picker(update, context)
        return

    data = query.data or ""
    match = re.match(r"^admin_set_mode:(manual|day_windows)$", data)
    if not match:
        await query.message.reply_text("Не удалось определить режим.")
        return

    mode = match.group(1)
    success, message = update_experiment_config_value(
        context,
        experiment_id=experiment_id,
        field_name="slot_mode",
        value=mode,
    )
    if not success:
        await query.message.reply_text(
            f"Не удалось обновить параметр: {message}"
        )
        return

    context.user_data.pop("admin_step", None)
    context.user_data.pop("admin_field", None)
    await send_researcher_fields_menu(
        update,
        context,
        prefix=f"Параметр обновлен: {ADMIN_FIELD_LABELS['slot_mode']} = {mode}",
    )


async def admin_visibility_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    context.user_data["role"] = "researcher"
    if not await ensure_researcher_access(update, context):
        return

    experiment_id = context.user_data.get("admin_experiment_id")
    if not experiment_id:
        await send_researcher_experiment_picker(update, context)
        return
    if experiment_id not in {
        exp.experiment_id for exp in get_researcher_experiments(update, context)
    }:
        await query.message.reply_text(
            "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
        )
        await send_researcher_experiment_picker(update, context)
        return

    exp = get_experiment_by_id(context, experiment_id)
    if not exp:
        await send_researcher_experiment_picker(update, context)
        return

    data = query.data or ""
    match = re.match(r"^admin_visibility:(publish|hide|cancel)$", data)
    if not match:
        await query.message.reply_text("Не удалось определить действие видимости.")
        return

    action = match.group(1)
    if action == "cancel":
        context.user_data.pop("admin_step", None)
        context.user_data.pop("admin_field", None)
        await send_researcher_fields_menu(
            update,
            context,
            prefix="Публикация отменена. Вы можете отредактировать параметры.",
        )
        return

    make_visible = action == "publish"
    success, message = update_experiment_config_value(
        context,
        experiment_id=experiment_id,
        field_name="participant_visible",
        value=make_visible,
    )
    if not success:
        await query.message.reply_text(
            f"Не удалось обновить параметр: {message}"
        )
        return

    context.user_data.pop("admin_step", None)
    context.user_data.pop("admin_field", None)
    if make_visible:
        status_message = "Эксперимент опубликован и теперь виден участникам."
    else:
        status_message = "Эксперимент переведен в режим «Невидимка»."
    await send_researcher_fields_menu(update, context, prefix=status_message)


async def accept_terms_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer("Спасибо!")

    mark_terms_accepted(context)

    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        logger.debug("Could not remove terms keyboard", exc_info=True)

    await send_main_menu(update, context)


async def menu_main_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not await ensure_participant_access(update, context):
        await query.answer()
        return
    if not get_current_experiment(context):
        await query.answer("Сначала выберите эксперимент.", show_alert=True)
        await send_experiments_menu(update, context)
        return
    if not terms_accepted(context):
        await query.answer("Сначала подтвердите условия.", show_alert=True)
        await send_terms_prompt(update, context)
        return
    await query.answer()
    clear_booking_context(context)
    await send_main_menu(update, context)


async def menu_book_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not await ensure_participant_access(update, context):
        await query.answer()
        return
    if not get_current_experiment(context):
        await query.answer("Сначала выберите эксперимент.", show_alert=True)
        await send_experiments_menu(update, context)
        return
    if not terms_accepted(context):
        await query.answer("Сначала подтвердите условия.", show_alert=True)
        await send_terms_prompt(update, context)
        return
    await query.answer()
    await start_book_flow(update, context)


async def menu_move_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not await ensure_participant_access(update, context):
        await query.answer()
        return
    if not get_current_experiment(context):
        await query.answer("Сначала выберите эксперимент.", show_alert=True)
        await send_experiments_menu(update, context)
        return
    if not terms_accepted(context):
        await query.answer("Сначала подтвердите условия.", show_alert=True)
        await send_terms_prompt(update, context)
        return
    await query.answer()
    await start_move_flow(update, context)


async def menu_experiments_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()
    clear_booking_context(context)
    await send_experiments_menu(update, context)


async def select_experiment_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    data = query.data or ""
    match = re.match(r"^exp_select:([a-z0-9_]+)$", data)
    if not match:
        await query.answer(support_error_message(context), show_alert=True)
        return

    exp_id = match.group(1)
    experiments = get_experiments(context)
    target_exp = None
    for exp in experiments:
        if exp.experiment_id == exp_id:
            target_exp = exp
            break

    if not target_exp:
        await query.answer(support_error_message(context), show_alert=True)
        await send_experiments_menu(update, context)
        return
    if current_role(context) != "researcher" and not target_exp.participant_visible:
        await query.answer("Этот эксперимент пока не опубликован.", show_alert=True)
        await send_experiments_menu(update, context)
        return

    await query.answer()
    clear_booking_context(context)
    context.user_data["experiment_id"] = target_exp.experiment_id

    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        logger.debug("Could not remove experiment keyboard", exc_info=True)

    if terms_accepted(context):
        await send_main_menu(update, context)
    else:
        await send_terms_prompt(update, context)


async def open_slots_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not await ensure_participant_access(update, context):
        await query.answer()
        return
    if not get_current_experiment(context):
        await query.answer("Сначала выберите эксперимент.", show_alert=True)
        await send_experiments_menu(update, context)
        return

    if not terms_accepted(context):
        await query.answer(
            "Сначала подтвердите, что прочитали условия.",
            show_alert=True,
        )
        await send_terms_prompt(update, context)
        return

    await query.answer()
    await start_book_flow(update, context)


async def page_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not await ensure_participant_access(update, context):
        await query.answer()
        return
    if not get_current_experiment(context):
        await query.answer("Сначала выберите эксперимент.", show_alert=True)
        await send_experiments_menu(update, context)
        return

    if not terms_accepted(context):
        await query.answer("Сначала подтвердите условия.", show_alert=True)
        await send_terms_prompt(update, context)
        return

    await query.answer()

    data = query.data or ""
    match = re.match(r"^page:(book|move):(\d+)$", data)
    if not match:
        await query.message.reply_text(support_error_message(context))
        return

    mode = match.group(1)
    offset = int(match.group(2))
    await show_slot_window(update, context, mode=mode, offset=offset)


async def day_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not await ensure_participant_access(update, context):
        await query.answer()
        return
    if not get_current_experiment(context):
        await query.answer("Сначала выберите эксперимент.", show_alert=True)
        await send_experiments_menu(update, context)
        return

    if not terms_accepted(context):
        await query.answer("Сначала подтвердите условия.", show_alert=True)
        await send_terms_prompt(update, context)
        return

    await query.answer()

    data = query.data or ""
    match = re.match(r"^day:(book|move):(\d{8}):(\d+)$", data)
    if not match:
        await query.message.reply_text(support_error_message(context))
        return

    mode = match.group(1)
    day_raw = match.group(2)
    offset = int(match.group(3))
    try:
        day_value = datetime.strptime(day_raw, "%Y%m%d").date()
    except ValueError:
        await query.message.reply_text(support_error_message(context))
        return

    storage = current_storage(context)
    if not storage:
        await send_experiments_menu(update, context)
        return

    max_weekly_hours, default_slot_duration_hours = current_experiment_limits(context)
    (
        slot_mode,
        working_hours,
        slot_duration_hours,
        min_gap_hours,
        slot_step_minutes,
    ) = current_slot_generation_config(context)
    excluded_weekdays = current_excluded_weekdays(context)
    available_days_ahead = current_available_days_ahead(context)
    slots, error = await asyncio.to_thread(
        get_available_slots_with_retry,
        storage,
        max_weekly_hours=max_weekly_hours,
        default_slot_duration_hours=default_slot_duration_hours,
        slot_mode=slot_mode,
        working_hours=working_hours,
        excluded_weekdays=excluded_weekdays,
        slot_duration_hours=slot_duration_hours,
        min_gap_hours=min_gap_hours,
        slot_step_minutes=slot_step_minutes,
        days_ahead=available_days_ahead,
    )
    if error:
        exp = get_current_experiment(context)
        logger.warning(
            "day_callback get_available_slots failed: experiment=%s mode=%s offset=%s day=%s slot_mode=%s error=%s",
            exp.experiment_id if exp else "none",
            mode,
            offset,
            day_raw,
            slot_mode,
            error,
        )
        await query.message.reply_text(support_error_message(context))
        return

    window_days = max(1, min(SLOT_WINDOW_DAYS, available_days_ahead))
    start, end = get_window_bounds(mode, offset, window_days)
    day_slots = [
        slot for slot in slots
        if slot["slot_date"] == day_value and start <= slot["slot_date"] < end
    ]
    if not day_slots:
        await show_slot_window(update, context, mode=mode, offset=offset)
        return

    date_label = day_slots[0].get("date_label") or date_label_with_weekday(day_value)
    title = "Слоты на день" if mode == "book" else "Слоты для переноса на день"
    await query.message.reply_text(
        f"{title} {date_label}:\nВыберите удобное время.",
        reply_markup=build_day_slots_keyboard(mode, day_slots, offset),
    )


async def slot_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not await ensure_participant_access(update, context):
        await query.answer()
        return
    storage = current_storage(context)
    if not storage:
        await query.answer("Сначала выберите эксперимент.", show_alert=True)
        await send_experiments_menu(update, context)
        return

    if not terms_accepted(context):
        await query.answer("Сначала подтвердите условия.", show_alert=True)
        await send_terms_prompt(update, context)
        return

    data = query.data or ""
    manual_row_match = re.match(r"^slot:(book|move):(\d+)$", data)
    manual_key_match = re.match(r"^slotm:(book|move):(\d{12}|\d{16})$", data)
    generated_match = re.match(r"^slotg:(book|move):(\d{12})$", data)
    if not manual_row_match and not manual_key_match and not generated_match:
        await query.answer(support_error_message(context), show_alert=True)
        await query.message.reply_text(support_error_message(context))
        return

    mode = (manual_row_match or manual_key_match or generated_match).group(1)
    slot_kind = "manual" if (manual_row_match or manual_key_match) else "generated"
    slot_token = (
        manual_row_match.group(2)
        if manual_row_match
        else manual_key_match.group(2)
        if manual_key_match
        else generated_match.group(2)
    )

    max_weekly_hours, default_slot_duration_hours = current_experiment_limits(context)
    (
        slot_mode,
        working_hours,
        slot_duration_hours,
        min_gap_hours,
        slot_step_minutes,
    ) = current_slot_generation_config(context)
    excluded_weekdays = current_excluded_weekdays(context)
    available_days_ahead = current_available_days_ahead(context)
    slots, error = await asyncio.to_thread(
        get_available_slots_with_retry,
        storage,
        max_weekly_hours=max_weekly_hours,
        default_slot_duration_hours=default_slot_duration_hours,
        slot_mode=slot_mode,
        working_hours=working_hours,
        excluded_weekdays=excluded_weekdays,
        slot_duration_hours=slot_duration_hours,
        min_gap_hours=min_gap_hours,
        slot_step_minutes=slot_step_minutes,
        days_ahead=available_days_ahead,
    )
    if error:
        exp = get_current_experiment(context)
        logger.warning(
            "slot_callback get_available_slots failed: experiment=%s mode=%s slot_kind=%s token=%s slot_mode=%s error=%s",
            exp.experiment_id if exp else "none",
            mode,
            slot_kind,
            slot_token,
            slot_mode,
            error,
        )
        await query.message.reply_text(support_error_message(context))
        return

    if slot_kind == "manual":
        available_slots_by_row = {
            str(slot["row"]): slot
            for slot in slots
            if slot.get("kind", "manual") == "manual" and "row" in slot
        }
        available_slots_by_key = {
            slot["manual_key"]: slot
            for slot in slots
            if (
                slot.get("kind", "manual") == "manual"
                and isinstance(slot.get("manual_key"), str)
                and re.fullmatch(r"\d{12}|\d{16}", slot["manual_key"])
            )
        }
        available_slots = (
            available_slots_by_key if manual_key_match else available_slots_by_row
        )
    else:
        available_slots = {
            slot["key"]: slot
            for slot in slots
            if slot.get("kind") == "generated" and "key" in slot
        }

    if slot_token not in available_slots:
        await query.answer("Этот слот уже занят. Выберите другой.", show_alert=True)
        offset = context.user_data.get("book_offset", 0) if mode == "book" else context.user_data.get("move_offset", 0)
        await show_slot_window(update, context, mode=mode, offset=offset)
        return

    selected_slot = available_slots[slot_token]
    await query.answer()

    if mode == "move":
        _, handles = get_user_handles(update)
        scientist_id = current_scientist_id(context)
        if slot_kind == "generated":
            if slot_duration_hours is None:
                await query.message.reply_text(support_error_message(context))
                return
            exp = get_current_experiment(context)
            if should_use_labshake_writeback(exp) and slot_mode == "day_windows":
                await query.message.reply_text(
                    "Пожалуйста, подождите минутку. Проверяю слот и завершаю перенос, "
                    "ничего не нажимайте."
                )
                assert exp is not None
                success, message, new_label = await asyncio.to_thread(
                    move_user_booking_generated_with_labshake,
                    exp,
                    handles,
                    slot_key=selected_slot["key"],
                    working_hours=working_hours,
                    excluded_weekdays=excluded_weekdays,
                    slot_duration_hours=slot_duration_hours,
                    min_gap_hours=min_gap_hours,
                    slot_step_minutes=slot_step_minutes,
                    max_weekly_hours=max_weekly_hours,
                    default_slot_duration_hours=default_slot_duration_hours,
                    days_ahead=available_days_ahead,
                )
            else:
                success, message, new_label = move_user_booking_generated(
                    storage,
                    handles,
                    slot_key=selected_slot["key"],
                    working_hours=working_hours,
                    excluded_weekdays=excluded_weekdays,
                    slot_duration_hours=slot_duration_hours,
                    min_gap_hours=min_gap_hours,
                    slot_step_minutes=slot_step_minutes,
                    max_weekly_hours=max_weekly_hours,
                    default_slot_duration_hours=default_slot_duration_hours,
                    days_ahead=available_days_ahead,
                )
        else:
            success, message, new_label = move_user_booking(
                storage,
                handles,
                selected_slot["row"],
                max_weekly_hours=max_weekly_hours,
                default_slot_duration_hours=default_slot_duration_hours,
            )
        if not success:
            logger.warning("Move booking failed: %s", message)
            lowered = str(message or "").lower()
            if any(
                marker in lowered
                for marker in (
                    "недоступ",
                    "занял",
                    "занят",
                    "labshake",
                    "выберите другой",
                )
            ):
                await query.message.reply_text(str(message))
            else:
                await query.message.reply_text(support_error_message(context))
            await show_slot_window(
                update, context, mode="move", offset=context.user_data.get("move_offset", 0)
            )
            return

        clear_booking_context(context)
        await query.message.reply_text(
            (
                f"Готово, запись перенесена на {new_label}.\n"
                "Экспериментатор напишет вам за сутки до исследования для подтверждения записи. "
                f"Вы также можете связаться с экспериментатором самостоятельно: {scientist_id}"
            ),
            reply_markup=post_booking_keyboard(),
        )
        return

    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        logger.debug("Could not remove old keyboard", exc_info=True)

    context.user_data["selected_mode"] = "book"
    context.user_data["selected_slot_kind"] = slot_kind
    context.user_data["selected_row"] = selected_slot["row"] if slot_kind == "manual" else None
    context.user_data["selected_generated_key"] = slot_token if slot_kind == "generated" else None
    context.user_data["selected_label"] = selected_slot["label"]
    context.user_data["booking_step"] = "await_full_name"

    await query.message.reply_text(
        f"Вы выбрали: {selected_slot['label']}\nВведите, пожалуйста, ваше полное ФИО:",
    )


async def edit_booking_data_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    await query.answer()

    if not await ensure_participant_access(update, context):
        return

    step = context.user_data.get("booking_step")
    selected_mode = context.user_data.get("selected_mode")
    if selected_mode != "book" or step not in {"await_full_name", "await_phone"}:
        await query.message.reply_text("Сейчас нечего исправлять.")
        return

    context.user_data["booking_step"] = "await_full_name"
    context.user_data.pop("full_name", None)
    await query.message.reply_text(
        "Хорошо, введите ФИО заново:",
    )


async def cancel_my_booking_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    query = update.callback_query
    if not await ensure_participant_access(update, context):
        await query.answer()
        return
    storage = current_storage(context)
    if not storage:
        await query.answer("Сначала выберите эксперимент.", show_alert=True)
        await send_experiments_menu(update, context)
        return

    if not terms_accepted(context):
        await query.answer("Сначала подтвердите условия.", show_alert=True)
        await send_terms_prompt(update, context)
        return

    await query.answer()

    _, handles = get_user_handles(update)
    slot_mode, _, _, _, _ = current_slot_generation_config(context)
    exp = get_current_experiment(context)
    if should_use_labshake_writeback(exp) and slot_mode == "day_windows":
        await query.message.reply_text(
            "Пожалуйста, подождите минутку. Отменяю запись, "
            "ничего не нажимайте."
        )
        assert exp is not None
        success, message, old_label = await asyncio.to_thread(
            cancel_user_booking_with_labshake,
            exp,
            handles,
            slot_mode=slot_mode,
            default_slot_duration_hours=exp.slot_duration_hours
            if exp.slot_duration_hours and exp.slot_duration_hours > 0
            else exp.default_slot_duration_hours,
        )
    else:
        success, message, old_label = cancel_user_booking(
            storage, handles, slot_mode=slot_mode
        )
    if not success:
        logger.warning("Cancel booking failed: %s", message)
        lowered = str(message or "").lower()
        if any(
            marker in lowered
            for marker in (
                "недоступ",
                "занял",
                "занят",
                "labshake",
                "выберите другой",
                "cancel",
            )
        ):
            await query.message.reply_text(str(message))
        else:
            await query.message.reply_text(support_error_message(context))
        await show_slot_window(
            update, context, mode="move", offset=context.user_data.get("move_offset", 0)
        )
        return

    clear_booking_context(context)
    await query.message.reply_text(
        f"Запись на {old_label} отменена.",
        reply_markup=post_booking_keyboard(),
    )


def is_valid_phone(value: str) -> bool:
    if not re.match(r"^[0-9+()\-\s]{5,25}$", value):
        return False
    return bool(re.search(r"\d", value))


async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    step = context.user_data.get("booking_step")
    admin_step = context.user_data.get("admin_step")

    text = (update.message.text or "").strip()

    # Role switching must always work, regardless of current wizard step.
    if text == ROLE_PARTICIPANT_BUTTON_TEXT:
        clear_booking_context(context)
        clear_admin_context(context)
        context.user_data["role"] = "participant"
        await send_experiments_menu(update, context)
        return

    if text == ROLE_RESEARCHER_BUTTON_TEXT:
        clear_booking_context(context)
        clear_admin_context(context)
        context.user_data["role"] = "researcher"
        if not await ensure_researcher_access(update, context):
            return
        await send_researcher_experiment_picker(update, context)
        return

    if text == EXPERIMENT_SELECTOR_BUTTON_TEXT:
        clear_booking_context(context)
        clear_admin_context(context)
        context.user_data["role"] = "participant"
        await send_experiments_menu(update, context)
        return

    if admin_step == "await_labshake_url_for_sync":
        if not await ensure_researcher_access(update, context):
            return

        experiment_id = context.user_data.get("admin_experiment_id")
        if not experiment_id:
            clear_admin_context(context)
            await send_researcher_experiment_picker(
                update,
                context,
                prefix="Сессия синхронизации сброшена. Выберите эксперимент заново.",
            )
            return
        if experiment_id not in {
            exp.experiment_id for exp in get_researcher_experiments(update, context)
        }:
            clear_admin_context(context)
            await update.message.reply_text(
                "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
            )
            await send_researcher_experiment_picker(update, context)
            return

        if not text:
            await update.message.reply_text(
                "Ссылка пустая. Отправьте ссылку на страницу расписания LabShake."
            )
            return

        normalized_url, validation_error = normalize_labshake_schedule_url(text)
        if validation_error:
            await update.message.reply_text(validation_error)
            return
        assert normalized_url is not None

        success, message = update_experiment_config_value(
            context,
            experiment_id=experiment_id,
            field_name="labshake_schedule_url",
            value=normalized_url,
        )
        if not success:
            await update.message.reply_text(f"Не удалось сохранить ссылку: {message}")
            return

        context.user_data.pop("admin_step", None)
        pending_sync = bool(
            context.user_data.pop("admin_pending_sync_after_link", False)
        )

        exp = get_experiment_by_id(context, experiment_id)
        if not exp:
            await send_researcher_experiment_picker(
                update,
                context,
                prefix="Эксперимент не найден после обновления. Выберите снова.",
            )
            return

        if not pending_sync:
            await send_researcher_experiment_actions(
                update,
                context,
                prefix="Ссылка LabShake сохранена.",
            )
            return

        await update.message.reply_text("Ссылка сохранена. Запускаю синхронизацию...")
        await run_labshake_sync_and_report(update, context, exp)
        return

    if admin_step == "await_admin_delete_interval":
        if not await ensure_researcher_access(update, context):
            return

        experiment_id = context.user_data.get("admin_experiment_id")
        if not experiment_id:
            clear_admin_context(context)
            await send_researcher_experiment_picker(
                update,
                context,
                prefix="Сессия удаления сброшена. Выберите эксперимент заново.",
            )
            return
        if experiment_id not in {
            exp.experiment_id for exp in get_researcher_experiments(update, context)
        }:
            clear_admin_context(context)
            await update.message.reply_text(
                "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
            )
            await send_researcher_experiment_picker(update, context)
            return

        exp = get_experiment_by_id(context, experiment_id)
        if not exp:
            clear_admin_context(context)
            await send_researcher_experiment_picker(
                update,
                context,
                prefix="Эксперимент не найден. Выберите снова.",
            )
            return

        day_raw = str(context.user_data.get("admin_delete_day_raw") or "")
        day_value = parse_admin_day_raw(day_raw)
        if not day_value:
            context.user_data.pop("admin_step", None)
            await send_researcher_delete_slots(
                update,
                context,
                prefix="Не удалось определить день. Выберите его снова.",
            )
            return

        normalized = text.replace(" ", "")
        parsed_range = parse_time_range(normalized)
        if not parsed_range:
            await update.message.reply_text(
                "Неверный формат интервала. Пример: 13:00-16:00"
            )
            return

        start_time, end_time = parsed_range
        page = int(context.user_data.get("admin_delete_day_page", 0))
        success, message, canceled, removed_labels = delete_slots_for_admin_interval(
            exp,
            slot_date=day_value,
            start_time=start_time,
            end_time=end_time,
        )
        if not success:
            await update.message.reply_text(
                message if message == "На выбранный интервал слоты не найдены." else support_error_message(context)
            )
            await send_researcher_delete_slots(update, context, page=page)
            return

        notify_status = await notify_canceled_participants(context, exp, canceled)
        summary = (
            f"Готово. Интервал {date_label_with_weekday(day_value)} "
            f"{format_time_range(start_time, end_time)} недоступен для записи."
        )
        if removed_labels:
            summary += f"\nУдалено занятых слотов: {len(removed_labels)}."
        if notify_status:
            summary += f"\n{notify_status}"

        context.user_data.pop("admin_step", None)
        await send_researcher_delete_slots(
            update,
            context,
            page=page,
            prefix=summary,
        )
        return

    if admin_step == "await_admin_value":
        if not await ensure_researcher_access(update, context):
            return

        field_name = context.user_data.get("admin_field")
        experiment_id = context.user_data.get("admin_experiment_id")
        if not field_name or not experiment_id:
            clear_admin_context(context)
            await send_researcher_experiment_picker(
                update,
                context,
                prefix="Сессия настройки сброшена. Выберите эксперимент заново.",
            )
            return
        if experiment_id not in {
            exp.experiment_id for exp in get_researcher_experiments(update, context)
        }:
            clear_admin_context(context)
            await update.message.reply_text(
                "У вас нет доступа к этому эксперименту. Можно редактировать только свои."
            )
            await send_researcher_experiment_picker(update, context)
            return

        if not text:
            await update.message.reply_text("Сообщение пустое. Введите значение параметра.")
            return

        if field_name == "title":
            if len(text) < 3:
                await update.message.reply_text("Название слишком короткое.")
                return
            value = text
        elif field_name == "default_terms_text":
            if len(text) < 10:
                await update.message.reply_text("Текст условий слишком короткий.")
                return
            value = text
        elif field_name == "working_hours":
            lowered = text.strip().lower()
            if lowered in {"none", "null", "пусто", "нет"}:
                value = None
            else:
                normalized = text.replace(" ", "")
                if parse_time_range(normalized) is None:
                    await update.message.reply_text(
                        "Неверный формат. Пример: 10:00-17:00"
                    )
                    return
                value = normalized
        elif field_name == "excluded_days":
            try:
                parsed_days = parse_excluded_weekdays_value(text)
            except ValueError as exc:
                await update.message.reply_text(
                    f"{exc}\nПример: Суббота, Воскресенье"
                )
                return
            value = sorted(parsed_days) if parsed_days else None
        elif field_name == "scientist_id":
            value = text if text.startswith("@") else f"@{text.lstrip('@')}"
            if len(value) < 3:
                await update.message.reply_text("Введите корректный Telegram, например @ivanov.")
                return
        elif field_name == "labshake_booking_comment":
            lowered = text.strip().lower()
            if lowered in {"none", "null", "пусто", "нет"}:
                value = None
            else:
                if len(text) > 400:
                    await update.message.reply_text(
                        "Комментарий слишком длинный. Максимум 400 символов."
                    )
                    return
                value = text
        else:
            try:
                value = parse_admin_numeric_value(text, field_name)
            except ValueError as exc:
                await update.message.reply_text(str(exc))
                return

        success, message = update_experiment_config_value(
            context,
            experiment_id=experiment_id,
            field_name=field_name,
            value=value,
        )
        if not success:
            await update.message.reply_text(f"Не удалось сохранить параметр: {message}")
            return

        clear_booking_context(context)
        context.user_data.pop("admin_step", None)
        context.user_data.pop("admin_field", None)
        if field_name in {"title", "default_terms_text"}:
            label = "Название" if field_name == "title" else "Условия"
            await send_researcher_experiment_actions(
                update,
                context,
                prefix=f"{label} обновлено.",
            )
        else:
            await send_researcher_fields_menu(
                update,
                context,
                prefix=(
                    f"Параметр обновлен: {ADMIN_FIELD_LABELS.get(field_name, field_name)} = "
                    f"{value if value is not None else 'не задано'}"
                ),
            )
        return

    if not step:
        if current_role(context) == "researcher":
            experiments = get_experiments(context)
        else:
            experiments = get_participant_experiments(context)
        matched = next((exp for exp in experiments if exp.title == text), None)
        if matched:
            clear_booking_context(context)
            clear_admin_context(context)
            context.user_data["role"] = "participant"
            context.user_data["experiment_id"] = matched.experiment_id
            if terms_accepted(context):
                await send_main_menu(update, context)
            else:
                await send_terms_prompt(update, context)
            return

        role = current_role(context)
        if role == "researcher":
            if not await ensure_researcher_access(update, context):
                return
            await send_researcher_experiment_picker(update, context)
            return
        if role == "participant":
            await update.message.reply_text(
                "Нажмите кнопку «Выбрать эксперимент» внизу или выберите через меню команд.",
                reply_markup=entry_keyboard(include_experiment_selector=True),
            )
            await send_experiments_menu(update, context)
            return

        await send_role_prompt(update, context, prefix="Сначала выберите роль.")
        return

    if not text:
        await update.message.reply_text("Сообщение пустое. Попробуйте снова.")
        return

    if step == "await_full_name":
        if len(text) < 5:
            await update.message.reply_text("Введите ФИО полностью (минимум 5 символов).")
            return
        if len(text) > 120:
            await update.message.reply_text("Слишком длинное ФИО. Сократите и попробуйте снова.")
            return

        context.user_data["full_name"] = text
        context.user_data["booking_step"] = "await_phone"
        await update.message.reply_text(
            "Теперь отправьте номер телефона (например: +7 999 123-45-67).",
            reply_markup=edit_booking_data_keyboard(),
        )
        return

    if step == "await_phone":
        if not is_valid_phone(text):
            await update.message.reply_text(
                "Номер выглядит некорректно. Пример: +7 999 123-45-67"
            )
            return

        row = context.user_data.get("selected_row")
        generated_key = context.user_data.get("selected_generated_key")
        selected_slot_kind = context.user_data.get("selected_slot_kind")
        full_name = context.user_data.get("full_name")
        slot_label = context.user_data.get("selected_label", "выбранный слот")
        selected_mode = context.user_data.get("selected_mode")

        if selected_mode != "book":
            clear_booking_context(context)
            await update.message.reply_text("Сессия записи сброшена. Нажмите /menu и начните заново.")
            return

        if selected_slot_kind not in {"manual", "generated"} or not full_name:
            clear_booking_context(context)
            await update.message.reply_text("Сессия записи сброшена. Отправьте /book заново.")
            return

        storage = current_storage(context)
        if not storage:
            clear_booking_context(context)
            await send_experiments_menu(
                update,
                context,
                prefix="Сначала выберите эксперимент.",
            )
            return

        exp = get_current_experiment(context)
        max_weekly_hours, default_slot_duration_hours = current_experiment_limits(context)
        (
            slot_mode,
            working_hours,
            slot_duration_hours,
            min_gap_hours,
            slot_step_minutes,
        ) = current_slot_generation_config(context)
        excluded_weekdays = current_excluded_weekdays(context)
        available_days_ahead = current_available_days_ahead(context)
        telegram_handle, user_handles = get_user_handles(update)

        if selected_slot_kind == "manual":
            if not row:
                clear_booking_context(context)
                await update.message.reply_text("Сессия записи сброшена. Отправьте /book заново.")
                return
            success, message = reserve_slot(
                storage,
                row=row,
                telegram_handle=telegram_handle,
                full_name=full_name,
                phone=text,
                user_handles=user_handles,
                max_weekly_hours=max_weekly_hours,
                default_slot_duration_hours=default_slot_duration_hours,
            )
        else:
            if not generated_key or slot_duration_hours is None:
                clear_booking_context(context)
                await update.message.reply_text("Сессия записи сброшена. Отправьте /book заново.")
                return
            if should_use_labshake_writeback(exp) and slot_mode == "day_windows":
                await update.message.reply_text(
                    "Пожалуйста, подождите минутку. Проверяю слот и завершаю запись, "
                    "ничего не нажимайте."
                )
                assert exp is not None
                success, message = await asyncio.to_thread(
                    reserve_generated_slot_with_labshake,
                    exp,
                    slot_key=generated_key,
                    working_hours=working_hours,
                    excluded_weekdays=excluded_weekdays,
                    telegram_handle=telegram_handle,
                    full_name=full_name,
                    phone=text,
                    user_handles=user_handles,
                    slot_duration_hours=slot_duration_hours,
                    min_gap_hours=min_gap_hours,
                    slot_step_minutes=slot_step_minutes,
                    max_weekly_hours=max_weekly_hours,
                    default_slot_duration_hours=default_slot_duration_hours,
                    days_ahead=available_days_ahead,
                )
            else:
                success, message = reserve_generated_slot(
                    storage,
                    slot_key=generated_key,
                    working_hours=working_hours,
                    excluded_weekdays=excluded_weekdays,
                    telegram_handle=telegram_handle,
                    full_name=full_name,
                    phone=text,
                    user_handles=user_handles,
                    slot_duration_hours=slot_duration_hours,
                    min_gap_hours=min_gap_hours,
                    slot_step_minutes=slot_step_minutes,
                    max_weekly_hours=max_weekly_hours,
                    default_slot_duration_hours=default_slot_duration_hours,
                    days_ahead=available_days_ahead,
                )

        if not success:
            logger.warning("Reserve booking failed: %s", message)
            clear_booking_context(context)
            lowered = str(message or "").lower()
            if any(
                marker in lowered
                for marker in (
                    "недоступ",
                    "занял",
                    "занят",
                    "labshake",
                    "выберите другой",
                )
            ):
                await update.message.reply_text(str(message))
                await start_book_flow(update, context)
            else:
                await update.message.reply_text(support_error_message(context))
            return

        clear_booking_context(context)
        scientist_id = current_scientist_id(context)
        await update.message.reply_text(
            (
                f"Готово, вы записаны на {slot_label}.\n"
                "Экспериментатор напишет вам за сутки до исследования для подтверждения записи. "
                f"Вы также можете связаться с экспериментатором самостоятельно: {scientist_id}"
            ),
            reply_markup=post_booking_keyboard(),
        )


async def cancel_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    clear_booking_context(context)
    clear_admin_context(context)
    await update.message.reply_text(
        "Текущее действие отменено. Для записи используйте /book, для меню /menu."
    )


async def configure_bot_ui(app: Application) -> None:
    try:
        await app.bot.set_my_commands(BOT_COMMANDS)
    except Exception:
        logger.exception("Could not set bot commands")

    try:
        await app.bot.set_chat_menu_button(menu_button=MenuButtonCommands())
    except Exception:
        logger.exception("Could not set chat menu button")


def release_single_instance_lock() -> None:
    global single_instance_lock_fd, single_instance_lock_path
    if single_instance_lock_fd is not None:
        try:
            os.close(single_instance_lock_fd)
        except OSError:
            pass
        single_instance_lock_fd = None
    if single_instance_lock_path:
        try:
            if os.path.exists(single_instance_lock_path):
                os.remove(single_instance_lock_path)
        except OSError:
            pass


def process_exists(pid: int) -> bool:
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
    except (OSError, SystemError, ValueError):
        return False
    return True


def acquire_single_instance_lock() -> None:
    global single_instance_lock_fd, single_instance_lock_path
    lock_path_raw = os.getenv("BOT_LOCK_FILE", ".bot.lock").strip() or ".bot.lock"
    lock_path = os.path.abspath(lock_path_raw)

    while True:
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            existing_pid: int | None = None
            try:
                with open(lock_path, "r", encoding="utf-8") as f:
                    content = f.read().strip()
                if content:
                    existing_pid = int(content)
            except Exception:
                existing_pid = None

            if existing_pid and process_exists(existing_pid):
                raise RuntimeError(
                    f"Bot already running (PID {existing_pid}). "
                    "Stop old process before starting a new one."
                )

            try:
                os.remove(lock_path)
            except OSError as exc:
                raise RuntimeError(
                    f"Could not remove stale lock file: {lock_path} ({exc})"
                ) from exc
            continue

        os.write(fd, str(os.getpid()).encode("utf-8"))
        single_instance_lock_fd = fd
        single_instance_lock_path = lock_path
        atexit.register(release_single_instance_lock)
        break


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--check-storage",
        action="store_true",
        help="Проверить доступ к хранилищу Excel и выйти.",
    )
    args = parser.parse_args()

    load_dotenv()
    acquire_single_instance_lock()
    default_storage = build_storage_config()
    experiments = load_experiments_config(default_storage)
    researchers_access = load_researchers_access(experiments)
    experiments_file = os.getenv("EXPERIMENTS_FILE", "experiments.json").strip()

    if args.check_storage:
        failures: list[str] = []
        for exp in experiments:
            ok, message = check_storage_access(exp.storage)
            status_prefix = f"[{exp.title}]"
            if ok:
                print(f"{status_prefix} {message}")
            else:
                print(f"{status_prefix} ERROR: {message}")
                failures.append(f"{exp.title}: {message}")

        if failures:
            raise RuntimeError(" / ".join(failures))
        return

    bot_token = os.getenv("BOT_TOKEN")
    if not bot_token:
        raise RuntimeError("BOT_TOKEN not found. Add it to .env file.")

    for exp in experiments:
        success, error = ensure_storage_workbook_exists(exp.storage)
        if not success:
            raise RuntimeError(f"[{exp.title}] {error}")

    app = Application.builder().token(bot_token).post_init(configure_bot_ui).build()
    app.bot_data["experiments"] = experiments
    app.bot_data["default_storage"] = default_storage
    app.bot_data["experiments_file"] = experiments_file
    app.bot_data["researchers_access"] = researchers_access

    for exp in experiments:
        storage = exp.storage
        if storage.mode == "local":
            logger.info(
                "Experiment '%s' storage: local (%s)", exp.title, storage.excel_path
            )
        else:
            logger.info(
                "Experiment '%s' storage: yadisk (%s)",
                exp.title,
                storage.yadisk_path,
            )

    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("experiments", experiments_command))
    app.add_handler(CommandHandler("researcher", researcher_command))
    app.add_handler(CommandHandler("book", book_command))
    app.add_handler(CommandHandler("move", move_command))
    app.add_handler(CommandHandler("menu", menu_command))
    app.add_handler(CommandHandler("cancel", cancel_command))
    app.add_handler(
        CallbackQueryHandler(select_experiment_callback, pattern=r"^exp_select:[a-z0-9_]+$")
    )
    app.add_handler(CallbackQueryHandler(accept_terms_callback, pattern=r"^accept_terms$"))
    app.add_handler(CallbackQueryHandler(admin_menu_callback, pattern=r"^admin_menu$"))
    app.add_handler(CallbackQueryHandler(admin_edit_callback, pattern=r"^admin_edit$"))
    app.add_handler(
        CallbackQueryHandler(
            admin_to_participant_callback, pattern=r"^admin_to_participant$"
        )
    )
    app.add_handler(CallbackQueryHandler(admin_exp_callback, pattern=r"^admin_exp:[a-z0-9_]+$"))
    app.add_handler(CallbackQueryHandler(admin_back_actions_callback, pattern=r"^admin_back_actions$"))
    app.add_handler(CallbackQueryHandler(admin_back_fields_callback, pattern=r"^admin_back_fields$"))
    app.add_handler(
        CallbackQueryHandler(
            admin_action_callback,
            pattern=r"^admin_action:(edit_title|edit_terms|table_link|sync_labshake|params|delete_slot(?::\d+)?)$",
        )
    )
    app.add_handler(
        CallbackQueryHandler(
            admin_labshake_link_callback,
            pattern=r"^admin_labshake_link:(use_current|change)$",
        )
    )
    app.add_handler(CallbackQueryHandler(admin_field_callback, pattern=r"^admin_field:[a-z_]+$"))
    app.add_handler(
        CallbackQueryHandler(
            admin_visibility_callback, pattern=r"^admin_visibility:(publish|hide|cancel)$"
        )
    )
    app.add_handler(
        CallbackQueryHandler(
            admin_delete_row_callback, pattern=r"^admin_delete_row:\d+:\d+$"
        )
    )
    app.add_handler(
        CallbackQueryHandler(
            admin_delete_day_callback,
            pattern=r"^admin_delete_day:(?:[a-z0-9_]+:)?\d{8}:\d+$",
        )
    )
    app.add_handler(
        CallbackQueryHandler(
            admin_delete_scope_callback,
            pattern=r"^admin_delete_scope:(day|interval):(?:[a-z0-9_]+:)?\d{8}:\d+$",
        )
    )
    app.add_handler(
        CallbackQueryHandler(
            admin_set_mode_callback, pattern=r"^admin_set_mode:(manual|day_windows)$"
        )
    )
    app.add_handler(CallbackQueryHandler(menu_main_callback, pattern=r"^menu_main$"))
    app.add_handler(CallbackQueryHandler(menu_book_callback, pattern=r"^menu_book$"))
    app.add_handler(CallbackQueryHandler(menu_move_callback, pattern=r"^menu_move$"))
    app.add_handler(
        CallbackQueryHandler(menu_experiments_callback, pattern=r"^menu_experiments$")
    )
    app.add_handler(CallbackQueryHandler(open_slots_callback, pattern=r"^open_slots$"))
    app.add_handler(CallbackQueryHandler(page_callback, pattern=r"^page:(book|move):\d+$"))
    app.add_handler(CallbackQueryHandler(day_callback, pattern=r"^day:(book|move):\d{8}:\d+$"))
    app.add_handler(
        CallbackQueryHandler(
            slot_callback,
            pattern=r"^(slot:(book|move):\d+|slotm:(book|move):(\d{12}|\d{16})|slotg:(book|move):\d{12})$",
        )
    )
    app.add_handler(
        CallbackQueryHandler(edit_booking_data_callback, pattern=r"^edit_booking_data$")
    )
    app.add_handler(CallbackQueryHandler(cancel_my_booking_callback, pattern=r"^cancel_my_booking$"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))

    logger.info("Bot started.")
    app.run_polling()


if __name__ == "__main__":
    main()

